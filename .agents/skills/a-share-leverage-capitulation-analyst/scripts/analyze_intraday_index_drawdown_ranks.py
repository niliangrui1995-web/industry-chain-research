from __future__ import annotations

import argparse
import hashlib
import json
import struct
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DAY_STRUCT = struct.Struct("<IIIIIfII")
CNINDEX_COLUMNS = (
    "date",
    "_1",
    "high",
    "open",
    "low",
    "close",
    "_2",
    "return_pct",
    "amount",
    "volume",
    "_3",
)
INDEX_SPECS = {
    "399106": {"label": "深证综指", "signal_file": "signals_sz_triple.csv"},
    "399006": {"label": "创业板指", "signal_file": "signals_chinext_triple.csv"},
    "399102": {"label": "创业板综指", "signal_file": "signals_chinext_comp_triple.csv"},
    "880003": {"label": "平均股价指数", "signal_file": None},
}
FONT_NAME = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def latest_matching(folder: Path, pattern: str) -> Path:
    matches = sorted(folder.glob(pattern), key=lambda item: item.stat().st_mtime)
    if not matches:
        raise FileNotFoundError(f"No source snapshot matches {folder / pattern}")
    return matches[-1]


def validate_ohlc(frame: pd.DataFrame, source: Path) -> pd.DataFrame:
    raw_dates = pd.to_datetime(frame["date"])
    if frame.empty or raw_dates.duplicated().any():
        raise ValueError(f"Empty or duplicate index dates: {source}")
    if not (raw_dates.is_monotonic_increasing or raw_dates.is_monotonic_decreasing):
        raise ValueError(f"Index dates are not consistently ordered: {source}")
    output = frame.sort_values("date").reset_index(drop=True)
    prices = output[["open", "high", "low", "close"]]
    if prices.isna().any().any() or prices.le(0).any().any():
        raise ValueError(f"Missing or non-positive OHLC values: {source}")
    invalid = output["high"].lt(prices[["open", "low", "close"]].max(axis=1)) | output[
        "low"
    ].gt(prices[["open", "high", "close"]].min(axis=1))
    if invalid.any():
        raise ValueError(f"Invalid OHLC relationship in {source}: {int(invalid.sum())} rows")
    return output


def read_cnindex_snapshot(path: Path) -> pd.DataFrame:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("code") != 200:
        raise ValueError(f"CNIndex snapshot failed its response gate: {path}")
    rows = payload.get("data", {}).get("data", [])
    frame = pd.DataFrame(rows, columns=CNINDEX_COLUMNS)
    frame = frame[["date", "open", "high", "low", "close", "return_pct"]].copy()
    frame["date"] = pd.to_datetime(frame["date"], errors="raise")
    for column in ("open", "high", "low", "close"):
        frame[column] = pd.to_numeric(frame[column], errors="raise")
    frame["return_pct"] = pd.to_numeric(
        frame["return_pct"].astype(str).str.replace("%", "", regex=False), errors="raise"
    )
    return validate_ohlc(frame, path)


def read_tdx_day(path: Path) -> pd.DataFrame:
    raw = path.read_bytes()
    if not raw or len(raw) % DAY_STRUCT.size:
        raise ValueError(f"Invalid TDX day file size: {path}")
    rows: list[tuple[int, float, float, float, float]] = []
    for offset in range(0, len(raw), DAY_STRUCT.size):
        date_value, open_value, high_value, low_value, close_value, _, _, _ = DAY_STRUCT.unpack_from(
            raw, offset
        )
        rows.append(
            (
                date_value,
                open_value / 100.0,
                high_value / 100.0,
                low_value / 100.0,
                close_value / 100.0,
            )
        )
    frame = pd.DataFrame(rows, columns=["date", "open", "high", "low", "close"])
    frame["date"] = pd.to_datetime(frame["date"].astype(str), format="%Y%m%d", errors="raise")
    return validate_ohlc(frame, path)


def causal_rank(
    dates: pd.Series,
    values: pd.Series,
    *,
    years: int = 3,
    min_observations: int = 600,
    descending: bool = False,
    eligible: pd.Series | None = None,
) -> pd.DataFrame:
    parsed_dates = pd.to_datetime(dates).reset_index(drop=True)
    numeric = pd.to_numeric(values, errors="coerce").reset_index(drop=True)
    allowed = (
        pd.Series(True, index=numeric.index)
        if eligible is None
        else eligible.fillna(False).astype(bool).reset_index(drop=True)
    )
    ranks = pd.Series(pd.NA, index=numeric.index, dtype="Int64")
    observations = pd.Series(pd.NA, index=numeric.index, dtype="Int64")
    starts = pd.Series(pd.NaT, index=numeric.index, dtype="datetime64[ns]")

    for position, (date, value) in enumerate(zip(parsed_dates, numeric)):
        if pd.isna(value) or not allowed.iloc[position]:
            continue
        start = date - pd.DateOffset(years=years)
        if parsed_dates.iloc[0] > start:
            continue
        mask = (
            parsed_dates.iloc[: position + 1].ge(start)
            & allowed.iloc[: position + 1]
            & numeric.iloc[: position + 1].notna()
        )
        window = numeric.iloc[: position + 1][mask]
        if len(window) < min_observations:
            continue
        ranks.iloc[position] = 1 + int((window > value).sum() if descending else (window < value).sum())
        observations.iloc[position] = len(window)
        starts.iloc[position] = start

    return pd.DataFrame(
        {
            "rank_3y": ranks,
            "rank_window_observations": observations,
            "rank_window_start": starts,
        }
    )


def add_return_and_rank(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.sort_values("date").reset_index(drop=True).copy()
    output["previous_close"] = output["close"].shift(1)
    output["close_return_pct"] = (output["close"] / output["previous_close"] - 1.0) * 100.0
    output["intraday_max_drawdown_pct"] = (
        output["low"] / output["previous_close"] - 1.0
    ) * 100.0
    intraday = causal_rank(output["date"], output["intraday_max_drawdown_pct"])
    close = causal_rank(output["date"], output["close_return_pct"])
    output["intraday_rank_3y"] = intraday["rank_3y"]
    output["intraday_rank_window_observations"] = intraday["rank_window_observations"]
    output["intraday_rank_window_start"] = intraday["rank_window_start"]
    output["close_return_rank_3y"] = close["rank_3y"]
    return output


def compare_official_to_local(official: pd.DataFrame, local: pd.DataFrame) -> dict[str, object]:
    compared = official.merge(local, on="date", how="outer", suffixes=("_official", "_local"), indicator=True)
    both = compared["_merge"].eq("both")
    output: dict[str, object] = {
        "official_rows": len(official),
        "official_start": official["date"].min().strftime("%Y-%m-%d"),
        "official_end": official["date"].max().strftime("%Y-%m-%d"),
        "missing_official_dates_in_local": int(compared["_merge"].eq("left_only").sum()),
        "local_dates_absent_from_official_within_range": int(
            (
                compared["_merge"].eq("right_only")
                & compared["date"].between(official["date"].min(), official["date"].max())
            ).sum()
        ),
    }
    for field in ("open", "high", "low", "close"):
        errors = (compared[f"{field}_official"] - compared[f"{field}_local"]).abs()
        output[f"{field}_mismatches_over_0_011"] = int((both & errors.gt(0.011)).sum())
        output[f"max_{field}_abs_error"] = float(errors.loc[both].max())
    if (
        output["missing_official_dates_in_local"]
        or output["local_dates_absent_from_official_within_range"]
        or any(
        output[f"{field}_mismatches_over_0_011"] for field in ("open", "high", "low", "close")
        )
    ):
        raise ValueError("Local official-index OHLC failed the CNIndex comparison gate")
    return output


def parse_bool(series: pd.Series) -> pd.Series:
    if pd.api.types.is_bool_dtype(series):
        return series.fillna(False)
    normalized = series.astype(str).str.strip().str.lower()
    if not normalized.isin({"true", "false"}).all():
        raise ValueError("Invalid boolean field in factor panel")
    return normalized.eq("true")


def validate_verified_backtest_inputs(backtest_dir: Path) -> dict[str, object]:
    results_path = backtest_dir / "backtest_results.json"
    results = json.loads(results_path.read_text(encoding="utf-8"))
    metadata = results.get("metadata", {})
    margin_csv = Path(str(metadata.get("margin_csv", "")))
    margin_audit_path = Path(str(metadata.get("margin_audit_json", "")))
    if not margin_csv.exists() or not margin_audit_path.exists():
        raise FileNotFoundError("Verified margin inputs referenced by the backtest are unavailable")
    margin_audit = json.loads(margin_audit_path.read_text(encoding="utf-8"))
    margin_hash = sha256_file(margin_csv)
    if margin_audit.get("verified_snapshot_complete") is not True:
        raise ValueError("Verified margin audit did not pass")
    if margin_audit.get("verified_margin_balances_sha256") != margin_hash:
        raise ValueError("Verified margin CSV hash does not match its audit report")
    if metadata.get("margin_csv_sha256") != margin_hash:
        raise ValueError("Backtest margin hash does not match the verified margin CSV")
    if results.get("config", {}).get("start_date") != "2019-01-01":
        raise ValueError("Unexpected backtest evaluation start")
    return {
        "backtest_results": str(results_path.resolve()),
        "backtest_results_sha256": sha256_file(results_path),
        "margin_csv": str(margin_csv.resolve()),
        "margin_csv_sha256": margin_hash,
        "margin_audit_json": str(margin_audit_path.resolve()),
        "margin_audit_json_sha256": sha256_file(margin_audit_path),
        "verified_margin_rows": int(margin_audit["verified_rows"]),
        "verified_margin_start": margin_audit["verified_start"],
        "verified_margin_end": margin_audit["verified_end"],
    }


def load_signal_dates(backtest_dir: Path) -> tuple[pd.DataFrame, dict[str, set[pd.Timestamp]]]:
    memberships: dict[str, set[pd.Timestamp]] = {}
    dates: set[pd.Timestamp] = set()
    for code, spec in INDEX_SPECS.items():
        signal_file = spec["signal_file"]
        if signal_file is None:
            continue
        frame = pd.read_csv(backtest_dir / str(signal_file), parse_dates=["date"])
        values = set(frame["date"])
        memberships[code] = values
        dates.update(values)
    return pd.DataFrame({"date": sorted(dates)}), memberships


def build_signal_day_table(
    backtest_dir: Path,
    series: dict[str, pd.DataFrame],
) -> tuple[pd.DataFrame, dict[str, set[pd.Timestamp]]]:
    base, memberships = load_signal_dates(backtest_dir)
    base["sample_status"] = "formal_close_factor_signal"
    scenario_path = backtest_dir / "estimated_signal_scenarios.json"
    if scenario_path.exists():
        payload = json.loads(scenario_path.read_text(encoding="utf-8"))
        if payload.get("formal_statistics_inclusion") is not False:
            raise ValueError("Estimated scenarios cannot enter formal statistics")
        scenario_dates = [pd.Timestamp(str(row["date"])) for row in payload.get("scenarios", [])]
        if scenario_dates:
            base = pd.concat(
                [
                    base,
                    pd.DataFrame(
                        {
                            "date": scenario_dates,
                            "sample_status": "estimated_margin_scenario_not_formal",
                        }
                    ),
                ],
                ignore_index=True,
            )
    base = base.drop_duplicates("date", keep="first").sort_values("date").reset_index(drop=True)

    for code in ("399106", "399006", "399102"):
        base[f"signal_{code}_close_factor"] = base["date"].isin(memberships[code])
    for code, frame in series.items():
        selected = frame.set_index("date").reindex(base["date"])
        if selected["intraday_rank_3y"].isna().any():
            missing = base.loc[selected["intraday_rank_3y"].isna().to_numpy(), "date"]
            raise ValueError(f"Missing intraday ranks for {code}: {missing.dt.strftime('%Y-%m-%d').tolist()}")
        for column in (
            "previous_close",
            "low",
            "intraday_max_drawdown_pct",
            "intraday_rank_3y",
            "intraday_rank_window_observations",
            "close_return_pct",
            "close_return_rank_3y",
        ):
            base[f"{code}_{column}"] = selected[column].to_numpy()
        base[f"{code}_intraday_top15"] = base[f"{code}_intraday_rank_3y"].le(15)
    return base, memberships


def recompute_margin_rank(panel: pd.DataFrame) -> pd.Series:
    eligible = parse_bool(panel["margin_data_valid"]) & ~parse_bool(panel["long_break_eve"])
    ranks = causal_rank(
        panel["date"],
        panel["margin_outflow_pct"],
        descending=True,
        eligible=eligible,
    )
    return ranks["rank_3y"]


def build_factor_variant_comparison(
    backtest_dir: Path,
    series: dict[str, pd.DataFrame],
    memberships: dict[str, set[pd.Timestamp]],
) -> tuple[pd.DataFrame, dict[str, object]]:
    panel = pd.read_csv(backtest_dir / "factor_panel.csv", parse_dates=["date"])
    panel = panel.sort_values("date").reset_index(drop=True)
    panel["margin_rank_recomputed"] = recompute_margin_rank(panel)
    valid = (
        panel["date"].ge(pd.Timestamp("2019-01-01"))
        & parse_bool(panel["breadth_valid"])
        & parse_bool(panel["margin_data_valid"])
        & ~parse_bool(panel["long_break_eve"])
        & panel["margin_rank_recomputed"].le(15)
        & panel["down_pct"].ge(80.0)
    )
    comparison_dates: set[pd.Timestamp] = set()
    flags: dict[str, dict[str, set[pd.Timestamp]]] = {}
    summary: dict[str, object] = {}

    for code in INDEX_SPECS:
        ranked = series[code].set_index("date")
        aligned = ranked.reindex(panel["date"])
        close_rank = pd.to_numeric(aligned["close_return_rank_3y"], errors="coerce").reset_index(
            drop=True
        )
        intraday_rank = pd.to_numeric(aligned["intraday_rank_3y"], errors="coerce").reset_index(
            drop=True
        )
        close_mask = valid & close_rank.le(15)
        intraday_mask = valid & intraday_rank.le(15)
        close_dates = set(panel.loc[close_mask.fillna(False), "date"])
        intraday_dates = set(panel.loc[intraday_mask.fillna(False), "date"])
        if code in memberships and close_dates != memberships[code]:
            raise ValueError(f"Corrected close-factor reproduction changed the locked {code} signal dates")
        flags[code] = {"close": close_dates, "intraday": intraday_dates}
        comparison_dates.update(close_dates)
        comparison_dates.update(intraday_dates)
        summary[code] = {
            "close_factor_signal_count": len(close_dates),
            "intraday_factor_signal_count": len(intraday_dates),
            "common_count": len(close_dates & intraday_dates),
            "close_only_dates": sorted(date.strftime("%Y-%m-%d") for date in close_dates - intraday_dates),
            "intraday_only_dates": sorted(date.strftime("%Y-%m-%d") for date in intraday_dates - close_dates),
        }

    output = pd.DataFrame({"date": sorted(comparison_dates)})
    for code in INDEX_SPECS:
        output[f"{code}_close_factor_signal"] = output["date"].isin(flags[code]["close"])
        output[f"{code}_intraday_factor_signal"] = output["date"].isin(flags[code]["intraday"])
        output[f"{code}_membership_change"] = output.apply(
            lambda row: (
                "both"
                if row[f"{code}_close_factor_signal"] and row[f"{code}_intraday_factor_signal"]
                else "close_only"
                if row[f"{code}_close_factor_signal"]
                else "intraday_only"
                if row[f"{code}_intraday_factor_signal"]
                else "neither"
            ),
            axis=1,
        )
        ranks = series[code].set_index("date").reindex(output["date"])
        output[f"{code}_close_return_rank_3y"] = ranks["close_return_rank_3y"].to_numpy()
        output[f"{code}_intraday_rank_3y"] = ranks["intraday_rank_3y"].to_numpy()
    return output, summary


def append_table(ws, frame: pd.DataFrame, table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append(
            [
                value.to_pydatetime()
                if isinstance(value, pd.Timestamp)
                else None
                if pd.isna(value)
                else value
                for value in row
            ]
        )
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT_NAME)
    for index, column in enumerate(frame.columns, start=1):
        width = min(28, max(12, len(str(column)) * 1.7))
        ws.column_dimensions[get_column_letter(index)].width = width
        if column == "日期":
            for cell in ws[get_column_letter(index)][1:]:
                cell.number_format = "yyyy-mm-dd"
        elif "跌幅(%)" in str(column):
            for cell in ws[get_column_letter(index)][1:]:
                cell.number_format = "0.0000;[Red]-0.0000;-"
    if len(frame):
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(frame.columns))}{len(frame) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def workbook_signal_table(frame: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame(
        {
            "日期": frame["date"],
            "样本状态": frame["sample_status"].map(
                {
                    "formal_close_factor_signal": "原正式信号日",
                    "estimated_margin_scenario_not_formal": "融资估算场景（非正式）",
                }
            ),
        }
    )
    for code in INDEX_SPECS:
        label = f"{INDEX_SPECS[code]['label']}{code}"
        output[f"{label}盘中最大跌幅(%)"] = frame[f"{code}_intraday_max_drawdown_pct"]
        output[f"{label}三年排名"] = frame[f"{code}_intraday_rank_3y"].astype("Int64")
        output[f"{label}Top15"] = frame[f"{code}_intraday_top15"].map({True: "是", False: "否"})
    return output


def workbook_variant_table(frame: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame({"日期": frame["date"]})
    for code in INDEX_SPECS:
        label = f"{INDEX_SPECS[code]['label']}{code}"
        output[f"{label}收盘口径信号"] = frame[f"{code}_close_factor_signal"].map(
            {True: "是", False: "否"}
        )
        output[f"{label}盘中口径信号"] = frame[f"{code}_intraday_factor_signal"].map(
            {True: "是", False: "否"}
        )
        output[f"{label}变化"] = frame[f"{code}_membership_change"]
    return output


def write_workbook(
    path: Path,
    signal_days: pd.DataFrame,
    variants: pd.DataFrame,
    sources: dict[str, object],
    variant_summary: dict[str, object],
) -> None:
    workbook = Workbook()
    notes = workbook.active
    notes.title = "口径与审计"
    rows = [
        ("项目", "盘中最大跌幅及滚动三年排名"),
        ("盘中最大跌幅", "(T日最低价 / T-1交易日收盘价 - 1) × 100%"),
        ("排名", "1 + [T-3个日历年,T]内严格低于T日数值的交易日数；第1名最惨"),
        ("并列", "相同未四舍五入值采用并列最优名次"),
        ("未来函数", "低价排名只用T日及以前数据；T日融资余额仍到T+1才完整可得"),
        ("正式边界", "原正式信号日保留；盘中口径替换实验单列，不覆盖原信号"),
        ("399106/399006/399102", "国证官网OHLC快照；与本地TDX OHLC逐日核验"),
        ("880003", "通达信官网板块指数完整包，属于行情厂商指数"),
    ]
    for key, value in rows:
        notes.append([key, value])
    notes.append([])
    notes.append(["来源哈希", json.dumps(sources, ensure_ascii=False)])
    notes.append(["新旧信号摘要", json.dumps(variant_summary, ensure_ascii=False)])
    for row in notes.iter_rows():
        for cell in row:
            cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in notes[1]:
        cell.fill = NOTE_FILL
        cell.font = Font(name=FONT_NAME, bold=True)
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 120
    notes.sheet_view.showGridLines = False

    detail = workbook.create_sheet("原信号日盘中排名")
    append_table(detail, workbook_signal_table(signal_days), "IntradaySignalRanks")
    comparison = workbook.create_sheet("新旧口径信号比较")
    append_table(comparison, workbook_variant_table(variants), "IntradaySignalChanges")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    project_root = Path(__file__).resolve().parents[4]
    default_root = project_root / "artifacts" / "leverage_capitulation" / "verified_2016_present"
    parser = argparse.ArgumentParser(description="Audit intraday maximum index drawdowns on signal dates")
    parser.add_argument("--backtest-dir", type=Path, default=default_root / "backtest_2019_present")
    parser.add_argument("--snapshot-dir", type=Path, default=default_root / "source_snapshots")
    parser.add_argument("--ht-root", type=Path, default=Path(r"D:\HT"))
    parser.add_argument("--output-dir", type=Path, default=default_root / "intraday_index_drawdown_3y")
    args = parser.parse_args()

    verified_backtest = validate_verified_backtest_inputs(args.backtest_dir)

    source_paths = {
        code: latest_matching(args.snapshot_dir, f"cnindex_{code}_*.json")
        for code in ("399106", "399006", "399102")
    }
    source_paths["880003"] = latest_matching(args.snapshot_dir, "sh880003_complete_*.day")
    zip_path = latest_matching(args.snapshot_dir, "tdxzs_day_*.zip")

    series: dict[str, pd.DataFrame] = {}
    official_dates: pd.DatetimeIndex | None = None
    source_audit: dict[str, object] = {}
    for code in ("399106", "399006", "399102"):
        official = read_cnindex_snapshot(source_paths[code])
        dates = pd.DatetimeIndex(official["date"])
        if official_dates is None:
            official_dates = dates
        elif not dates.equals(official_dates):
            raise ValueError("The three CNIndex snapshots do not share an identical trading calendar")
        local_path = args.ht_root / "vipdoc" / "sz" / "lday" / f"sz{code}.day"
        local = read_tdx_day(local_path)
        comparison = compare_official_to_local(official, local)
        series[code] = add_return_and_rank(official)
        source_audit[code] = {
            "source": "CNIndex official daily OHLC JSON",
            "path": str(source_paths[code].resolve()),
            "sha256": sha256_file(source_paths[code]),
            "local_tdx_path": str(local_path.resolve()),
            "local_tdx_sha256": sha256_file(local_path),
            "official_local_comparison": comparison,
        }

    average = read_tdx_day(source_paths["880003"])
    assert official_dates is not None
    relevant_official = set(official_dates[official_dates >= average["date"].min()])
    missing_average_dates = sorted(relevant_official - set(average["date"]))
    if missing_average_dates:
        raise ValueError(
            "Complete 880003 snapshot is missing official trading dates: "
            f"{[date.strftime('%Y-%m-%d') for date in missing_average_dates[:10]]}"
        )
    series["880003"] = add_return_and_rank(average)
    current_average_path = args.ht_root / "vipdoc" / "sh" / "lday" / "sh880003.day"
    current_average = read_tdx_day(current_average_path)
    current_missing = sorted(set(average["date"]) - set(current_average["date"]))
    average_common = average.merge(
        current_average, on="date", how="inner", suffixes=("_complete", "_current")
    )
    current_mismatches: dict[str, int] = {}
    for field in ("open", "high", "low", "close"):
        errors = (average_common[f"{field}_complete"] - average_common[f"{field}_current"]).abs()
        current_mismatches[field] = int(errors.gt(0.011).sum())
    if any(current_mismatches.values()):
        raise ValueError("Current 880003 file disagrees with the official full package on common dates")
    source_audit["880003"] = {
        "source": "TDX official full board-index package",
        "path": str(source_paths["880003"].resolve()),
        "sha256": sha256_file(source_paths["880003"]),
        "rows": len(average),
        "current_local_path": str(current_average_path.resolve()),
        "current_local_sha256": sha256_file(current_average_path),
        "current_local_missing_dates": [date.strftime("%Y-%m-%d") for date in current_missing],
        "current_local_common_rows": len(average_common),
        "current_local_common_ohlc_mismatches_over_0_011": current_mismatches,
        "package_zip": str(zip_path.resolve()),
        "package_zip_sha256": sha256_file(zip_path),
        "package_url": "https://www.tdx.com.cn/products/data/data/vipdoc/tdxzs_day.zip",
    }

    signal_days, memberships = build_signal_day_table(args.backtest_dir, series)
    variants, variant_summary = build_factor_variant_comparison(
        args.backtest_dir, series, memberships
    )
    args.output_dir.mkdir(parents=True, exist_ok=True)
    daily = series["399106"][["date"]].copy()
    for code, frame in series.items():
        columns = [
            "date",
            "previous_close",
            "low",
            "intraday_max_drawdown_pct",
            "intraday_rank_3y",
            "intraday_rank_window_observations",
            "close_return_pct",
            "close_return_rank_3y",
        ]
        renamed = frame[columns].rename(
            columns={column: f"{code}_{column}" for column in columns if column != "date"}
        )
        daily = daily.merge(renamed, on="date", how="left", validate="one_to_one")

    daily_path = args.output_dir / "intraday_index_drawdown_rank_daily_3y.csv"
    signal_path = args.output_dir / "signal_day_intraday_index_drawdown_ranks.csv"
    comparison_path = args.output_dir / "intraday_factor_signal_comparison.csv"
    workbook_path = args.output_dir / "intraday_drawdown_rank_comparison_2019_present.xlsx"
    daily.to_csv(daily_path, index=False, encoding="utf-8-sig")
    signal_days.to_csv(signal_path, index=False, encoding="utf-8-sig")
    variants.to_csv(comparison_path, index=False, encoding="utf-8-sig")
    write_workbook(workbook_path, signal_days, variants, source_audit, variant_summary)

    formal = signal_days["sample_status"].eq("formal_close_factor_signal")
    top15_counts = {
        code: int(signal_days.loc[formal, f"{code}_intraday_top15"].sum()) for code in INDEX_SPECS
    }
    audit = {
        "schema_version": 1,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "formula": "(low_T / close_T_minus_1_trading_day - 1) * 100",
        "rank_definition": (
            "1 + count of strictly lower unrounded values within inclusive [T-3 calendar years, T]; "
            "rank 1 is the worst intraday drawdown"
        ),
        "min_window_observations": 600,
        "future_information_boundary": (
            "intraday low rank uses only T and earlier market data; T margin balance remains available only on T+1"
        ),
        "formal_signal_dates_preserved": True,
        "formal_signal_count": int(formal.sum()),
        "estimated_scenario_count": int((~formal).sum()),
        "formal_signal_intraday_top15_counts": top15_counts,
        "sources": source_audit,
        "verified_backtest": verified_backtest,
        "factor_variant_summary": variant_summary,
        "outputs": {
            "daily_csv": str(daily_path.resolve()),
            "daily_csv_sha256": sha256_file(daily_path),
            "signal_csv": str(signal_path.resolve()),
            "signal_csv_sha256": sha256_file(signal_path),
            "comparison_csv": str(comparison_path.resolve()),
            "comparison_csv_sha256": sha256_file(comparison_path),
            "workbook": str(workbook_path.resolve()),
            "workbook_sha256": sha256_file(workbook_path),
        },
    }
    audit_path = args.output_dir / "intraday_drawdown_analysis_audit.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
