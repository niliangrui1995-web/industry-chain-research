from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HORIZONS = (1, 2, 5, 10, 20, 40)
PERFORMANCE_SAMPLE_KEY = "terminal_10d"
PERFORMANCE_SAMPLE_LABEL = "10日连续簇最后信号（事后）"
SIGNALS = {
    "sz_triple": {
        "label": "深证综指",
        "code": "399106",
        "return_prefix": "sz_comp",
        "rank_column": "sz_comp_rank",
        "return_column": "sz_comp_return_pct",
        "sheet_prefix": "深证综三因子",
    },
    "chinext_triple": {
        "label": "创业板指",
        "code": "399006",
        "return_prefix": "chinext",
        "rank_column": "chinext_rank",
        "return_column": "chinext_return_pct",
        "sheet_prefix": "创业板指三因子",
    },
    "chinext_comp_triple": {
        "label": "创业板综指",
        "code": "399102",
        "return_prefix": "chinext_comp",
        "rank_column": "chinext_comp_rank",
        "return_column": "chinext_comp_return_pct",
        "sheet_prefix": "创业板综指三因子",
    },
}
OUTCOMES = {
    "sz_comp": "深证综指",
    "chinext": "创业板指",
    "chinext_comp": "创业板综指",
}
FONT_NAME = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
POSITIVE_FILL = PatternFill("solid", fgColor="E2F0D9")
NEGATIVE_FILL = PatternFill("solid", fgColor="FCE4D6")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFF2CC")
INPUT_BLUE = "0000FF"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def yes_no(values: pd.Series) -> pd.Series:
    return values.fillna(False).astype(bool).map({True: "是", False: "否"})


def load_inputs(backtest_dir: Path, margin_csv: Path, margin_audit: Path) -> dict[str, object]:
    audit = json.loads(margin_audit.read_text(encoding="utf-8"))
    margin_hash = sha256_file(margin_csv)
    if audit.get("verified_snapshot_complete") is not True:
        raise ValueError("融资余额审计未通过，禁止导出")
    if audit.get("verified_margin_balances_sha256") != margin_hash:
        raise ValueError("融资余额文件哈希与审计报告不一致")

    results = json.loads((backtest_dir / "backtest_results.json").read_text(encoding="utf-8"))
    if results["metadata"]["margin_csv_sha256"] != margin_hash:
        raise ValueError("回测使用的融资余额哈希与当前文件不一致")
    if results["metadata"].get("factor_index_source") not in {
        "cnindex",
        "cnindex_audited_snapshot",
    }:
        raise ValueError("指数因子不是国证官网或其已审计快照口径，禁止导出")
    expected_codes = {
        "sz_comp": "399106",
        "chinext": "399006",
        "chinext_comp": "399102",
    }
    if results["metadata"].get("factor_index_codes") != expected_codes:
        raise ValueError("回测指数代码不是399106、399006和399102三指数口径，禁止导出")

    panel = pd.read_csv(backtest_dir / "factor_panel.csv", parse_dates=["date"])
    margin = pd.read_csv(margin_csv, parse_dates=["date"])
    if panel["date"].duplicated().any() or margin["date"].duplicated().any():
        raise ValueError("因子面板或融资余额存在重复交易日")
    if set(panel["date"]) != set(margin["date"]):
        raise ValueError("因子面板与融资余额交易日集合不一致")

    scenario_path = backtest_dir / "estimated_signal_scenarios.json"
    scenarios: list[dict[str, object]] = []
    if scenario_path.exists():
        scenario_payload = json.loads(scenario_path.read_text(encoding="utf-8"))
        if scenario_payload.get("formal_statistics_inclusion") is not False:
            raise ValueError("估算场景不得并入正式统计")
        scenarios = list(scenario_payload.get("scenarios", []))
        formal_end = pd.Timestamp(results["metadata"]["data_end"])
        for scenario in scenarios:
            if scenario.get("sample_status") != "estimated_not_in_formal_statistics":
                raise ValueError("估算场景缺少未纳入正式统计标记")
            if pd.Timestamp(str(scenario["date"])) <= formal_end:
                raise ValueError("估算场景日期必须晚于正式回测截止日")

    rank_threshold = float(results["config"]["rank_threshold"])
    breadth_threshold = float(results["config"]["breadth_threshold"])
    panel_by_date = panel.set_index("date")
    signal_frames: dict[str, pd.DataFrame] = {}

    for key, spec in SIGNALS.items():
        frame = pd.read_csv(backtest_dir / f"signals_{key}.csv", parse_dates=["date"])
        dates = set(frame["date"])
        expected_signals = int(results["summaries"][key]["all_signals"]["signal_count"])
        if len(frame) != expected_signals or len(dates) != expected_signals:
            raise ValueError(f"{key} 信号数与回测汇总不一致")
        factor_rows = panel_by_date.loc[frame["date"]]
        invalid = (
            factor_rows[spec["rank_column"]].gt(rank_threshold)
            | factor_rows["margin_outflow_rank"].gt(rank_threshold)
            | factor_rows["down_pct"].lt(breadth_threshold)
            | ~factor_rows["margin_data_valid"].fillna(False)
            | ~factor_rows["breadth_valid"].fillna(False)
            | factor_rows["long_break_eve"].fillna(False)
        )
        if invalid.any():
            raise ValueError(f"{key} 中存在未同时满足三因子或长假规则的日期")
        signal_frames[key] = frame

    return {
        "audit": audit,
        "results": results,
        "panel": panel,
        "margin": margin,
        "signal_frames": signal_frames,
        "scenarios": scenarios,
    }


def scenario_flags(
    scenario: dict[str, object],
    results: dict[str, object],
) -> dict[str, bool]:
    rank_threshold = float(results["config"]["rank_threshold"])
    breadth_threshold = float(results["config"]["breadth_threshold"])
    margin_pass = float(scenario["margin_outflow_rank"]) <= rank_threshold
    breadth_pass = (
        float(scenario["down_pct"]) >= breadth_threshold
        and bool(scenario["breadth_valid"])
        and not bool(scenario["long_break_eve"])
    )
    flags = {
        "margin_pass": margin_pass,
        "breadth_pass": breadth_pass,
    }
    for key, spec in SIGNALS.items():
        flags[key] = (
            float(scenario[spec["rank_column"]]) <= rank_threshold
            and margin_pass
            and breadth_pass
        )
    return flags


def build_signal_scenario_row(
    data: dict[str, object],
    signal_key: str,
    scenario: dict[str, object],
) -> dict[str, object]:
    spec = SIGNALS[signal_key]
    results = data["results"]
    flags = scenario_flags(scenario, results)
    rank_threshold = int(float(results["config"]["rank_threshold"]))
    breadth_threshold = int(float(results["config"]["breadth_threshold"]))
    date = pd.Timestamp(str(scenario["date"]))
    index_pass = float(scenario[spec["rank_column"]]) <= rank_threshold
    row: dict[str, object] = {
        "信号类型": f"{spec['label']}三因子",
        "指数代码": spec["code"],
        "信号日": date,
        "收益起算日": date,
        "起算价格": "T日收盘",
        "样本状态": "估算场景（未纳入正式统计）",
        "前后期": "2024年后期",
        "指数当日涨跌幅(%)": scenario[spec["return_column"]],
        "指数跌幅三年排名": scenario[spec["rank_column"]],
        f"指数排名Top{rank_threshold}": "是" if index_pass else "否",
        "沪市融资余额(亿元)": scenario["sh_margin_y"],
        "深市融资余额(亿元)": scenario["sz_margin_y"],
        "两市融资余额(亿元)": scenario["total_margin_y"],
        "融资余额变动(亿元)": scenario["margin_change_amount"],
        "融资流出比例(%)": scenario["margin_outflow_pct"],
        "融资流出三年排名": scenario["margin_outflow_rank"],
        f"融资排名Top{rank_threshold}": "是" if flags["margin_pass"] else "否",
        "宽度可比股票数": scenario["breadth_total"],
        "宽度覆盖率": scenario["breadth_coverage"],
        "收跌股票占比(%)": scenario["down_pct"],
        f"收跌占比达到{breadth_threshold}%": "是" if flags["breadth_pass"] else "否",
        "宽度数据有效": "是" if bool(scenario["breadth_valid"]) else "否",
        "长假前日期": "是" if bool(scenario["long_break_eve"]) else "否",
        "三因子同时满足": "是" if flags[signal_key] else "否",
    }
    for outcome_label in OUTCOMES.values():
        for horizon in HORIZONS:
            row[f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"] = None
    return row


def build_signal_table(data: dict[str, object], signal_key: str) -> pd.DataFrame:
    spec = SIGNALS[signal_key]
    results = data["results"]
    raw = data["signal_frames"][signal_key].copy()
    margin = data["margin"][["date", "sh_margin_y", "sz_margin_y", "total_margin_y"]]
    panel_columns = ["date", "breadth_valid", "long_break_eve"]
    raw = raw.merge(margin, on="date", how="left", validate="one_to_one")
    raw = raw.merge(data["panel"][panel_columns], on="date", how="left", validate="one_to_one")

    raw["validation_period"] = raw["date"].ge(pd.Timestamp(results["config"]["validation_date"])).map(
        {True: "2024年后期", False: "2024年前期"}
    )
    rank_threshold = float(results["config"]["rank_threshold"])
    breadth_threshold = float(results["config"]["breadth_threshold"])
    index_pass = raw[spec["rank_column"]].le(rank_threshold)
    margin_pass = raw["margin_outflow_rank"].le(rank_threshold)
    breadth_pass = raw["down_pct"].ge(breadth_threshold)
    triple_pass = index_pass & margin_pass & breadth_pass

    output = pd.DataFrame(
        {
            "信号类型": f"{spec['label']}三因子",
            "指数代码": spec["code"],
            "信号日": raw["date"],
            "收益起算日": raw["date"],
            "起算价格": "T日收盘",
            "样本状态": "正式回测",
            "前后期": raw["validation_period"],
            "指数当日涨跌幅(%)": raw[spec["return_column"]],
            "指数跌幅三年排名": raw[spec["rank_column"]],
            f"指数排名Top{int(rank_threshold)}": yes_no(index_pass),
            "沪市融资余额(亿元)": raw["sh_margin_y"],
            "深市融资余额(亿元)": raw["sz_margin_y"],
            "两市融资余额(亿元)": raw["total_margin_y"],
            "融资余额变动(亿元)": raw["margin_change_amount"],
            "融资流出比例(%)": raw["margin_outflow_pct"],
            "融资流出三年排名": raw["margin_outflow_rank"],
            f"融资排名Top{int(rank_threshold)}": yes_no(margin_pass),
            "宽度可比股票数": raw["breadth_total"],
            "宽度覆盖率": raw["breadth_coverage"],
            "收跌股票占比(%)": raw["down_pct"],
            f"收跌占比达到{int(breadth_threshold)}%": yes_no(breadth_pass),
            "宽度数据有效": yes_no(raw["breadth_valid"]),
            "长假前日期": yes_no(raw["long_break_eve"]),
            "三因子同时满足": yes_no(triple_pass),
        }
    )
    for outcome_key, outcome_label in OUTCOMES.items():
        for horizon in HORIZONS:
            source = f"{outcome_key}_cc_t{horizon}"
            output[f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"] = raw[source]
    scenario_rows = [
        build_signal_scenario_row(data, signal_key, scenario)
        for scenario in data["scenarios"]
        if scenario_flags(scenario, results)[signal_key]
    ]
    if scenario_rows:
        output = pd.concat(
            [output, pd.DataFrame(scenario_rows, columns=output.columns)],
            ignore_index=True,
        )
    return output.sort_values("信号日").reset_index(drop=True)


def build_union_signal_table(data: dict[str, object]) -> pd.DataFrame:
    raw = (
        pd.concat(data["signal_frames"].values(), ignore_index=True)
        .drop_duplicates("date", keep="first")
        .sort_values("date")
        .reset_index(drop=True)
    )
    margin = data["margin"][["date", "sh_margin_y", "sz_margin_y", "total_margin_y"]]
    panel_columns = ["date", "breadth_valid", "long_break_eve"]
    raw = raw.merge(margin, on="date", how="left", validate="one_to_one")
    raw = raw.merge(data["panel"][panel_columns], on="date", how="left", validate="one_to_one")
    results = data["results"]
    rank_threshold = float(results["config"]["rank_threshold"])
    breadth_threshold = float(results["config"]["breadth_threshold"])

    output = pd.DataFrame(
        {
            "信号日": raw["date"],
            "收益起算日": raw["date"],
            "起算价格": "T日收盘",
            "样本状态": "正式回测",
            "前后期": raw["date"].ge(pd.Timestamp(results["config"]["validation_date"])).map(
                {True: "2024年后期", False: "2024年前期"}
            ),
        }
    )
    for key, spec in SIGNALS.items():
        raw_dates = set(data["signal_frames"][key]["date"])
        output[f"{spec['label']}三因子信号"] = yes_no(raw["date"].isin(raw_dates))

    for key, spec in SIGNALS.items():
        rank_column = spec["rank_column"]
        output[f"{spec['label']}当日涨跌幅(%)"] = raw[spec["return_column"]]
        output[f"{spec['label']}跌幅三年排名"] = raw[rank_column]
        output[f"{spec['label']}排名Top{int(rank_threshold)}"] = yes_no(
            raw[rank_column].le(rank_threshold)
        )

    output["沪市融资余额(亿元)"] = raw["sh_margin_y"]
    output["深市融资余额(亿元)"] = raw["sz_margin_y"]
    output["两市融资余额(亿元)"] = raw["total_margin_y"]
    output["融资余额变动(亿元)"] = raw["margin_change_amount"]
    output["融资流出比例(%)"] = raw["margin_outflow_pct"]
    output["融资流出三年排名"] = raw["margin_outflow_rank"]
    output[f"融资排名Top{int(rank_threshold)}"] = yes_no(
        raw["margin_outflow_rank"].le(rank_threshold)
    )
    output["宽度可比股票数"] = raw["breadth_total"]
    output["宽度覆盖率"] = raw["breadth_coverage"]
    output["收跌股票占比(%)"] = raw["down_pct"]
    output[f"收跌占比达到{int(breadth_threshold)}%"] = yes_no(
        raw["down_pct"].ge(breadth_threshold)
    )
    output["宽度数据有效"] = yes_no(raw["breadth_valid"])
    output["长假前日期"] = yes_no(raw["long_break_eve"])

    for outcome_key, outcome_label in OUTCOMES.items():
        for horizon in HORIZONS:
            output[f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"] = raw[
                f"{outcome_key}_cc_t{horizon}"
            ]

    scenario_rows: list[dict[str, object]] = []
    for scenario in data["scenarios"]:
        flags = scenario_flags(scenario, results)
        if not any(flags[key] for key in SIGNALS):
            continue
        date = pd.Timestamp(str(scenario["date"]))
        row: dict[str, object] = {
            "信号日": date,
            "收益起算日": date,
            "起算价格": "T日收盘",
            "样本状态": "估算场景（未纳入正式统计）",
            "前后期": "2024年后期",
        }
        for key, spec in SIGNALS.items():
            row[f"{spec['label']}三因子信号"] = "是" if flags[key] else "否"
            row[f"{spec['label']}当日涨跌幅(%)"] = scenario[spec["return_column"]]
            row[f"{spec['label']}跌幅三年排名"] = scenario[spec["rank_column"]]
            row[f"{spec['label']}排名Top{int(rank_threshold)}"] = (
                "是" if float(scenario[spec["rank_column"]]) <= rank_threshold else "否"
            )
        row.update(
            {
                "沪市融资余额(亿元)": scenario["sh_margin_y"],
                "深市融资余额(亿元)": scenario["sz_margin_y"],
                "两市融资余额(亿元)": scenario["total_margin_y"],
                "融资余额变动(亿元)": scenario["margin_change_amount"],
                "融资流出比例(%)": scenario["margin_outflow_pct"],
                "融资流出三年排名": scenario["margin_outflow_rank"],
                f"融资排名Top{int(rank_threshold)}": "是" if flags["margin_pass"] else "否",
                "宽度可比股票数": scenario["breadth_total"],
                "宽度覆盖率": scenario["breadth_coverage"],
                "收跌股票占比(%)": scenario["down_pct"],
                f"收跌占比达到{int(breadth_threshold)}%": (
                    "是" if flags["breadth_pass"] else "否"
                ),
                "宽度数据有效": "是" if bool(scenario["breadth_valid"]) else "否",
                "长假前日期": "是" if bool(scenario["long_break_eve"]) else "否",
            }
        )
        for outcome_label in OUTCOMES.values():
            for horizon in HORIZONS:
                row[f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"] = None
        scenario_rows.append(row)
    if scenario_rows:
        output = pd.concat(
            [output, pd.DataFrame(scenario_rows, columns=output.columns)],
            ignore_index=True,
        )
    return output.sort_values("信号日").reset_index(drop=True)


def dataframe_rows(frame: pd.DataFrame):
    for row in frame.itertuples(index=False, name=None):
        yield [
            None
            if pd.isna(value)
            else value.to_pydatetime()
            if isinstance(value, pd.Timestamp)
            else value
            for value in row
        ]


def add_table(ws, display_name: str, min_row: int, max_row: int, max_column: int) -> None:
    table = Table(
        displayName=display_name,
        ref=f"A{min_row}:{get_column_letter(max_column)}{max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_data_sheet(ws, frame: pd.DataFrame, table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    ws.row_dimensions[1].height = 36
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT_NAME, color=INPUT_BLUE)
            cell.alignment = Alignment(vertical="center")

    for index, header in enumerate(frame.columns, start=1):
        letter = get_column_letter(index)
        lengths = [len(str(header))]
        lengths.extend(len(str(value)) for value in frame.iloc[:, index - 1].dropna())
        ws.column_dimensions[letter].width = min(max(max(lengths, default=8) + 2, 11), 24)
        cells = ws[letter][1:]
        if "日期" in header or header in {"信号日", "收益起算日"}:
            for cell in cells:
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center")
        elif "覆盖率" in header:
            for cell in cells:
                cell.number_format = "0.0%"
        elif header in {
            "前后期",
            "样本状态",
            "宽度数据有效",
            "长假前日期",
            "三因子同时满足",
        } or "Top" in header or "达到" in header or "信号" in header:
            for cell in cells:
                cell.alignment = Alignment(horizontal="center")
        elif "排名" in header or "股票数" in header:
            for cell in cells:
                cell.number_format = "0"
        elif "亿元" in header:
            for cell in cells:
                cell.number_format = "#,##0.00;[Red](#,##0.00);-"
        elif "(%)" in header or "收益" in header:
            for cell in cells:
                cell.number_format = "0.00;[Red](0.00);-"
        if "收益" in header:
            data_range = f"{letter}2:{letter}{ws.max_row}"
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL),
            )
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL),
            )
    add_table(ws, table_name, 1, ws.max_row, ws.max_column)


def add_dataframe_sheet(
    wb: Workbook,
    name: str,
    frame: pd.DataFrame,
    table_name: str,
):
    ws = wb.create_sheet(name)
    ws.append(list(frame.columns))
    for row in dataframe_rows(frame):
        ws.append(row)
    style_data_sheet(ws, frame, table_name)
    return ws


def add_summary_header(ws, row: int, values: list[str]) -> None:
    for column, value in enumerate(values, start=1):
        cell = ws.cell(row, column, value)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_summary(
    wb: Workbook,
    data: dict[str, object],
) -> None:
    results = data["results"]
    audit = data["audit"]
    ws = wb.active
    ws.title = "回测汇总"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_title_rows = "20:20"
    ws.merge_cells("A1:I1")
    ws["A1"] = "三类指数严格三因子信号日与未来表现（2019年至今）"
    ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    if data["scenarios"]:
        scenario_note = "；".join(
            str(scenario.get("source_note") or f"{scenario['date']}为估算场景，未纳入正式统计")
            for scenario in data["scenarios"]
        )
    else:
        scenario_note = "无估算场景；未核验数据不得进入正式统计"

    notes = [
        ("回测区间", f"{results['metadata']['evaluation_start']} 至 {results['metadata']['data_end']}"),
        ("预热数据", f"自 {results['metadata']['warmup_data_start']} 起，用于滚动三年排名"),
        ("信号口径", "指数跌幅Top15 + 融资流出比例Top15 + 收跌股票占比>=80%，三项同时满足"),
        ("收益口径", "以信号日T日收盘价为起点、第N个交易日收盘价为终点；融资余额在收盘后才完整可得，因此含信号可得性未来函数，只能作事后统计"),
        ("指数代码", "深证综指399106；创业板指399006；创业板综指399102"),
        ("融资数据", f"{audit['verified_rows']}个共同交易日；完整性和文件哈希已通过审计门"),
        ("样本政策", "正式数据中所有满足三因子的交易日均计入，不做去聚类或冷却筛选"),
        ("统计限制", "连续信号及未来持有期可能重叠，样本并非相互独立；Wilson区间仅作描述"),
        ("限制/估算", scenario_note),
    ]
    for row, (label, value) in enumerate(notes, start=3):
        ws.cell(row, 1, label).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row, 2, value).font = Font(name=FONT_NAME, color=INPUT_BLUE)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[9].height = 36

    parameter_row = 12
    add_summary_header(ws, parameter_row, ["参数", "值"])
    parameters = [
        ("滚动窗口（年）", results["config"]["window_years"]),
        ("排名阈值", results["config"]["rank_threshold"]),
        ("收跌占比阈值（%）", results["config"]["breadth_threshold"]),
        ("最少窗口观测", results["config"]["min_window_observations"]),
    ]
    for row, (label, value) in enumerate(parameters, start=parameter_row + 1):
        ws.cell(row, 1, label).font = Font(name=FONT_NAME)
        ws.cell(row, 2, value).font = Font(name=FONT_NAME, color=INPUT_BLUE)
        ws.cell(row, 2).fill = ASSUMPTION_FILL

    count_row = 12
    add_summary_header(ws, count_row, ["参数", "值", "", "信号口径", "正式信号日数"])
    for offset, key in enumerate(SIGNALS, start=1):
        row = count_row + offset
        ws.cell(row, 4, f"{SIGNALS[key]['label']}三因子")
        ws.cell(row, 5, data["results"]["summaries"][key]["all_signals"]["signal_count"])
        ws.cell(row, 5).font = Font(name=FONT_NAME)
        ws.cell(row, 5).number_format = "0"

    performance_row = 20
    performance_headers = [
        "信号口径",
        "样本",
        "持有期",
        "有效样本",
        "胜率",
        "平均收益(%)",
        "中位收益(%)",
        "最差收益(%)",
        "最好收益(%)",
    ]
    add_summary_header(ws, performance_row, performance_headers)
    row = performance_row + 1
    for key, spec in SIGNALS.items():
        for horizon in HORIZONS:
            ws.cell(row, 1, f"{spec['label']}三因子")
            ws.cell(row, 2, PERFORMANCE_SAMPLE_LABEL)
            ws.cell(row, 3, f"T+{horizon}")
            stats = data["results"]["summaries"][key][PERFORMANCE_SAMPLE_KEY][
                spec["return_prefix"]
            ]["cc"][f"t{horizon}"]
            values = [
                stats["n"],
                None if stats["win_rate"] is None else stats["win_rate"] / 100.0,
                stats["mean"],
                stats["median"],
                stats["min"],
                stats["max"],
            ]
            for column, value in enumerate(values, start=4):
                cell = ws.cell(row, column, value)
                cell.font = Font(name=FONT_NAME)
                if column == 4:
                    cell.number_format = "0"
                elif column == 5:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "0.00;[Red](0.00);-"
            row += 1
    add_table(ws, "PerformanceSummary", performance_row, row - 1, len(performance_headers))

    widths = [20, 12, 12, 12, 12, 15, 15, 15, 15]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for cells in ws.iter_rows():
        for cell in cells:
            if cell.font.name is None:
                cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A21"


def create_cross_comparison(
    wb: Workbook,
    data: dict[str, object],
) -> None:
    rows: list[dict[str, object]] = []
    for key, spec in SIGNALS.items():
        for outcome_key, outcome_label in OUTCOMES.items():
            for horizon in HORIZONS:
                stats = data["results"]["summaries"][key][PERFORMANCE_SAMPLE_KEY][outcome_key][
                    "cc"
                ][f"t{horizon}"]
                rows.append(
                    {
                        "信号指数": spec["label"],
                        "样本": PERFORMANCE_SAMPLE_LABEL,
                        "表现指数": outcome_label,
                        "持有期": f"T+{horizon}",
                        "有效样本": stats["n"],
                        "胜率": None if stats["win_rate"] is None else stats["win_rate"] / 100.0,
                        "Wilson95%下限": (
                            None
                            if stats["win_rate_ci_low"] is None
                            else stats["win_rate_ci_low"] / 100.0
                        ),
                        "Wilson95%上限": (
                            None
                            if stats["win_rate_ci_high"] is None
                            else stats["win_rate_ci_high"] / 100.0
                        ),
                        "平均收益(%)": stats["mean"],
                        "中位收益(%)": stats["median"],
                        "最差收益(%)": stats["min"],
                        "最好收益(%)": stats["max"],
                    }
                )
    frame = pd.DataFrame(rows)
    ws = add_dataframe_sheet(wb, "交叉表现", frame, "CrossPerformance")
    for column in range(5, 13):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.font = Font(name=FONT_NAME)
                if column == 5:
                    item.number_format = "0"
                elif column in (6, 7, 8):
                    item.number_format = "0.0%"
                else:
                    item.number_format = "0.00;[Red](0.00);-"


def build_sensitivity(backtest_dir: Path) -> pd.DataFrame:
    frame = pd.read_csv(backtest_dir / "sensitivity.csv")
    labels = {key: spec["label"] for key, spec in SIGNALS.items()}
    frame["signal"] = frame["signal"].map(labels)
    return frame.rename(
        columns={
            "signal": "信号指数",
            "index_code": "指数代码",
            "window_years": "滚动窗口(年)",
            "rank_threshold": "排名阈值",
            "breadth_threshold": "收跌占比阈值(%)",
            "n": "有效样本",
            "t2_from_t_close_win_rate": "T收盘至T+2收盘胜率(%)",
            "t2_from_t_close_mean": "T收盘至T+2收盘平均收益(%)",
        }
    )


def build_dictionary(audit: dict[str, object]) -> pd.DataFrame:
    official_szse_checks = audit.get("official_szse_checks", "N/A")
    official_szse_scope = audit.get("official_szse_validation_scope", "N/A")
    rows = [
        ("信号定义", "对应指数跌幅三年排名Top15、融资流出比例三年排名Top15、收跌股票占比>=80%，三项同时满足"),
        ("信号日", "因子在当日收盘后满足条件的交易日"),
        ("收益起算日", "与信号日相同，记为T日"),
        ("起算价格", "按用户指定使用T日收盘价"),
        ("样本状态", "正式回测计入汇总统计；估算场景只展示、不纳入样本数、胜率和平均收益"),
        ("可执行性警告", "融资余额在T日收盘后才完整可得；T日收盘口径含信号可得性未来函数，不能视为可执行收益"),
        ("指数跌幅三年排名", "只使用当日及之前完整3个日历年数据；1为窗口内最差"),
        ("融资流出比例(%)", "-(当日两市融资余额-前一交易日余额)/前一交易日余额"),
        ("融资流出三年排名", "只使用当日及之前完整3个日历年数据；1为窗口内最大流出"),
        ("收跌股票占比(%)", "当日收跌股票数/当日有可比收盘价股票数"),
        ("长假前日期", "当日至下一交易日间隔至少5个日历日；此类日期已在信号生成前排除"),
        ("信号样本政策", "所有满足三因子的交易日均计入，不做去聚类或冷却筛选"),
        ("统计依赖性", "连续信号及未来持有期可能重叠；Wilson区间仅作描述，不能视为独立样本推断"),
        ("指数T收盘至T+N收盘收益(%)", "对应指数第N个交易日收盘价/T日收盘价-1；仅为事后条件统计"),
        ("空白未来收益", "数据截止日尚未走完对应持有期，不补写未来数据"),
        ("创业板口径", "创业板指399006与创业板综指399102分别回测，不混用"),
        (
            "数据限制",
            f"深市官方逐值核验覆盖{official_szse_checks}条（范围：{official_szse_scope}）；"
            "宽度使用当前本地证券池，存在幸存者偏差",
        ),
    ]
    return pd.DataFrame(rows, columns=["字段/项目", "说明"])


def build_parser() -> argparse.ArgumentParser:
    project_root = Path(__file__).resolve().parents[4]
    default_root = project_root / "artifacts" / "leverage_capitulation" / "verified_2016_present"
    parser = argparse.ArgumentParser(description="导出三类指数严格三因子信号日与未来表现工作簿")
    parser.add_argument("--backtest-dir", type=Path, default=default_root / "backtest_2019_present")
    parser.add_argument(
        "--margin-csv",
        type=Path,
        default=default_root / "official_margin_audit" / "verified_margin_balances.csv",
    )
    parser.add_argument(
        "--margin-audit",
        type=Path,
        default=default_root / "official_margin_audit" / "margin_audit.json",
    )
    parser.add_argument("--output", type=Path)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    output = args.output or args.backtest_dir / "three_index_triple_factor_comparison_2019_present.xlsx"
    data = load_inputs(args.backtest_dir, args.margin_csv, args.margin_audit)

    wb = Workbook()
    wb.properties.title = "三类指数严格三因子信号日与未来表现"
    wb.properties.creator = "Codex"
    wb.properties.description = (
        "2019年至今深证综指/创业板指/创业板综指三因子信号；原始日期完整保留，"
        "后期走势使用10交易日连续簇最后信号，T日收盘至T+N收盘"
    )
    counts: dict[str, dict[str, int]] = {}

    union = build_union_signal_table(data)
    add_dataframe_sheet(wb, "信号日总表", union, "SignalDayUnion")

    for index, (key, spec) in enumerate(SIGNALS.items(), start=1):
        signals = build_signal_table(data, key)
        add_dataframe_sheet(
            wb,
            spec["sheet_prefix"],
            signals,
            f"TripleSignals{index}",
        )
        counts[key] = {
            "formal": int(data["results"]["summaries"][key]["all_signals"]["signal_count"]),
            "displayed_with_scenarios": len(signals),
        }

    create_cross_comparison(wb, data)
    add_dataframe_sheet(
        wb,
        "敏感性分析",
        build_sensitivity(args.backtest_dir),
        "SensitivityAnalysis",
    )
    add_dataframe_sheet(wb, "字段说明", build_dictionary(data["audit"]), "FieldDictionary")
    create_summary(wb, data)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f"{output.stem}.tmp{output.suffix}")
    wb.save(temporary)
    temporary.replace(output)
    print(
        json.dumps(
            {
                "output": str(output.resolve()),
                "signals": counts,
                "formal_union_signal_dates": len(
                    set().union(*(set(frame["date"]) for frame in data["signal_frames"].values()))
                ),
                "displayed_union_signal_dates": len(union),
                "estimated_scenario_dates": len(data["scenarios"]),
                "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
