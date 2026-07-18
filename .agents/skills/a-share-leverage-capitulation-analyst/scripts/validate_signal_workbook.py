from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "回测汇总",
    "信号日总表",
    "深证综三因子",
    "创业板指三因子",
    "创业板综指三因子",
    "交叉表现",
    "敏感性分析",
    "字段说明",
]
SPECS = {
    "sz_triple": {
        "label": "深证综指",
        "sheet": "深证综三因子",
        "outcome": "sz_comp",
        "rank_column": "sz_comp_rank",
        "return_column": "sz_comp_return_pct",
    },
    "chinext_triple": {
        "label": "创业板指",
        "sheet": "创业板指三因子",
        "outcome": "chinext",
        "rank_column": "chinext_rank",
        "return_column": "chinext_return_pct",
    },
    "chinext_comp_triple": {
        "label": "创业板综指",
        "sheet": "创业板综指三因子",
        "outcome": "chinext_comp",
        "rank_column": "chinext_comp_rank",
        "return_column": "chinext_comp_return_pct",
    },
}
OUTCOMES = {
    "深证综指": "sz_comp",
    "创业板指": "chinext",
    "创业板综指": "chinext_comp",
}
HORIZONS = (1, 2, 5, 10, 20, 40)
PERFORMANCE_SAMPLE_KEY = "terminal_10d"
PERFORMANCE_SAMPLE_LABEL = "10日连续簇最后信号（事后）"
EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def close_enough(actual: object, expected: object) -> bool:
    if actual is None or expected is None:
        return actual is None and expected is None
    return math.isclose(float(actual), float(expected), rel_tol=0.0, abs_tol=1e-10)


def header_map(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}


def csv_dates(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [row["date"] for row in csv.DictReader(handle)]


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def close_return(
    panel_rows: list[dict[str, str]],
    panel_index: dict[str, int],
    date: str,
    outcome_key: str,
    horizon: int,
) -> float | None:
    start_index = panel_index[date]
    end_index = start_index + horizon
    if end_index >= len(panel_rows):
        return None
    start_close = float(panel_rows[start_index][f"{outcome_key}_close"])
    end_close = float(panel_rows[end_index][f"{outcome_key}_close"])
    return (end_close / start_close - 1.0) * 100.0


def scenario_qualifies(
    scenario: dict[str, object],
    signal_key: str,
    results: dict[str, object],
) -> bool:
    rank_threshold = float(results["config"]["rank_threshold"])
    breadth_threshold = float(results["config"]["breadth_threshold"])
    spec = SPECS[signal_key]
    return (
        float(scenario[spec["rank_column"]]) <= rank_threshold
        and float(scenario["margin_outflow_rank"]) <= rank_threshold
        and float(scenario["down_pct"]) >= breadth_threshold
        and bool(scenario["breadth_valid"])
        and not bool(scenario["long_break_eve"])
    )


def validate(workbook_path: Path, backtest_dir: Path) -> dict[str, object]:
    results = json.loads((backtest_dir / "backtest_results.json").read_text(encoding="utf-8"))
    scenario_path = backtest_dir / "estimated_signal_scenarios.json"
    scenarios: list[dict[str, object]] = []
    if scenario_path.exists():
        scenario_payload = json.loads(scenario_path.read_text(encoding="utf-8"))
        if scenario_payload.get("formal_statistics_inclusion") is not False:
            raise ValueError("估算场景错误地声明可纳入正式统计")
        scenarios = list(scenario_payload.get("scenarios", []))
    scenario_by_date = {str(item["date"]): item for item in scenarios}
    panel_rows = csv_rows(backtest_dir / "factor_panel.csv")
    panel_index = {row["date"]: index for index, row in enumerate(panel_rows)}
    wb = load_workbook(workbook_path, data_only=False)
    if wb.sheetnames != EXPECTED_SHEETS:
        raise ValueError(f"工作表顺序或名称不匹配: {wb.sheetnames}")

    formulas: list[str] = []
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{ws.title}!{cell.coordinate}")
                if isinstance(cell.value, str) and any(error in cell.value for error in EXCEL_ERRORS):
                    errors.append(f"{ws.title}!{cell.coordinate}")
    if formulas:
        raise ValueError(f"工作簿应为免重算静态统计，但仍有公式: {formulas[:10]}")
    if errors:
        raise ValueError(f"工作簿包含Excel错误: {errors[:10]}")

    union_dates: set[str] = set()
    for key, spec in SPECS.items():
        expected_signals = int(results["summaries"][key]["all_signals"]["signal_count"])
        expected_scenario_dates = sorted(
            str(item["date"])
            for item in scenarios
            if scenario_qualifies(item, key, results)
        )
        signal_ws = wb[spec["sheet"]]
        if signal_ws.max_row - 1 != expected_signals + len(expected_scenario_dates):
            raise ValueError(f"{key} 信号行数不匹配")

        source_dates = csv_dates(backtest_dir / f"signals_{key}.csv")
        union_dates.update(source_dates)
        headers = header_map(signal_ws)
        if "下一交易日开盘入场日" in headers:
            raise ValueError(f"{key} 仍包含旧的下一交易日开盘口径")
        for required in ("信号日", "收益起算日", "起算价格", "样本状态"):
            if required not in headers:
                raise ValueError(f"{key} 缺少T日收盘口径字段: {required}")
        formal_rows = [
            row
            for row in range(2, signal_ws.max_row + 1)
            if signal_ws.cell(row, headers["样本状态"]).value == "正式回测"
        ]
        scenario_rows = [
            row
            for row in range(2, signal_ws.max_row + 1)
            if signal_ws.cell(row, headers["样本状态"]).value == "估算场景（未纳入正式统计）"
        ]
        workbook_dates = [
            signal_ws.cell(row, headers["信号日"]).value.strftime("%Y-%m-%d")
            for row in formal_rows
        ]
        workbook_scenario_dates = [
            signal_ws.cell(row, headers["信号日"]).value.strftime("%Y-%m-%d")
            for row in scenario_rows
        ]
        if workbook_dates != source_dates:
            raise ValueError(f"{key} 信号日期与CSV不匹配")
        if workbook_scenario_dates != expected_scenario_dates:
            raise ValueError(f"{key} 估算信号日期与场景文件不匹配")
        if len(formal_rows) + len(scenario_rows) != signal_ws.max_row - 1:
            raise ValueError(f"{key} 存在未知样本状态")
        for row in range(2, signal_ws.max_row + 1):
            date = signal_ws.cell(row, headers["信号日"]).value.strftime("%Y-%m-%d")
            status = signal_ws.cell(row, headers["样本状态"]).value
            if signal_ws.cell(row, headers["收益起算日"]).value != signal_ws.cell(
                row, headers["信号日"]
            ).value:
                raise ValueError(f"{key} 收益起算日不是信号日: 第{row}行")
            if signal_ws.cell(row, headers["起算价格"]).value != "T日收盘":
                raise ValueError(f"{key} 起算价格不是T日收盘: 第{row}行")
            if signal_ws.cell(row, headers["三因子同时满足"]).value != "是":
                raise ValueError(f"{key} 存在未同时满足三因子的行")
            if signal_ws.cell(row, headers["长假前日期"]).value != "否":
                raise ValueError(f"{key} 存在未剔除的长假前日期")
            for outcome_label, outcome_key in OUTCOMES.items():
                for horizon in HORIZONS:
                    return_header = f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"
                    if return_header not in headers:
                        raise ValueError(f"{key} 缺少T日收盘收益字段: {return_header}")
                    actual = signal_ws.cell(row, headers[return_header]).value
                    expected = (
                        close_return(panel_rows, panel_index, date, outcome_key, horizon)
                        if status == "正式回测"
                        else None
                    )
                    if not close_enough(actual, expected):
                        raise ValueError(
                            f"{key} 明细收益不是close(T+N)/close(T)-1: 第{row}行 {return_header}"
                        )
            if status == "估算场景（未纳入正式统计）":
                scenario = scenario_by_date[date]
                expected_fields = {
                    "指数当日涨跌幅(%)": scenario[spec["return_column"]],
                    "指数跌幅三年排名": scenario[spec["rank_column"]],
                    "融资余额变动(亿元)": scenario["margin_change_amount"],
                    "融资流出比例(%)": scenario["margin_outflow_pct"],
                    "融资流出三年排名": scenario["margin_outflow_rank"],
                    "收跌股票占比(%)": scenario["down_pct"],
                }
                for name, target in expected_fields.items():
                    if not close_enough(signal_ws.cell(row, headers[name]).value, target):
                        raise ValueError(f"{key} 估算场景字段不匹配: 第{row}行 {name}")

    union_ws = wb["信号日总表"]
    union_scenario_dates = sorted(
        str(item["date"])
        for item in scenarios
        if any(scenario_qualifies(item, key, results) for key in SPECS)
    )
    if union_ws.max_row - 1 != len(union_dates) + len(union_scenario_dates):
        raise ValueError("信号日总表行数与三组信号并集不匹配")
    union_headers = header_map(union_ws)
    if "下一交易日开盘入场日" in union_headers:
        raise ValueError("信号日总表仍包含旧的下一交易日开盘口径")
    union_formal_dates: list[str] = []
    union_displayed_scenario_dates: list[str] = []
    for row in range(2, union_ws.max_row + 1):
        date = union_ws.cell(row, union_headers["信号日"]).value.strftime("%Y-%m-%d")
        status = union_ws.cell(row, union_headers["样本状态"]).value
        if status == "正式回测":
            union_formal_dates.append(date)
        elif status == "估算场景（未纳入正式统计）":
            union_displayed_scenario_dates.append(date)
        else:
            raise ValueError(f"信号日总表存在未知样本状态: 第{row}行")
        if union_ws.cell(row, union_headers["收益起算日"]).value != union_ws.cell(
            row, union_headers["信号日"]
        ).value:
            raise ValueError(f"信号日总表收益起算日不是信号日: 第{row}行")
        if union_ws.cell(row, union_headers["起算价格"]).value != "T日收盘":
            raise ValueError(f"信号日总表起算价格不是T日收盘: 第{row}行")
        if status == "估算场景（未纳入正式统计）":
            scenario = scenario_by_date[date]
            for key, spec in SPECS.items():
                expected_flag = "是" if scenario_qualifies(scenario, key, results) else "否"
                if union_ws.cell(row, union_headers[f"{spec['label']}三因子信号"]).value != expected_flag:
                    raise ValueError(f"信号日总表估算场景信号标记不匹配: 第{row}行 {key}")
            for outcome_label in OUTCOMES:
                for horizon in HORIZONS:
                    return_header = f"{outcome_label}T收盘至T+{horizon}收盘收益(%)"
                    if union_ws.cell(row, union_headers[return_header]).value is not None:
                        raise ValueError(f"估算场景未来收益必须为空: 第{row}行 {return_header}")
    if union_formal_dates != sorted(union_dates):
        raise ValueError("信号日总表正式日期与三组信号并集不匹配")
    if union_displayed_scenario_dates != union_scenario_dates:
        raise ValueError("信号日总表估算日期与场景文件不匹配")

    summary = wb["回测汇总"]
    summary_rows = {
        (
            summary.cell(row, 1).value,
            summary.cell(row, 2).value,
            summary.cell(row, 3).value,
        ): [summary.cell(row, column).value for column in range(4, 10)]
        for row in range(21, summary.max_row + 1)
    }
    for key, spec in SPECS.items():
        for horizon in HORIZONS:
            stats = results["summaries"][key][PERFORMANCE_SAMPLE_KEY][spec["outcome"]]["cc"][
                f"t{horizon}"
            ]
            actual = summary_rows[(f"{spec['label']}三因子", PERFORMANCE_SAMPLE_LABEL, f"T+{horizon}")]
            expected = [
                stats["n"],
                None if stats["win_rate"] is None else stats["win_rate"] / 100.0,
                stats["mean"],
                stats["median"],
                stats["min"],
                stats["max"],
            ]
            if not all(close_enough(item, target) for item, target in zip(actual, expected)):
                raise ValueError(f"回测汇总统计不匹配: {key} 全部信号日 T+{horizon}")

    cross = wb["交叉表现"]
    if cross.max_row - 1 != len(SPECS) * len(OUTCOMES) * len(HORIZONS):
        raise ValueError("交叉表现行数不匹配")
    cross_headers = header_map(cross)
    signal_lookup = {spec["label"]: key for key, spec in SPECS.items()}
    for row in range(2, cross.max_row + 1):
        signal_key = signal_lookup[cross.cell(row, cross_headers["信号指数"]).value]
        if cross.cell(row, cross_headers["样本"]).value != PERFORMANCE_SAMPLE_LABEL:
            raise ValueError(f"交叉表现样本口径不匹配: 第{row}行")
        outcome_key = OUTCOMES[cross.cell(row, cross_headers["表现指数"]).value]
        horizon_key = cross.cell(row, cross_headers["持有期"]).value.replace("T+", "t")
        stats = results["summaries"][signal_key][PERFORMANCE_SAMPLE_KEY][outcome_key]["cc"][horizon_key]
        expected = {
            "有效样本": stats["n"],
            "胜率": stats["win_rate"] / 100.0,
            "Wilson95%下限": stats["win_rate_ci_low"] / 100.0,
            "Wilson95%上限": stats["win_rate_ci_high"] / 100.0,
            "平均收益(%)": stats["mean"],
            "中位收益(%)": stats["median"],
            "最差收益(%)": stats["min"],
            "最好收益(%)": stats["max"],
        }
        for name, target in expected.items():
            if not close_enough(cross.cell(row, cross_headers[name]).value, target):
                raise ValueError(f"交叉表现不匹配: 第{row}行 {name}")

    sensitivity_rows = wb["敏感性分析"].max_row - 1
    if sensitivity_rows != len(results["sensitivity"]):
        raise ValueError("敏感性分析行数与JSON不匹配")
    wb.close()
    return {
        "status": "success",
        "sheets": len(EXPECTED_SHEETS),
        "formal_union_signal_dates": len(union_dates),
        "displayed_union_signal_dates": len(union_dates) + len(union_scenario_dates),
        "estimated_scenario_dates": len(union_scenario_dates),
        "cross_rows": cross.max_row - 1,
        "sensitivity_rows": sensitivity_rows,
        "formula_cells": len(formulas),
        "error_cells": len(errors),
    }


def build_parser() -> argparse.ArgumentParser:
    project_root = Path(__file__).resolve().parents[4]
    default_backtest = (
        project_root
        / "artifacts"
        / "leverage_capitulation"
        / "verified_2016_present"
        / "backtest_2019_present"
    )
    parser = argparse.ArgumentParser(description="校验三指数三因子回测工作簿")
    parser.add_argument("--backtest-dir", type=Path, default=default_backtest)
    parser.add_argument("--workbook", type=Path)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    workbook = args.workbook or args.backtest_dir / "three_index_triple_factor_comparison_2019_present.xlsx"
    print(json.dumps(validate(workbook, args.backtest_dir), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
