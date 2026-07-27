from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "validate_company_tracking_run.py"

WATCHLIST_HEADERS = [
    "enabled",
    "ticker",
    "exchange",
    "name",
    "aliases",
    "industry_tags",
    "priority",
    "baseline_status",
    "tracking_focus",
    "official_sources_hint",
    "last_baseline_date",
    "last_update_date",
    "notes",
]

COMPLETION_HEADERS = [
    "ticker",
    "name",
    "batch_no",
    "queue_status",
    "collection_scope",
    "announcements_checked",
    "lhb_checked",
    "block_trade_checked",
    "announcement_window_checked",
    "open_web_search_status",
    "state_change",
    "miss_risk_notes",
]


def current_event(ticker: str, name: str, title: str = "正式公告") -> dict[str, object]:
    attribution = {
        dimension: {
            "evidence": "N/A",
            "counterevidence": "N/A",
            "confidence": "low",
            "persistence_window": "30d",
            "next_validation": "下一份公告",
        }
        for dimension in ("company", "regulatory", "peer", "industry", "market")
    }
    return {
        "date": "2026-07-27",
        "ticker": ticker,
        "name": name,
        "source_type": "official_announcement",
        "source_name": "CNINFO",
        "title": title,
        "url": f"https://example.test/{ticker}/{title}",
        "summary": "正式公告摘要",
        "verification_status": "confirmed_official",
        "change_type": "fact_change",
        "hard_evidence_new": True,
        "assumption_ids": ["A1"],
        "thesis_effect": "unchanged",
        "previous_commercialization_stage": "N/A",
        "new_commercialization_stage": "N/A",
        "stage_evidence": "N/A",
        "stage_evidence_date": "N/A",
        "stage_source": "N/A",
        "revenue_materiality": "N/A",
        "attribution_dimensions": attribution,
        "evidence": ["CNINFO 原文"],
        "counterevidence": ["未披露财务影响"],
        "confidence": "high",
        "persistence_window": "90d",
        "next_validation": "下一期定期报告",
    }


class CompanyTrackingRunE2ETests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary.cleanup)
        self.base = Path(self.temporary.name)
        self.watchlist = self.base / "watchlist.xlsx"
        self.events_root = self.base / "company_tracking"
        self.run_status = self.base / "run_status.md"
        self.snapshot = self.base / "snapshot.json"
        self._write_workbook()
        self._write_initial_events()
        result = self._run(
            "snapshot",
            "--watchlist",
            str(self.watchlist),
            "--events-root",
            str(self.events_root),
            "--output",
            str(self.snapshot),
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def _run(self, *arguments: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), *arguments],
            cwd=ROOT,
            text=True,
            encoding="utf-8",
            capture_output=True,
            check=False,
        )

    def _validate(self) -> subprocess.CompletedProcess[str]:
        return self._run(
            "validate",
            "--snapshot",
            str(self.snapshot),
            "--watchlist",
            str(self.watchlist),
            "--events-root",
            str(self.events_root),
            "--run-status",
            str(self.run_status),
        )

    def _write_workbook(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "watchlist"
        worksheet.append(WATCHLIST_HEADERS)
        worksheet.append(
            [
                "Y",
                "000001.SZ",
                "SZSE",
                "甲公司",
                "甲",
                "测试",
                1,
                "pending",
                "公告",
                "CNINFO",
                "",
                "",
                "重点",
            ]
        )
        worksheet.append(
            [
                "Y",
                "600001.SH",
                "SSE",
                "乙公司",
                "乙",
                "测试",
                2,
                "done",
                "公告",
                "SSE",
                "2026-07-01",
                "2026-07-26",
                "重点",
            ]
        )
        worksheet.append(
            [
                "N",
                "300001.SZ",
                "SZSE",
                "丙公司",
                "丙",
                "测试",
                3,
                "disabled",
                "公告",
                "CNINFO",
                "",
                "",
                "不跟踪",
            ]
        )
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = "A1:M4"
        worksheet.column_dimensions["B"].width = 14
        enabled_validation = DataValidation(type="list", formula1='"Y,N"')
        worksheet.add_data_validation(enabled_validation)
        enabled_validation.add("A2:A100")

        schema = workbook.create_sheet("schema")
        schema.append(["field", "description"])
        schema.append(["enabled", "Y means tracked"])
        schema.freeze_panes = "A2"
        workbook.save(self.watchlist)
        workbook.close()

    def _write_initial_events(self) -> None:
        companies = (("000001.SZ", "甲公司"), ("600001.SH", "乙公司"))
        for ticker, name in companies:
            directory = self.events_root / ticker
            directory.mkdir(parents=True)
            legacy = {"date": "2026-07-26", "ticker": ticker, "name": name, "legacy": True}
            (directory / "events.jsonl").write_text(
                json.dumps(legacy, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )

    def _write_completion(self, rows: list[list[str]]) -> None:
        lines = [
            "# run status",
            "",
            "## Per-company completion table",
            "",
            "| " + " | ".join(COMPLETION_HEADERS) + " |",
            "| " + " | ".join("---" for _ in COMPLETION_HEADERS) + " |",
        ]
        lines.extend("| " + " | ".join(row) + " |" for row in rows)
        self.run_status.write_text("\n".join(lines) + "\n", encoding="utf-8")

    @staticmethod
    def _completion_row(ticker: str, name: str, batch: int) -> list[str]:
        return [
            ticker,
            name,
            str(batch),
            "completed",
            "controller_open_web_only",
            "checked",
            "checked",
            "checked",
            "T_and_T_plus_1",
            "no_signal",
            "no_change",
            "none",
        ]

    def _append_event(self, ticker: str, event: dict[str, object]) -> None:
        with (self.events_root / ticker / "events.jsonl").open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(event, ensure_ascii=False) + "\n")

    def test_excel_jsonl_and_completion_table_round_trip_passes(self) -> None:
        workbook = load_workbook(self.watchlist)
        worksheet = workbook["watchlist"]
        worksheet["H2"] = "done"
        worksheet["K2"] = "2026-07-27"
        worksheet["L2"] = "2026-07-27"
        worksheet["L3"] = "2026-07-27"
        workbook.save(self.watchlist)
        workbook.close()

        self._append_event("000001.SZ", current_event("000001.SZ", "甲公司"))
        self._write_completion(
            [
                self._completion_row("000001.SZ", "甲公司", 1),
                self._completion_row("600001.SH", "乙公司", 2),
            ]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual(payload["status"], "passed")
        self.assertEqual(payload["enabled_company_count"], 2)
        self.assertEqual(payload["completion_table_count"], 2)
        self.assertEqual(payload["new_event_count"], 1)
        self.assertEqual(payload["workbook_round_trip"], "passed")

    def test_missing_or_duplicate_completion_ticker_fails(self) -> None:
        duplicate = self._completion_row("000001.SZ", "甲公司", 2)
        self._write_completion(
            [self._completion_row("000001.SZ", "甲公司", 1), duplicate]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 1)
        self.assertIn("未与 enabled watchlist 一一按序覆盖", result.stdout)
        self.assertIn("600001.SH", result.stdout)

    def test_new_event_missing_required_field_fails(self) -> None:
        event = current_event("000001.SZ", "甲公司")
        del event["change_type"]
        self._append_event("000001.SZ", event)
        self._write_completion(
            [
                self._completion_row("000001.SZ", "甲公司", 1),
                self._completion_row("600001.SH", "乙公司", 2),
            ]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 1)
        self.assertIn("新增 event 缺少字段", result.stdout)
        self.assertIn("change_type", result.stdout)

    def test_duplicate_new_event_identity_fails(self) -> None:
        event = current_event("000001.SZ", "甲公司")
        self._append_event("000001.SZ", event)
        self._append_event("000001.SZ", event)
        self._write_completion(
            [
                self._completion_row("000001.SZ", "甲公司", 1),
                self._completion_row("600001.SH", "乙公司", 2),
            ]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 1)
        self.assertIn("新增 event 身份重复", result.stdout)

    def test_excel_non_allowed_change_fails(self) -> None:
        workbook = load_workbook(self.watchlist)
        worksheet = workbook["watchlist"]
        worksheet["I2"] = "未经允许改写的跟踪重点"
        workbook.save(self.watchlist)
        workbook.close()
        self._write_completion(
            [
                self._completion_row("000001.SZ", "甲公司", 1),
                self._completion_row("600001.SH", "乙公司", 2),
            ]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 1)
        self.assertIn("Excel 工作簿关键结构或非允许内容发生变化", result.stdout)

    def test_excel_structure_change_fails(self) -> None:
        workbook = load_workbook(self.watchlist)
        worksheet = workbook["watchlist"]
        worksheet.freeze_panes = None
        workbook.save(self.watchlist)
        workbook.close()
        self._write_completion(
            [
                self._completion_row("000001.SZ", "甲公司", 1),
                self._completion_row("600001.SH", "乙公司", 2),
            ]
        )

        result = self._validate()

        self.assertEqual(result.returncode, 1)
        self.assertIn("Excel 工作簿关键结构或非允许内容发生变化", result.stdout)
        self.assertIn("freeze_panes", result.stdout)


if __name__ == "__main__":
    unittest.main()
