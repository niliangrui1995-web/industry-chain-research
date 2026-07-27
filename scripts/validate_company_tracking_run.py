from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SNAPSHOT_SCHEMA_VERSION = 1

REQUIRED_WATCHLIST_COLUMNS = {
    "enabled",
    "ticker",
    "exchange",
    "name",
    "aliases",
    "industry_tags",
    "priority",
    "baseline_status",
    "tracking_focus",
    "official_sources_hint",
    "last_baseline_date",
    "last_update_date",
    "notes",
}

MUTABLE_WATCHLIST_COLUMNS = {
    "baseline_status",
    "last_baseline_date",
    "last_update_date",
}

REQUIRED_COMPLETION_COLUMNS = (
    "ticker",
    "name",
    "batch_no",
    "queue_status",
    "collection_scope",
    "announcements_checked",
    "lhb_checked",
    "block_trade_checked",
    "announcement_window_checked",
    "open_web_search_status",
    "state_change",
    "miss_risk_notes",
)

TERMINAL_QUEUE_STATUSES = {
    "completed",
    "completed_with_open_web_gap",
    "failed_with_reason",
}

REQUIRED_EVENT_FIELDS = {
    "date",
    "ticker",
    "name",
    "source_type",
    "source_name",
    "title",
    "url",
    "summary",
    "verification_status",
    "change_type",
    "hard_evidence_new",
    "assumption_ids",
    "thesis_effect",
    "previous_commercialization_stage",
    "new_commercialization_stage",
    "stage_evidence",
    "stage_evidence_date",
    "stage_source",
    "revenue_materiality",
    "attribution_dimensions",
    "evidence",
    "counterevidence",
    "confidence",
    "persistence_window",
    "next_validation",
}

REQUIRED_ATTRIBUTION_DIMENSIONS = {
    "company",
    "regulatory",
    "peer",
    "industry",
    "market",
}

REQUIRED_ATTRIBUTION_FIELDS = {
    "evidence",
    "counterevidence",
    "confidence",
    "persistence_window",
    "next_validation",
}

EVENT_IDENTITY_FIELDS = ("date", "ticker", "source_type", "title", "url")


class ContractError(ValueError):
    """Raised when a company-tracking run violates its durable artifact contract."""


@dataclass(frozen=True)
class EnabledCompany:
    ticker: str
    name: str
    row: int


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _json_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return {"type": type(value).__name__, "value": str(value)}


def _style_signature(cell) -> list[int]:
    if cell._style is None:
        return []
    values = list(cell._style)
    return [] if not any(values) else values


def _header_map(worksheet) -> dict[str, int]:
    headers = [_clean_text(cell.value) for cell in worksheet[1]]
    if any(not header for header in headers):
        raise ContractError("watchlist 表头包含空字段")
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise ContractError(f"watchlist 表头重复: {', '.join(duplicates)}")
    missing = sorted(REQUIRED_WATCHLIST_COLUMNS - set(headers))
    if missing:
        raise ContractError(f"watchlist 缺少必需字段: {', '.join(missing)}")
    return {header: index + 1 for index, header in enumerate(headers)}


def read_enabled_companies(path: Path) -> tuple[list[EnabledCompany], dict[str, int]]:
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl supplies many exception types
        raise ContractError(f"Excel watchlist 无法打开: {path}: {exc}") from exc
    try:
        if "watchlist" not in workbook.sheetnames:
            raise ContractError("Excel 缺少 watchlist 工作表")
        worksheet = workbook["watchlist"]
        columns = _header_map(worksheet)
        enabled: list[EnabledCompany] = []
        seen: set[str] = set()
        for row in range(2, worksheet.max_row + 1):
            flag = _clean_text(worksheet.cell(row, columns["enabled"]).value).upper()
            if flag != "Y":
                continue
            ticker = _clean_text(worksheet.cell(row, columns["ticker"]).value)
            name = _clean_text(worksheet.cell(row, columns["name"]).value)
            if not ticker or not name:
                raise ContractError(f"watchlist 第 {row} 行 enabled 公司缺少 ticker 或 name")
            if ticker in seen:
                raise ContractError(f"watchlist enabled ticker 重复: {ticker}")
            seen.add(ticker)
            enabled.append(EnabledCompany(ticker=ticker, name=name, row=row))
        return enabled, columns
    finally:
        workbook.close()


def _data_validation_signature(worksheet) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for validation in worksheet.data_validations.dataValidation:
        result.append(
            {
                "sqref": str(validation.sqref),
                "type": validation.type,
                "operator": validation.operator,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "allow_blank": validation.allow_blank,
                "show_error_message": validation.showErrorMessage,
                "show_input_message": validation.showInputMessage,
                "error": validation.error,
                "error_title": validation.errorTitle,
                "prompt": validation.prompt,
                "prompt_title": validation.promptTitle,
            }
        )
    return sorted(result, key=lambda item: (item["sqref"], _clean_text(item["type"])))


def _dimension_signature(dimensions: Iterable[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, dimension in dimensions:
        result[str(key)] = {
            "width": getattr(dimension, "width", None),
            "height": getattr(dimension, "height", None),
            "hidden": getattr(dimension, "hidden", False),
            "outline_level": getattr(dimension, "outlineLevel", 0),
            "collapsed": getattr(dimension, "collapsed", False),
            "best_fit": getattr(dimension, "bestFit", False),
        }
    return result


def workbook_fingerprint(
    path: Path, mutable_cells: set[str] | None = None
) -> dict[str, Any]:
    mutable_cells = mutable_cells or set()
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl supplies many exception types
        raise ContractError(f"Excel watchlist 无法重开: {path}: {exc}") from exc
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            cells: list[dict[str, Any]] = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    qualified = f"{worksheet.title}!{cell.coordinate}"
                    entry = {
                        "coordinate": cell.coordinate,
                        "style": _style_signature(cell),
                    }
                    if qualified not in mutable_cells:
                        entry["value"] = _json_value(cell.value)
                    cells.append(entry)
            sheets.append(
                {
                    "title": worksheet.title,
                    "state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "freeze_panes": _clean_text(worksheet.freeze_panes),
                    "auto_filter": _clean_text(worksheet.auto_filter.ref),
                    "merged_cells": sorted(str(item) for item in worksheet.merged_cells.ranges),
                    "data_validations": _data_validation_signature(worksheet),
                    "column_dimensions": _dimension_signature(
                        sorted(worksheet.column_dimensions.items())
                    ),
                    "row_dimensions": _dimension_signature(
                        sorted(worksheet.row_dimensions.items())
                    ),
                    "cells": cells,
                }
            )
        return {"sheetnames": list(workbook.sheetnames), "sheets": sheets}
    finally:
        workbook.close()


def _read_jsonl(path: Path) -> tuple[bytes, list[dict[str, Any]]]:
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ContractError(f"events.jsonl 不是 UTF-8: {path}: {exc}") from exc
    records: list[dict[str, Any]] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        if not line.strip():
            raise ContractError(f"events.jsonl 含空行: {path}:{line_number}")
        try:
            record = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ContractError(f"events.jsonl JSON 无效: {path}:{line_number}: {exc}") from exc
        if not isinstance(record, dict):
            raise ContractError(f"events.jsonl 每行必须是 JSON object: {path}:{line_number}")
        records.append(record)
    return raw, records


def _event_prefix(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"exists": False, "byte_length": 0, "sha256": None, "record_count": 0}
    raw, records = _read_jsonl(path)
    return {
        "exists": True,
        "byte_length": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "record_count": len(records),
    }


def _mutable_cells(
    enabled: list[EnabledCompany], columns: dict[str, int]
) -> set[str]:
    return {
        f"watchlist!{load_workbook_column_letter(columns[column])}{company.row}"
        for company in enabled
        for column in MUTABLE_WATCHLIST_COLUMNS
    }


def create_snapshot(watchlist: Path, events_root: Path) -> dict[str, Any]:
    enabled, columns = read_enabled_companies(watchlist)
    mutable_cells = _mutable_cells(enabled, columns)
    return {
        "schema_version": SNAPSHOT_SCHEMA_VERSION,
        "enabled_companies": [
            {"ticker": company.ticker, "name": company.name, "row": company.row}
            for company in enabled
        ],
        "mutable_workbook_cells": sorted(mutable_cells),
        "workbook_fingerprint": workbook_fingerprint(watchlist, mutable_cells),
        "event_prefixes": {
            company.ticker: _event_prefix(events_root / company.ticker / "events.jsonl")
            for company in enabled
        },
    }


def load_workbook_column_letter(index: int) -> str:
    if index < 1:
        raise ValueError("Excel column index must be positive")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if not stripped.startswith("|") or not stripped.endswith("|"):
        raise ContractError(f"completion table 行必须以 | 开始和结束: {line}")
    cells: list[str] = []
    current: list[str] = []
    escaped = False
    for character in stripped[1:-1]:
        if escaped:
            current.append(character)
            escaped = False
        elif character == "\\":
            escaped = True
        elif character == "|":
            cells.append("".join(current).strip())
            current = []
        else:
            current.append(character)
    if escaped:
        current.append("\\")
    cells.append("".join(current).strip())
    return cells


def _is_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells)


def parse_completion_table(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    required = set(REQUIRED_COMPLETION_COLUMNS)
    for index in range(len(lines) - 1):
        if not lines[index].strip().startswith("|"):
            continue
        try:
            headers = _split_markdown_row(lines[index])
            separator = _split_markdown_row(lines[index + 1])
        except ContractError:
            continue
        if not required.issubset(headers) or len(headers) != len(separator):
            continue
        if not _is_separator_row(separator):
            continue
        if len(headers) != len(set(headers)):
            raise ContractError("completion table 表头重复")
        rows: list[dict[str, str]] = []
        for line in lines[index + 2 :]:
            if not line.strip():
                break
            if not line.strip().startswith("|"):
                break
            values = _split_markdown_row(line)
            if len(values) != len(headers):
                raise ContractError("completion table 数据列数与表头不一致")
            rows.append(dict(zip(headers, values)))
        return rows
    raise ContractError(f"未找到包含必需字段的 completion table: {path}")


def validate_completion_table(
    enabled: list[EnabledCompany], rows: list[dict[str, str]]
) -> None:
    for row_number, row in enumerate(rows, start=1):
        missing_values = [column for column in REQUIRED_COMPLETION_COLUMNS if not row.get(column, "").strip()]
        if missing_values:
            raise ContractError(
                f"completion table 第 {row_number} 行字段为空: {', '.join(missing_values)}"
            )
        if row["queue_status"] not in TERMINAL_QUEUE_STATUSES:
            raise ContractError(
                f"completion table {row['ticker']} queue_status 非终态: {row['queue_status']}"
            )

    actual_tickers = [row["ticker"] for row in rows]
    expected_tickers = [company.ticker for company in enabled]
    duplicates = sorted(
        {ticker for ticker in actual_tickers if actual_tickers.count(ticker) > 1}
    )
    missing = [ticker for ticker in expected_tickers if ticker not in actual_tickers]
    extra = [ticker for ticker in actual_tickers if ticker not in expected_tickers]
    if duplicates or missing or extra or actual_tickers != expected_tickers:
        raise ContractError(
            "completion table 未与 enabled watchlist 一一按序覆盖: "
            f"missing={missing or 'none'}, extra={extra or 'none'}, "
            f"duplicates={duplicates or 'none'}, order_match={actual_tickers == expected_tickers}"
        )
    expected_names = {company.ticker: company.name for company in enabled}
    for row in rows:
        if row["name"] != expected_names[row["ticker"]]:
            raise ContractError(
                f"completion table 公司名称不匹配: {row['ticker']} "
                f"expected={expected_names[row['ticker']]}, actual={row['name']}"
            )


def _require_nonempty_event_fields(record: dict[str, Any], context: str) -> None:
    missing = sorted(REQUIRED_EVENT_FIELDS - record.keys())
    if missing:
        raise ContractError(f"新增 event 缺少字段 {context}: {', '.join(missing)}")
    empty = sorted(
        field
        for field in REQUIRED_EVENT_FIELDS - {"assumption_ids", "attribution_dimensions"}
        if record[field] is None or (isinstance(record[field], str) and not record[field].strip())
    )
    if empty:
        raise ContractError(f"新增 event 字段为空 {context}: {', '.join(empty)}")
    if not isinstance(record["assumption_ids"], list):
        raise ContractError(f"新增 event assumption_ids 必须是数组 {context}")
    try:
        dt.date.fromisoformat(str(record["date"]))
    except ValueError as exc:
        raise ContractError(f"新增 event date 必须是 YYYY-MM-DD {context}") from exc

    dimensions = record["attribution_dimensions"]
    if not isinstance(dimensions, dict):
        raise ContractError(f"新增 event attribution_dimensions 必须是对象 {context}")
    missing_dimensions = sorted(REQUIRED_ATTRIBUTION_DIMENSIONS - dimensions.keys())
    if missing_dimensions:
        raise ContractError(
            f"新增 event 五维归因不完整 {context}: {', '.join(missing_dimensions)}"
        )
    for dimension in sorted(REQUIRED_ATTRIBUTION_DIMENSIONS):
        value = dimensions[dimension]
        if not isinstance(value, dict):
            raise ContractError(f"新增 event 归因维度 {dimension} 必须是对象 {context}")
        missing_fields = sorted(REQUIRED_ATTRIBUTION_FIELDS - value.keys())
        if missing_fields:
            raise ContractError(
                f"新增 event 归因维度 {dimension} 缺少字段 {context}: "
                f"{', '.join(missing_fields)}"
            )
        empty_fields = sorted(
            field
            for field in REQUIRED_ATTRIBUTION_FIELDS
            if value[field] is None
            or (isinstance(value[field], str) and not value[field].strip())
        )
        if empty_fields:
            raise ContractError(
                f"新增 event 归因维度 {dimension} 字段为空 {context}: "
                f"{', '.join(empty_fields)}"
            )


def _event_identity(record: dict[str, Any]) -> tuple[str, ...] | None:
    values = tuple(_clean_text(record.get(field)) for field in EVENT_IDENTITY_FIELDS)
    return values if all(values) else None


def validate_events(
    enabled: list[EnabledCompany],
    events_root: Path,
    prefixes: dict[str, Any],
) -> int:
    new_event_count = 0
    for company in enabled:
        path = events_root / company.ticker / "events.jsonl"
        if not path.exists():
            raise ContractError(f"enabled 公司缺少 events.jsonl: {company.ticker}")
        raw, records = _read_jsonl(path)
        prefix = prefixes.get(company.ticker)
        if not isinstance(prefix, dict):
            raise ContractError(f"快照缺少 event 前缀: {company.ticker}")
        byte_length = int(prefix.get("byte_length", 0))
        if len(raw) < byte_length:
            raise ContractError(f"events.jsonl 被截断，违反 append-only: {company.ticker}")
        if prefix.get("exists"):
            digest = hashlib.sha256(raw[:byte_length]).hexdigest()
            if digest != prefix.get("sha256"):
                raise ContractError(f"events.jsonl 历史前缀被改写: {company.ticker}")
        start = int(prefix.get("record_count", 0))
        if len(records) < start:
            raise ContractError(f"events.jsonl 记录数减少: {company.ticker}")

        seen: dict[tuple[str, ...], int] = {}
        for line_number, record in enumerate(records[:start], start=1):
            identity = _event_identity(record)
            if identity is not None and identity not in seen:
                seen[identity] = line_number

        for line_number, record in enumerate(records[start:], start=start + 1):
            context = f"{path}:{line_number}"
            _require_nonempty_event_fields(record, context)
            if _clean_text(record["ticker"]) != company.ticker:
                raise ContractError(
                    f"新增 event ticker 与目录不匹配 {context}: {record['ticker']}"
                )
            if _clean_text(record["name"]) != company.name:
                raise ContractError(
                    f"新增 event name 与 watchlist 不匹配 {context}: {record['name']}"
                )
            identity = _event_identity(record)
            assert identity is not None
            if identity in seen:
                raise ContractError(
                    f"新增 event 身份重复 {context}; 首次出现于第 {seen[identity]} 行"
                )
            seen[identity] = line_number
            new_event_count += 1
    return new_event_count


def _first_difference(expected: Any, actual: Any, path: str = "root") -> str | None:
    if type(expected) is not type(actual):
        return f"{path}: type {type(expected).__name__} != {type(actual).__name__}"
    if isinstance(expected, dict):
        if expected.keys() != actual.keys():
            return f"{path}: keys {sorted(expected.keys())} != {sorted(actual.keys())}"
        for key in expected:
            difference = _first_difference(expected[key], actual[key], f"{path}.{key}")
            if difference:
                return difference
        return None
    if isinstance(expected, list):
        if len(expected) != len(actual):
            return f"{path}: length {len(expected)} != {len(actual)}"
        for index, (left, right) in enumerate(zip(expected, actual)):
            difference = _first_difference(left, right, f"{path}[{index}]")
            if difference:
                return difference
        return None
    if expected != actual:
        return f"{path}: {expected!r} != {actual!r}"
    return None


def validate_run(
    snapshot: dict[str, Any],
    watchlist: Path,
    events_root: Path,
    run_status: Path,
) -> dict[str, Any]:
    if snapshot.get("schema_version") != SNAPSHOT_SCHEMA_VERSION:
        raise ContractError(
            f"快照 schema_version 不支持: {snapshot.get('schema_version')}"
        )
    enabled, columns = read_enabled_companies(watchlist)
    expected_enabled = [
        EnabledCompany(
            ticker=_clean_text(item.get("ticker")),
            name=_clean_text(item.get("name")),
            row=int(item.get("row")),
        )
        for item in snapshot.get("enabled_companies", [])
    ]
    if enabled != expected_enabled:
        raise ContractError(
            "运行期间 enabled watchlist 发生变化: "
            f"expected={expected_enabled}, actual={enabled}"
        )

    mutable_cells = _mutable_cells(enabled, columns)
    if mutable_cells != set(snapshot.get("mutable_workbook_cells", [])):
        raise ContractError("快照中的 Excel 可变单元格合同与当前 watchlist 不一致")
    current_fingerprint = workbook_fingerprint(watchlist, mutable_cells)
    difference = _first_difference(snapshot.get("workbook_fingerprint"), current_fingerprint)
    if difference:
        raise ContractError(f"Excel 工作簿关键结构或非允许内容发生变化: {difference}")

    completion_rows = parse_completion_table(run_status)
    validate_completion_table(enabled, completion_rows)
    new_event_count = validate_events(
        enabled,
        events_root,
        snapshot.get("event_prefixes", {}),
    )
    return {
        "status": "passed",
        "enabled_company_count": len(enabled),
        "completion_table_count": len(completion_rows),
        "new_event_count": new_event_count,
        "workbook_round_trip": "passed",
        "event_append_only": "passed",
    }


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f"{path.name}.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="A 股公司跟踪 Excel、events.jsonl 与 completion table 端到端验证"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    snapshot_parser = subparsers.add_parser("snapshot", help="运行前保存验证快照")
    snapshot_parser.add_argument("--watchlist", required=True, type=Path)
    snapshot_parser.add_argument("--events-root", required=True, type=Path)
    snapshot_parser.add_argument("--output", required=True, type=Path)

    validate_parser = subparsers.add_parser("validate", help="运行后执行端到端验证")
    validate_parser.add_argument("--snapshot", required=True, type=Path)
    validate_parser.add_argument("--watchlist", required=True, type=Path)
    validate_parser.add_argument("--events-root", required=True, type=Path)
    validate_parser.add_argument("--run-status", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "snapshot":
            payload = create_snapshot(args.watchlist, args.events_root)
            _write_json(args.output, payload)
            result = {
                "status": "snapshot_created",
                "snapshot": str(args.output),
                "enabled_company_count": len(payload["enabled_companies"]),
            }
        else:
            snapshot = json.loads(args.snapshot.read_text(encoding="utf-8"))
            result = validate_run(
                snapshot=snapshot,
                watchlist=args.watchlist,
                events_root=args.events_root,
                run_status=args.run_status,
            )
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
        return 0
    except ContractError as exc:
        error = str(exc)
    except json.JSONDecodeError as exc:
        error = f"验证快照 JSON 无效: {exc}"
    except UnicodeError as exc:
        error = f"验证输入不是有效 UTF-8: {exc}"
    except OSError as exc:
        error = f"验证输入输出失败: {exc}"
    print(
        json.dumps(
            {"status": "failed", "error": error},
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 1


if __name__ == "__main__":
    sys.exit(main())
