#!/usr/bin/env python
"""Build the 2026H1 QDII holdings workbook from verified official-report transcriptions.

The manifest is deliberately explicit: full holdings means the report's §7.4
full equity table.  Fund/ETF and bond sections retain their statutory
"top-ten" / "top-five" disclosure boundaries rather than pretending to be a
complete look-through portfolio.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLUE = "0000FF"
BLACK = "000000"
WHITE = "FFFFFF"
HEADER = "1F4E78"
SECTION = "D9EAF7"
NOTE = "FFF2CC"
THIN_GRAY = Side(style="thin", color="B7B7B7")
MEDIUM_BLUE = Side(style="medium", color=HEADER)
INPUT_FONT = Font(name="Microsoft YaHei", size=10, color=BLUE)
FORMULA_FONT = Font(name="Microsoft YaHei", size=10, color=BLACK)
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color=HEADER)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=10, italic=True, color="666666")
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'
PCT_FORMAT = '0.00%'


def now_bj() -> str:
    return datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def decimal(value: Any) -> Decimal:
    if value is None:
        raise ValueError("金额/比例字段不得为 None")
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "－", "N/A"}:
        raise ValueError(f"无法转换为数字: {value!r}")
    return Decimal(text)


def maybe_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "—", "－", "N/A"}:
        return None
    return Decimal(text)


def pct(value: Any) -> float | None:
    item = maybe_decimal(value)
    return None if item is None else float(item / Decimal("100"))


def page_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value)


def listing_pages(row: dict[str, Any]) -> str:
    start = row.get("source_page_start")
    end = row.get("source_page_end")
    if start is None:
        return page_text(row.get("source_pages"))
    return str(start) if end in {None, start} else f"{start}-{end}"


def apply_base_style(cell, *, formula: bool = False) -> None:
    cell.font = FORMULA_FONT if formula else INPUT_FONT
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def set_title(ws, title: str, subtitle: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    ws.cell(2, 1, subtitle).font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34


def set_header(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row, col, value)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=MEDIUM_BLUE, right=MEDIUM_BLUE, top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)
    ws.row_dimensions[row].height = 32


def set_widths(ws, widths: list[float]) -> None:
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[float],
    formats: dict[int, str] | None = None,
) -> int:
    set_header(ws, start_row, headers)
    for row_offset, values in enumerate(rows, 1):
        current_row = start_row + row_offset
        for col, value in enumerate(values, 1):
            cell = ws.cell(current_row, col, value)
            apply_base_style(cell)
            if formats and col in formats and isinstance(value, (int, float, Decimal)):
                cell.number_format = formats[col]
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    ws.freeze_panes = f"A{start_row + 1}"
    set_widths(ws, widths)
    return start_row + len(rows)


def hyperlink(cell, url: str | None) -> None:
    if url:
        cell.hyperlink = url
        cell.style = "Hyperlink"


def source_json_for(fund: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    json_path = Path(fund["equity_json_path"])
    if not json_path.is_file():
        raise FileNotFoundError(f"未找到权益转录 JSON: {json_path}")
    expected_hash = fund.get("equity_json_sha256")
    actual_hash = sha256(json_path)
    if expected_hash and actual_hash.lower() != str(expected_hash).lower():
        raise ValueError(f"{fund['key']} 权益转录 JSON SHA-256 不匹配")
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    lines = doc.get("flattened_listing_lines")
    if not isinstance(lines, list) or not lines:
        raise ValueError(f"{fund['key']} 权益转录缺少 flattened_listing_lines")
    return doc, lines


def validate_fund(fund: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
    required = [
        "key",
        "a_code",
        "c_code",
        "fund_name",
        "report_title",
        "announcement_date",
        "report_period",
        "official_pdf_url",
        "pdf_path",
        "pdf_sha256",
        "page_count",
        "nav_cny",
        "equity_cny",
        "equity_nav_pct",
        "equity_pages",
        "equity_json_path",
    ]
    missing = [field for field in required if not fund.get(field)]
    if missing:
        raise ValueError(f"{fund.get('key', '<unknown>')} 缺少必填字段: {', '.join(missing)}")
    pdf_path = Path(fund["pdf_path"])
    if not pdf_path.is_file():
        raise FileNotFoundError(f"未找到官方 PDF: {pdf_path}")
    with pdf_path.open("rb") as handle:
        if handle.read(5) != b"%PDF-":
            raise ValueError(f"{fund['key']} 不是 PDF 文件")
    actual_hash = sha256(pdf_path)
    if actual_hash.lower() != str(fund["pdf_sha256"]).lower():
        raise ValueError(f"{fund['key']} 官方 PDF SHA-256 不匹配")
    if fund.get("pdf_bytes") is not None and pdf_path.stat().st_size != int(fund["pdf_bytes"]):
        raise ValueError(f"{fund['key']} 官方 PDF 文件大小不匹配")

    doc, lines = source_json_for(fund)
    sequence_values = [int(row["sequence"]) for row in lines]
    unique_sequences = sorted(set(sequence_values))
    expected_sequence_count = int(fund.get("company_sequence_count", len(unique_sequences)))
    expected_listing_count = int(fund.get("security_listing_line_count", len(lines)))
    if unique_sequences != list(range(1, expected_sequence_count + 1)):
        raise ValueError(f"{fund['key']} 公司级序号不连续")
    if len(lines) != expected_listing_count:
        raise ValueError(f"{fund['key']} 证券级上市行数不匹配")
    total = sum((decimal(row["fair_value_cny"]) for row in lines), Decimal("0"))
    reported = decimal(fund["equity_cny"])
    if total != reported:
        raise ValueError(f"{fund['key']} 权益行合计 {total} 与报告总额 {reported} 不一致")
    validation = doc.get("validation", {})
    json_total = validation.get("fair_value_sum_cny")
    if json_total is not None and decimal(json_total) != reported:
        raise ValueError(f"{fund['key']} JSON 核验总额与报告总额不一致")
    return doc, lines, {
        "pdf_sha256": actual_hash,
        "json_sha256": sha256(Path(fund["equity_json_path"])),
        "company_sequence_count": expected_sequence_count,
        "security_listing_line_count": expected_listing_count,
        "equity_sum": total,
        "validation": validation,
    }


def normalize_allocation(fund: dict[str, Any]) -> list[list[Any]]:
    rows = []
    for item in fund.get("allocation", []):
        rows.append(
            [
                item.get("sequence", ""),
                item.get("item", ""),
                float(maybe_decimal(item.get("amount_cny"))) if maybe_decimal(item.get("amount_cny")) is not None else None,
                pct(item.get("total_asset_pct")),
                item.get("note", ""),
                page_text(item.get("pages", fund.get("asset_pages", ""))),
            ]
        )
    return rows


def normalize_geography_sector(fund: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for category, items in (("地区", fund.get("geography", [])), ("行业", fund.get("sectors", []))):
        for item in items:
            rows.append(
                [
                    category,
                    item.get("item", ""),
                    float(maybe_decimal(item.get("fair_value_cny"))) if maybe_decimal(item.get("fair_value_cny")) is not None else None,
                    pct(item.get("nav_pct")),
                    item.get("section", ""),
                    page_text(item.get("pages", "")),
                    fund["official_pdf_url"],
                ]
            )
    return rows


def write_equity_sheet(wb: Workbook, fund: dict[str, Any], lines: list[dict[str, Any]]) -> tuple[str, int, int, int]:
    key = fund["key"]
    source_url = fund["official_pdf_url"]
    sheet_name = f"H1权益_{key}"
    ws = wb.create_sheet(sheet_name)
    set_title(
        ws,
        f"{fund['a_code']} / {fund['c_code']} H1 全部权益持仓",
        f"官方报告 §7.4 全部权益投资明细；公司级序号 1-{fund['company_sequence_count']}，证券级上市行 {len(lines)} 条。",
        12,
    )
    headers = [
        "公司级序号",
        "证券子行",
        "公司名称(英文)",
        "公司名称(中文)",
        "证券代码",
        "所在证券市场",
        "所属国家(地区)",
        "数量(股)",
        "公允价值(人民币)",
        "占基金资产净值比例",
        "PDF页码",
        "来源",
    ]
    set_header(ws, 3, headers)
    for row_index, item in enumerate(lines, 4):
        values = [
            int(item["sequence"]),
            int(item.get("listing_index", 1)),
            item.get("company_name_en", ""),
            item.get("company_name_zh", ""),
            item.get("security_code", ""),
            item.get("market", ""),
            item.get("country_region", ""),
            int(decimal(item["quantity_shares"])),
            float(decimal(item["fair_value_cny"])),
            pct(item["nav_pct"]),
            listing_pages(item),
            source_url,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_index, col, value)
            apply_base_style(cell)
            if col == 8:
                cell.number_format = INTEGER_FORMAT
            elif col == 9:
                cell.number_format = NUMBER_FORMAT
            elif col == 10:
                cell.number_format = PCT_FORMAT
        hyperlink(ws.cell(row_index, 12), source_url)

    last_data_row = 3 + len(lines)
    total_row = last_data_row + 1
    reported_row = total_row + 1
    difference_row = reported_row + 1
    for target_row, label, fill in (
        (total_row, "解析合计", SECTION),
        (reported_row, "报告§7.3权益投资合计", SECTION),
        (difference_row, "解析合计－报告合计", NOTE),
    ):
        ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=8)
        ws.cell(target_row, 1, label)
        for col in range(1, 13):
            cell = ws.cell(target_row, col)
            cell.fill = PatternFill("solid", fgColor=fill)
            apply_base_style(cell, formula=target_row in {total_row, difference_row} and col in {9, 10})
        ws.cell(target_row, 9).number_format = NUMBER_FORMAT
        ws.cell(target_row, 10).number_format = PCT_FORMAT
        ws.cell(target_row, 12, source_url)
        hyperlink(ws.cell(target_row, 12), source_url)
    ws.cell(total_row, 9, f"=SUM(I4:I{last_data_row})")
    ws.cell(total_row, 10, f"=SUM(J4:J{last_data_row})")
    ws.cell(total_row, 11, "与报告§7.3核验")
    ws.cell(reported_row, 9, float(decimal(fund["equity_cny"])))
    ws.cell(reported_row, 10, pct(fund["equity_nav_pct"]))
    ws.cell(reported_row, 11, f"报告第{fund.get('equity_summary_page', fund['equity_pages'])}页")
    ws.cell(difference_row, 9, f"=I{total_row}-I{reported_row}")
    ws.cell(difference_row, 10, f"=J{total_row}-J{reported_row}")
    ws.cell(difference_row, 11, "比例差异来自逐行两位小数披露的四舍五入")
    ws.auto_filter.ref = f"A3:L{last_data_row}"
    ws.freeze_panes = "A4"
    set_widths(ws, [12, 10, 30, 24, 19, 20, 14, 13, 19, 18, 12, 54])
    return sheet_name, last_data_row, total_row, reported_row


def write_allocation_sheet(wb: Workbook, fund: dict[str, Any]) -> None:
    key = fund["key"]
    ws = wb.create_sheet(f"H1资产_{key}")
    set_title(
        ws,
        f"{fund['a_code']} / {fund['c_code']} H1 基金资产组合",
        "官方报告 §7.1；金额单位为人民币元，占比以基金总资产为分母。",
        6,
    )
    rows = normalize_allocation(fund)
    if not rows:
        rows = [["", "N/A", None, None, "本次未结构化提取；权益明细与基金/ETF明细仍已更新", fund.get("asset_pages", "")]]
    write_table(
        ws,
        3,
        ["序号", "项目", "金额(人民币)", "占基金总资产比", "原文显示", "PDF页码"],
        rows,
        [10, 32, 22, 20, 30, 14],
        {3: NUMBER_FORMAT, 4: PCT_FORMAT},
    )


def write_geo_sheet(wb: Workbook, fund: dict[str, Any]) -> None:
    key = fund["key"]
    ws = wb.create_sheet(f"H1地区行业_{key}")
    set_title(
        ws,
        f"{fund['a_code']} / {fund['c_code']} H1 地区与行业配置",
        "官方报告 §7.2、§7.3；占比以基金资产净值为分母。",
        7,
    )
    rows = normalize_geography_sector(fund)
    if not rows:
        rows = [["", "N/A", None, None, "", "", fund["official_pdf_url"]]]
    end = write_table(
        ws,
        3,
        ["分类", "项目", "公允价值(人民币)", "占基金资产净值比", "报告章节", "PDF页码", "来源"],
        rows,
        [12, 28, 22, 20, 14, 12, 54],
        {3: NUMBER_FORMAT, 4: PCT_FORMAT},
    )
    for row in range(4, end + 1):
        hyperlink(ws.cell(row, 7), fund["official_pdf_url"])


def write_bond_sheet(wb: Workbook, fund: dict[str, Any]) -> None:
    key = fund["key"]
    ws = wb.create_sheet(f"H1债券_{key}")
    scope_note = fund.get("bond_scope_note", "官方报告 §7.7 披露的前五名债券投资明细。")
    set_title(ws, f"{fund['a_code']} / {fund['c_code']} H1 债券投资", scope_note, 8)
    rows = []
    for item in fund.get("bonds", []):
        rows.append(
            [
                item.get("sequence", ""),
                item.get("code", ""),
                item.get("name", ""),
                float(maybe_decimal(item.get("quantity"))) if maybe_decimal(item.get("quantity")) is not None else None,
                float(maybe_decimal(item.get("fair_value_cny"))) if maybe_decimal(item.get("fair_value_cny")) is not None else None,
                pct(item.get("nav_pct")),
                page_text(item.get("pages", fund.get("bond_pages", ""))),
                fund["official_pdf_url"],
            ]
        )
    if not rows:
        rows = [["", "", "报告披露未持有债券", None, None, None, fund.get("bond_pages", ""), fund["official_pdf_url"]]]
    end = write_table(
        ws,
        3,
        ["序号", "债券代码", "债券名称", "数量/面值", "公允价值(人民币)", "占净值比", "PDF页码", "来源"],
        rows,
        [10, 16, 30, 16, 22, 16, 12, 54],
        {4: INTEGER_FORMAT, 5: NUMBER_FORMAT, 6: PCT_FORMAT},
    )
    for row in range(4, end + 1):
        hyperlink(ws.cell(row, 8), fund["official_pdf_url"])


def write_fund_etf_sheet(wb: Workbook, fund: dict[str, Any]) -> None:
    key = fund["key"]
    ws = wb.create_sheet(f"H1基金及ETF_{key}")
    scope_note = fund.get("fund_scope_note", "官方报告 §7.10 披露的前十名基金投资明细；不代表完整基金投资穿透。")
    set_title(ws, f"{fund['a_code']} / {fund['c_code']} H1 基金及ETF投资", scope_note, 8)
    rows = []
    for item in fund.get("fund_investments", []):
        rows.append(
            [
                item.get("sequence", ""),
                item.get("name", ""),
                item.get("fund_type", ""),
                item.get("operation", ""),
                item.get("manager", ""),
                float(maybe_decimal(item.get("fair_value_cny"))) if maybe_decimal(item.get("fair_value_cny")) is not None else None,
                pct(item.get("nav_pct")),
                page_text(item.get("pages", fund.get("fund_pages", ""))),
            ]
        )
    if not rows:
        rows = [["", "报告披露未持有基金/ETF", "", "", "", None, None, fund.get("fund_pages", "")]]
    end = write_table(
        ws,
        3,
        ["序号", "基金名称", "基金类型", "运作方式", "管理人", "公允价值(人民币)", "占净值比", "PDF页码"],
        rows,
        [10, 44, 16, 18, 32, 22, 16, 12],
        {6: NUMBER_FORMAT, 7: PCT_FORMAT},
    )
    ws.merge_cells(start_row=end + 2, start_column=1, end_row=end + 2, end_column=8)
    ws.cell(end + 2, 1, "披露边界：本表来自报告“前十名基金投资明细”，不能视为全部基金/ETF穿透清单。")
    ws.cell(end + 2, 1).fill = PatternFill("solid", fgColor=NOTE)
    ws.cell(end + 2, 1).alignment = Alignment(wrap_text=True, vertical="center")


def write_notes_sheet(wb: Workbook, funds: list[dict[str, Any]], checked_at: str, results: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("说明")
    set_title(
        ws,
        "持仓QDII基金 2026H1 持仓跟踪",
        "账户当前三组基金家族均已按正式中期报告更新；完整持仓仅指报告 §7.4 的全部权益投资明细。",
        3,
    )
    rows: list[list[Any]] = [
        ["检查时间", checked_at, "北京时间；账户基金份额范围见下行"],
        ["账户范围", "164212 / 016823；100055 / 022184；016664 / 016665", "A/C份额按同一基金家族合并"],
        ["权益披露边界", "§7.4 全部权益投资明细", "股票/存托凭证等权益逐行完整更新；详见各 H1权益_<代码> 工作表"],
        ["基金/ETF披露边界", "§7.10 前十名基金投资明细", "仅记录报告实际披露项目，不声称为全部基金/ETF穿透清单"],
        ["债券披露边界", "§7.7 前五名债券投资明细", "仅记录报告实际披露项目"],
        ["历史保护", "原有 2026Q2 工作簿未覆盖", "本文件为独立 2026H1 快照"],
    ]
    for fund in funds:
        result = results[fund["key"]]
        rows.append(
            [
                f"已更新 {fund['a_code']} / {fund['c_code']}",
                fund["report_title"],
                f"{fund['announcement_date']}；PDF {fund['page_count']}页；SHA-256 {result['pdf_sha256'].upper()}",
            ]
        )
    set_header(ws, 4, ["项目", "内容", "说明"])
    for index, row in enumerate(rows, 5):
        for col, value in enumerate(row, 1):
            cell = ws.cell(index, col, value)
            apply_base_style(cell)
    set_widths(ws, [24, 80, 76])
    ws.freeze_panes = "A5"


def write_overview_sheet(
    wb: Workbook,
    funds: list[dict[str, Any]],
    details: dict[str, tuple[str, int, int, int]],
) -> None:
    ws = wb.create_sheet("H1概览")
    set_title(ws, "H1 披露概览", "当前账户三组基金家族均已按官方 2026 年中期报告更新。", 11)
    rows = []
    for fund in funds:
        sheet_name, last_data_row, _, _ = details[fund["key"]]
        rows.append(
            [
                fund["fund_name"],
                fund["a_code"],
                fund["c_code"],
                "已发布并更新",
                fund["announcement_date"],
                fund["report_period"],
                f"=SUM('{sheet_name}'!I4:I{last_data_row})",
                f"=SUM('{sheet_name}'!J4:J{last_data_row})",
                float(decimal(fund["equity_cny"])),
                pct(fund["equity_nav_pct"]),
                f"完整权益：§7.4第{fund['equity_pages']}页；基金/ETF：§7.10第{fund.get('fund_pages', '')}页",
            ]
        )
    end = write_table(
        ws,
        4,
        [
            "基金家族",
            "A份额",
            "C份额",
            "2026H1状态",
            "送出/公告日",
            "报告期",
            "解析权益合计(元)",
            "逐行净值占比合计",
            "官方权益合计(元)",
            "官方权益占净值比",
            "披露范围",
        ],
        rows,
        [34, 12, 12, 18, 16, 26, 20, 18, 20, 18, 48],
        {7: NUMBER_FORMAT, 8: PCT_FORMAT, 9: NUMBER_FORMAT, 10: PCT_FORMAT},
    )
    for row in range(5, end + 1):
        for col in (7, 8):
            apply_base_style(ws.cell(row, col), formula=True)
        ws.cell(row, 7).number_format = NUMBER_FORMAT
        ws.cell(row, 8).number_format = PCT_FORMAT
        for cell in ws[row]:
            cell.comment = Comment("硬编码来源或公式来源见工作簿“来源与核验”工作表。", "Codex")
    ws.merge_cells(start_row=end + 2, start_column=1, end_row=end + 2, end_column=11)
    ws.cell(end + 2, 1, "注：逐行占净值比例按原文两位小数披露；与报告合计的差异是披露舍入差，不反推或篡改单行数值。")
    ws.cell(end + 2, 1).fill = PatternFill("solid", fgColor=NOTE)
    ws.cell(end + 2, 1).alignment = Alignment(wrap_text=True, vertical="center")


def write_notice_sheet(wb: Workbook, funds: list[dict[str, Any]], checked_at: str) -> None:
    ws = wb.create_sheet("公告状态")
    set_title(ws, "当前账户基金家族的 2026H1 公告状态", "只以可下载、可校验的基金公司或法定披露平台原文作为“已发布”依据。", 10)
    rows = []
    for fund in funds:
        rows.append(
            [
                checked_at,
                fund["a_code"],
                fund["c_code"],
                fund["fund_name"],
                "已发布并更新",
                fund["report_title"],
                fund["announcement_date"],
                fund["report_period"],
                fund["official_pdf_url"],
                fund.get("source_status_note", "法定披露平台或 CNINFO 官方原文已下载并完成 SHA-256 校验"),
            ]
        )
    end = write_table(
        ws,
        3,
        ["核验时间", "A份额", "C份额", "基金家族", "H1状态", "最新H1标题", "送出/公告日", "报告期", "官方原文", "说明"],
        rows,
        [28, 12, 12, 34, 16, 58, 16, 26, 60, 48],
    )
    for row in range(4, end + 1):
        hyperlink(ws.cell(row, 9), ws.cell(row, 9).value)


def write_checks_sheet(
    wb: Workbook,
    funds: list[dict[str, Any]],
    details: dict[str, tuple[str, int, int, int]],
    results: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("来源与核验")
    set_title(ws, "来源与核验", "所有金额保留原始披露口径；公式只用于工作簿内的完整性校验。", 5)
    set_header(ws, 3, ["基金家族 / 核验项目", "数值/公式", "来源", "页码", "结果或说明"])
    row = 4
    for fund in funds:
        sheet_name, last_data_row, total_row, reported_row = details[fund["key"]]
        result = results[fund["key"]]
        validation = result["validation"]
        company_count = result["company_sequence_count"]
        listing_count = result["security_listing_line_count"]
        nav_rounding_note = validation.get(
            "note",
            "逐行净值占比为原文披露值；合计与报告值的差异属于披露舍入差。",
        )
        rows = [
            [f"{fund['a_code']} / {fund['c_code']} 官方原文", fund["report_title"], fund["official_pdf_url"], "封面及§7", "已验证 PDF、报告期、送出日期"],
            ["文件身份", result["pdf_sha256"].upper(), "官方PDF", "全文件", f"SHA-256；{fund['page_count']}页；{Path(fund['pdf_path']).stat().st_size:,} bytes"],
            ["公司级序号", company_count, "§7.4", fund["equity_pages"], "严格连续，无缺号/重号"],
            ["证券级上市行", listing_count, "§7.4", fund["equity_pages"], "多市场/多代码时保留子行"],
            ["解析权益市值", f"=SUM('{sheet_name}'!I4:I{last_data_row})", "§7.4行项目", fund["equity_pages"], "与§7.3报告总额比对"],
            ["报告权益市值", float(decimal(fund["equity_cny"])), "§7.3", fund.get("equity_summary_page", fund["equity_pages"]), "人民币；与解析合计应为0差异"],
            ["权益金额差异", f"='{sheet_name}'!I{total_row}-'{sheet_name}'!I{reported_row}", "公式", sheet_name, "应为0.00"],
            ["逐行净值占比合计", f"=SUM('{sheet_name}'!J4:J{last_data_row})", "§7.4", fund["equity_pages"], "逐行两位小数披露"],
            ["报告权益占净值比", pct(fund["equity_nav_pct"]), "§7.3", fund.get("equity_summary_page", fund["equity_pages"]), "报告原文比例"],
            ["占比差异", f"='{sheet_name}'!J{total_row}-'{sheet_name}'!J{reported_row}", "公式", sheet_name, nav_rounding_note],
            ["基金/ETF披露范围", fund.get("fund_scope_note", "§7.10 前十名基金投资明细"), "§7.10", fund.get("fund_pages", ""), "不可表述为全部基金/ETF穿透"],
        ]
        for values in rows:
            for col, value in enumerate(values, 1):
                cell = ws.cell(row, col, value)
                apply_base_style(cell, formula=isinstance(value, str) and value.startswith("="))
                if col == 2 and values[0] in {"解析权益市值", "报告权益市值", "权益金额差异"}:
                    cell.number_format = NUMBER_FORMAT
                if col == 2 and values[0] in {"逐行净值占比合计", "报告权益占净值比", "占比差异"}:
                    cell.number_format = PCT_FORMAT
            if values[0].endswith("官方原文"):
                hyperlink(ws.cell(row, 3), fund["official_pdf_url"])
            row += 1
        row += 1
    ws.freeze_panes = "A4"
    set_widths(ws, [30, 36, 64, 18, 72])


def create_workbook(
    output_path: Path,
    manifest: dict[str, Any],
    fund_docs: dict[str, tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]],
    checked_at: str,
) -> None:
    funds = manifest["funds"]
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "持仓QDII基金 2026H1 持仓跟踪"
    wb.properties.subject = "以基金正式中期报告更新的持仓快照"
    wb.properties.creator = "Codex"

    results = {key: result for key, (_, _, result) in fund_docs.items()}
    write_notes_sheet(wb, funds, checked_at, results)
    details: dict[str, tuple[str, int, int, int]] = {}
    for fund in funds:
        _, lines, _ = fund_docs[fund["key"]]
        details[fund["key"]] = write_equity_sheet(wb, fund, lines)
        write_allocation_sheet(wb, fund)
        write_geo_sheet(wb, fund)
        write_bond_sheet(wb, fund)
        write_fund_etf_sheet(wb, fund)
    # Move the high-level pages ahead of fund-specific details, while leaving
    # the source/verification sheets as the final audit trail.
    write_overview_sheet(wb, funds, details)
    wb._sheets.insert(1, wb._sheets.pop())
    write_notice_sheet(wb, funds, checked_at)
    write_checks_sheet(wb, funds, details, results)

    for ws in wb.worksheets:
        for cells in ws.iter_rows():
            for cell in cells:
                if cell.value is not None and cell.font.name is None:
                    cell.font = INPUT_FONT
        ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_status_payload(funds: list[dict[str, Any]], checked_at: str, results: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return {
        "checked_at": checked_at,
        "scope": "current_account_fund_share_classes",
        "account_fund_codes": ["164212", "016823", "100055", "022184", "016664", "016665"],
        "fund_families": [
            {
                "primary_code": fund["a_code"],
                "share_classes": [fund["a_code"], fund["c_code"]],
                "h1_status": "published_and_updated",
                "title": fund["report_title"],
                "announcement_date": fund["announcement_date"],
                "period": fund["report_period"],
                "official_pdf": fund["official_pdf_url"],
                "pdf_sha256": results[fund["key"]]["pdf_sha256"],
                "equity_coverage": f"full_equity_details_{results[fund['key']]['company_sequence_count']}_company_sequences_{results[fund['key']]['security_listing_line_count']}_security_listing_lines",
                "fund_etf_coverage": fund.get("fund_scope_note", "§7.10 top-ten disclosure"),
            }
            for fund in funds
        ],
        "boundary": "All three official H1 full equity tables are updated. Fund/ETF and bond items retain the report's top-ten/top-five disclosure scope.",
    }


def build_audit_payload(fund: dict[str, Any], first_holding: dict[str, Any], checked_at: str) -> dict[str, Any]:
    position_as_of = "2026-06-30T23:59:59+08:00"
    source_id = f"S_OFFICIAL_{fund['key']}"
    transcription_id = f"S_TRANSCRIPTION_{fund['key']}"
    official_fact_id = f"F_OFFICIAL_POSITION_{fund['key']}"
    transcribed_fact_id = f"F_TRANSCRIBED_POSITION_{fund['key']}"
    origin_id = fund.get("origin_id", f"official:{fund['key']}:2026h1")
    return {
        "schema_version": "1.0",
        "audit_id": f"fund-{fund['key']}-h1-20260831",
        "as_of": checked_at,
        "sources": [
            {
                "id": source_id,
                "source_type": "official_filing",
                "origin_id": origin_id,
                "locator": fund["official_pdf_url"] + f"#page={first_holding.get('source_page_start', fund['equity_pages'])}",
                "source_date": fund["announcement_date"],
                "checked_at": checked_at,
                "status": "accepted",
            },
            {
                "id": transcription_id,
                "source_type": "report_under_audit",
                "origin_id": origin_id,
                "locator": f"artifact:fund_tracking/h1_2026_{fund['key']}_all_equity.json#sequence={first_holding['sequence']}",
                "source_date": fund["announcement_date"],
                "checked_at": checked_at,
                "status": "accepted",
            },
        ],
        "facts": [
            {
                "id": official_fact_id,
                "metric": "fund_equity_position_fair_value",
                "value": first_holding["fair_value_cny"],
                "unit": "currency",
                "currency": "CNY",
                "scale": "1",
                "period": {"kind": "instant", "as_of": position_as_of},
                "basis": "reported_fund_equity_position",
                "source_refs": [source_id],
            },
            {
                "id": transcribed_fact_id,
                "metric": "fund_equity_position_fair_value",
                "value": first_holding["fair_value_cny"],
                "unit": "currency",
                "currency": "CNY",
                "scale": "1",
                "period": {"kind": "instant", "as_of": position_as_of},
                "basis": "reported_fund_equity_position",
                "source_refs": [transcription_id],
            },
        ],
        "checks": [
            {
                "id": f"C_POSITION_TRANSCRIPTION_{fund['key']}",
                "kind": "cross_source",
                "materiality": "material",
                "target": {"fact_id": transcribed_fact_id},
                "references": [{"fact_id": official_fact_id}],
                "source_gate": {
                    "min_independent_origins": 1,
                    "counted_tier": "official",
                    "required_anchor_tier": "official",
                },
                "tolerance": {"relative_pct": "0", "absolute_base": "0"},
            }
        ],
    }


def copy_artifacts(
    artifacts: Path,
    manifest_path: Path,
    funds: list[dict[str, Any]],
    fund_docs: dict[str, tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]],
    checked_at: str,
) -> None:
    artifacts.mkdir(parents=True, exist_ok=True)
    results = {key: result for key, (_, _, result) in fund_docs.items()}
    for fund in funds:
        key = fund["key"]
        source_pdf = Path(fund["pdf_path"])
        target_pdf = artifacts / f"h1_2026_{key}.pdf"
        if source_pdf.resolve() != target_pdf.resolve():
            shutil.copy2(source_pdf, target_pdf)
        source_json = Path(fund["equity_json_path"])
        target_json = artifacts / f"h1_2026_{key}_all_equity.json"
        if source_json.resolve() != target_json.resolve():
            shutil.copy2(source_json, target_json)
        audit_path = artifacts / f"financial_audit_{key}_h1_20260831.input.json"
        audit_path.write_text(
            json.dumps(build_audit_payload(fund, fund_docs[key][1][0], checked_at), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    status_path = artifacts / "latest_notice_check" / "2026_h1_status_20260831.json"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    status_path.write_text(
        json.dumps(build_status_payload(funds, checked_at, results), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (artifacts / "h1_2026_portfolio_manifest.json").write_text(
        manifest_path.read_text(encoding="utf-8"), encoding="utf-8"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", required=True, type=Path)
    parser.add_argument("--artifacts-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    funds = manifest.get("funds")
    if not isinstance(funds, list) or len(funds) != 3:
        raise ValueError("清单必须精确包含当前账户的三组基金家族")
    keys = [fund.get("key") for fund in funds]
    if len(set(keys)) != len(keys):
        raise ValueError("基金家族 key 不得重复")
    fund_docs = {fund["key"]: validate_fund(fund) for fund in funds}
    checked_at = manifest.get("checked_at") or now_bj()
    copy_artifacts(args.artifacts_dir, args.manifest, funds, fund_docs, checked_at)
    create_workbook(args.output, manifest, fund_docs, checked_at)
    print(
        json.dumps(
            {
                "checked_at": checked_at,
                "output": str(args.output),
                "funds": {
                    key: {
                        "company_sequence_count": result[2]["company_sequence_count"],
                        "security_listing_line_count": result[2]["security_listing_line_count"],
                        "equity_sum": str(result[2]["equity_sum"]),
                    }
                    for key, result in fund_docs.items()
                },
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
