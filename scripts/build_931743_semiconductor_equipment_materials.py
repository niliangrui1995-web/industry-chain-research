"""构建中证半导体材料设备主题指数（931743）半年报跟踪表。

数据边界：
* 东方财富：成分行情快照、财务候选数值和业绩预告范围；
* 巨潮资讯：预约披露、报告/预告原文和 PDF 链接；
* 中证指数：成分与权重快照；
* 产品场景：公司 2025 年年报直接表述，保留页面与证据边界。

本脚本不把业绩预告当成已实现利润。预告公司的 Q2、增速与 PE 均以区间
展示，并显式标为“预告推导”。“用户定义 PE”始终是：
总市值 / （Q2 单季扣非归母净利润 × 4），不是 TTM PE。
"""

from __future__ import annotations

import argparse
import http.client
import io
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover - bundled runtime normally contains pypdf
    PdfReader = None


CN_TZ = ZoneInfo("Asia/Shanghai")
EASTMONEY_COMPONENT_URL = (
    "https://push2.eastmoney.com/api/qt/slist/get?"
    "secid=2.931743&fltt=1&fields=f1,f2,f3,f4,f12,f13,f14,f152,f20,f21&"
    "invt=2&fid=f3&pz=100&po=1&np=1&spt=11&pn=1&"
    "ut=f057cbcbce2a86e2866ab8877db1d059"
)
EASTMONEY_FINANCE_ENDPOINT = "https://datacenter-web.eastmoney.com/api/data/v1/get"
CNINFO_PRBOOK_ENDPOINT = "https://www.cninfo.com.cn/new/information/getPrbookInfo"
CNINFO_ANNOUNCEMENT_ENDPOINT = "https://www.cninfo.com.cn/new/hisAnnouncement/query"
CNINFO_STATIC_PREFIX = "https://static.cninfo.com.cn/"
CSINDEX_COMPONENT_URL = (
    "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/file/"
    "autofile/cons/931743cons.xls"
)
CSINDEX_WEIGHT_URL = (
    "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/file/"
    "autofile/closeweight/931743closeweight.xls?20260810134526"
)


# 中证官方 2026-07-31 收盘权重快照；用作标准名称与指数权重交叉核验。
WEIGHT_ROWS = """
002119\t康强电子\t0.387
002371\t北方华创\t15.267
002409\t雅克科技\t2.413
300236\t上海新阳\t1.169
300346\t南大光电\t2.111
300604\t长川科技\t7.281
300655\t晶瑞电材\t0.837
300666\t江丰电子\t2.726
301095\t广立微\t0.501
301369\t联动科技\t0.292
301611\t珂玛科技\t0.932
301629\t矽电股份\t0.297
600206\t有研新材\t1.324
603061\t金海通\t0.868
603078\t江化微\t0.409
603690\t至纯科技\t0.420
605358\t立昂微\t1.388
688012\t中微公司\t16.282
688019\t安集科技\t2.728
688037\t芯源微\t2.988
688072\t拓荆科技\t8.309
688082\t盛美上海\t2.486
688120\t华海清科\t6.422
688126\t沪硅产业\t2.450
688146\t中船特气\t2.668
688200\t华峰测控\t3.171
688233\t神工股份\t0.603
688234\t天岳先进\t1.176
688361\t中科飞测\t5.365
688409\t富创精密\t1.914
688432\t有研硅\t0.691
688478\t晶升股份\t0.178
688535\t华海诚科\t0.479
688545\t兴福电子\t0.757
688549\t中巨芯\t0.729
688584\t上海合晶\t0.302
688605\t先锋精科\t0.429
688652\t京仪装备\t0.817
688720\t艾森股份\t0.290
688721\t龙图光罩\t0.147
""".strip()


# 只保留年报直接支持的“产品—工艺/器件应用”表述；N/A 是明确的证据缺口，
# 不是以指数标签或行业新闻补出来的结论。
APPLICATION_ROWS = """
688082\tB.TEBO 单片兆声波清洗\t集成电路图形晶圆（≤28nm、2D/3D 结构）清洗\thttps://static.cninfo.com.cn/finalpage/2026-02-27/1224986438.PDF\t2025年报 p14\t中\t仅覆盖 B.TEBO 产品，不能外推为公司全部产品
600206\tN/A\tN/A（未在本轮年报定位到产品—应用直接表述）\thttps://static.cninfo.com.cn/finalpage/2026-04-25/1225177379.PDF\t2025年报 p11-12\t低\t集团多业务概述；指数成分不能证明全部业务均属半导体材料
688409\t精密零部件、特种涂层及翻新\t半导体前道刻蚀、沉积、清洗、光刻等设备\thttps://static.cninfo.com.cn/finalpage/2026-04-30/1225265029.PDF\t2025年报 p19\t高\t设备零部件场景，不写成芯片终端应用
688652\t温控、工艺废气处理、晶圆传片/定位/排序设备\t晶圆制造产线辅助工艺\thttps://static.cninfo.com.cn/finalpage/2026-03-31/1225056304.PDF\t2025年报 p6\t高\t限晶圆制造辅助设备
688432\t半导体硅材料、刻蚀设备硅材料/零部件\t集成电路、分立/功率器件、传感器、光学器件制造\thttps://static.cninfo.com.cn/finalpage/2026-03-27/1225037327.PDF\t2025年报 p11\t高\t新能源车/航天仅可作广泛应用背景，不代表客户或收入
688012\t刻蚀、MOCVD 等半导体设备\t逻辑、DRAM、3D NAND 制造；先进封装/MEMS；LED 外延\thttps://static.cninfo.com.cn/finalpage/2026-03-31/1225062431.PDF\t2025年报 p43-46\t高\t7nm以下、Micro-LED、GaN/SiC 的阶段/量产状态须逐项判断
603690\t湿法清洗、湿法刻蚀设备\t逻辑、高密度存储、先进封装、化合物半导体的晶圆前道\thttps://static.cninfo.com.cn/finalpage/2026-04-30/1225267363.PDF\t2025年报 p17\t高\t节点和量产状态应逐产品写
688072\t半导体薄膜沉积及三维集成设备\tN/A（HBM/Chiplet、3D存储、背照式CIS等细分场景待补公司产品直接证据）\thttps://static.cninfo.com.cn/finalpage/2026-04-28/1225222226.PDF\t2025年报 p8\t中\t该页为混合键合技术定义，不能据此断言公司设备已覆盖全部细分场景
688545\t湿电子化学品\t集成电路、TFT-LCD 制造的清洗/蚀刻；通用湿电子化学品亦用于光伏\thttps://static.cninfo.com.cn/finalpage/2026-03-31/1225059665.PDF\t2025年报 p6\t高\t光伏/面板业务不可混同为半导体收入
002409\tN/A\tN/A（待补电子材料分部产品页/官网原文）\thttps://static.cninfo.com.cn/finalpage/2026-04-28/1225216877.PDF\t2025年报扫描\t低\t本轮未定位电子材料产品—应用直接句，不能以概念标签填充
301611\t先进陶瓷零部件\t晶圆前道设备腔室内应用\thttps://static.cninfo.com.cn/finalpage/2026-04-28/1225216551.PDF\t2025年报 p13\t高\t新能源是另一应用，不并入半导体场景
688037\t前/后段清洗设备\t晶圆沉积前、刻蚀/离子注入/CMP后清洗；2.5D/3D封装 Frame 晶圆清洗\thttps://static.cninfo.com.cn/finalpage/2026-04-18/1225120170.PDF\t2025年报 p5,p13\t高\t不含客户或出货结论
300666\t超高纯金属溅射靶材、精密零部件\t芯片/面板产线 PVD；PVD、CVD、刻蚀、离子注入设备零部件/耗材\thttps://static.cninfo.com.cn/finalpage/2026-04-16/1225108970.PDF\t2025年报 p11\t高\t靶材与零部件两条业务分开呈现
688605\t精密金属工艺/结构部件\t半导体刻蚀、薄膜沉积设备的腔体、内衬、加热器等\thttps://static.cninfo.com.cn/finalpage/2026-04-10/1225089537.PDF\t2025年报 p14-15\t高\t仅设备零部件层
688200\t集成电路测试系统\t模拟/数模混合、分立器件及功率模块；GaN/SiC 等功率器件测试\thttps://static.cninfo.com.cn/finalpage/2026-04-29/1225240540.PDF\t2025年报 p11\t高\t下游新能源/通信仅作器件下游背景
688361\t检测量测设备\t硅片出厂/入厂质控、晶圆制程污染监控、图形晶圆二维/三维缺陷检测\thttps://static.cninfo.com.cn/finalpage/2026-04-25/1225182753.PDF\t2025年报 p15-16\t高\t不推断具体芯片类型
688146\t电子特种气体\t集成电路/显示面板制造的光刻、刻蚀、成膜、清洗、掺杂、沉积工艺\thttps://static.cninfo.com.cn/finalpage/2026-04-21/1225127119.PDF\t2025年报 p6\t高\t电子特气是工艺材料，非终端产品
688233\t硅电极及相关硅零部件\t存储芯片厂等离子刻蚀耗材；12英寸晶圆制造线\thttps://static.cninfo.com.cn/finalpage/2026-03-21/1225023708.PDF\t2025年报 p11,p21\t高\t不推断客户名称
688721\t光罩\t功率、模拟、逻辑芯片；中低精度半导体制造/封装、光学器件、触控屏及 PCB\thttps://static.cninfo.com.cn/finalpage/2026-03-27/1225038344.PDF\t2025年报 p13-14\t高\t按光罩基材/精度区别应用
300346\tMO 源、电子特气、ArF 光刻胶\tLED；集成电路安全源及 ArF 光刻胶；分立器件精细图形加工\thttps://static.cninfo.com.cn/finalpage/2026-04-10/1225087469.PDF\t2025年报 p7\t高\t气体其他下游不能自动归为半导体业务
300604\t测试机、分选机、老化设备\tSoC 高可靠性老化（车规、AI算法芯片）；MOS/IGBT/SiC；GPU、服务器CPU、AI芯片测试分选\thttps://static.cninfo.com.cn/finalpage/2026-04-25/1225193261.PDF\t2025年报 p11\t中高\t适用对象不等于全部收入，型号/量产须另列
300236\t半导体光刻胶/CMP 相关材料\t逻辑、模拟、存储芯片制造光刻；STI/钨/铜/氧化层/多晶硅 CMP\thttps://static.cninfo.com.cn/finalpage/2026-03-13/1225006867.PDF\t2025年报 p13\t高\t排除非半导体功能涂料业务
688120\tCMP、减薄等装备/耗材\t集成电路、先进封装/3D IC、大硅片、第三代半导体、MEMS、MicroLED 制造\thttps://static.cninfo.com.cn/finalpage/2026-04-23/1225156951.PDF\t2025年报 p12-13\t高\t工艺设备应用，不代表全部终端收入
688549\t电子湿化学品、电子特气\tIC、平板显示、光伏的清洗/刻蚀；IC/面板/LED/光纤/光伏的清洗、刻蚀、掺杂、沉积\thttps://static.cninfo.com.cn/finalpage/2026-04-22/1225139979.PDF\t2025年报 p5-6\t高\t多下游并列，不应按指数只保留半导体而遗漏边界
688234\tN/A\tN/A（待补碳化硅衬底产品—下游/器件直接证据）\thttps://static.cninfo.com.cn/finalpage/2026-03-28/1225049103.PDF\t2025年报扫描\t低\t不能以新能源车、光伏等概念替代直接证据
688720\t电镀液\t传统/先进封装电镀；28nm及5–14nm晶圆先进制程电镀\thttps://static.cninfo.com.cn/finalpage/2026-04-25/1225196094.PDF\t2025年报 p12\t中高\t限对应电镀液产品线
002371\t刻蚀、薄膜沉积、热处理、清洗等装备\t集成电路、功率/化合物半导体、3D集成/先进封装、新型显示制造\thttps://static.cninfo.com.cn/finalpage/2026-04-18/1225122918.PDF\t2025年报 p14,p19-20\t高\tAI/汽车为行业合作或需求表述，不能写成订单/收入
603061\t测试分选机\t集成电路后道封装测试\thttps://static.cninfo.com.cn/finalpage/2026-03-11/1225004246.PDF\t2025年报 p11-12\t高\t芯片类别由客户经营决定，不写汽车/5G/AI终端类别
688019\tCMP 抛光液、清洗液、刻蚀后清洗液\t逻辑、3D NAND、DRAM、CIS 制造；2.5D、3D TSV、混合键合先进封装\thttps://static.cninfo.com.cn/finalpage/2026-04-16/1225108553.PDF\t2025年报 p16-18,p28-29\t高\t不同产品量产/验证状态须分开；官网产品页用于交叉
688584\t硅外延片\tMOSFET/IGBT 等功率器件、PMIC/CIS 模拟芯片；CIS及车载充电/充电桩功率器件\thttps://static.cninfo.com.cn/finalpage/2026-04-04/1225079952.PDF\t2025年报 p11-12\t高\t器件/应用场景，不写成终端客户
301629\t探针台\t晶圆 CP/WAT、设计验证、成品测试（FT）\thttps://static.cninfo.com.cn/finalpage/2026-04-22/1225138756.PDF\t2025年报 p11\t高\t不外推具体终端芯片
688535\t环氧塑封料、芯片粘结材料\t传统/先进半导体封装、芯片级塑封、板级组装\thttps://static.cninfo.com.cn/finalpage/2026-03-18/1225014871.PDF\t2025年报 p15\t高\t封装材料用途，非下游整机用途
688478\t单晶硅炉\t8–12英寸半导体硅片制造；碳化硅单晶生长\thttps://static.cninfo.com.cn/finalpage/2026-04-30/1225262156.PDF\t2025年报 p16\t高\t不据此推断功率器件出货
603078\t湿电子化学品、光刻胶配套试剂\t芯片、显示面板、电池板制造的清洗、刻蚀、显影、去膜、掺杂\thttps://static.cninfo.com.cn/finalpage/2026-04-25/1225176561.PDF\t2025年报 p4,p8\t高\t电池/面板为并列下游，不等于半导体收入
301369\t半导体测试系统、激光打标设备\t功率半导体、模拟/数模混合IC；IGBT/SiC/GaN、晶圆、车规KGD、功率模块；芯片激光打标\thttps://static.cninfo.com.cn/finalpage/2026-03-31/1225054584.PDF\t2025年报 p10\t高\t“批量销售”仅对应 QT-8400，不能泛化
002119\t引线框架、键合丝\tIC、LED、功率电子和分立器件封装；晶体管/集成电路/大规模集成电路封装\thttps://static.cninfo.com.cn/finalpage/2026-03-31/1225052706.PDF\t2025年报 p11-12\t高\t封装材料层
301095\tEDA 软件及晶圆级电性测试相关产品\t集成电路设计、制造、封测的数据分析/工艺良率管理\thttps://static.cninfo.com.cn/finalpage/2026-04-23/1225153967.PDF\t2025年报 p19\t中\t具体产品与量产范围待补官网产品页
300655\t湿电子化学品、光刻胶\t半导体、显示面板、LED 的光刻/显影/刻蚀/清洗/去膜\thttps://static.cninfo.com.cn/finalpage/2026-04-28/1225199721.PDF\t2025年报 p16\t高\t锂电材料另列，不能混用
605358\t硅抛光片、外延片\t逻辑、存储、功率、模拟、传感器芯片；8英寸功率/PMIC/汽车电子，12英寸先进制程/高端功率\thttps://static.cninfo.com.cn/finalpage/2026-04-28/1225214695.PDF\t2025年报 p11-12\t高\t限硅片/外延片业务，勿混同器件业务
688126\t抛光片、外延片、SOI 等硅片\t存储、逻辑、硅光、图像处理、通用处理器、功率、射频、模拟、分立芯片制造\thttps://static.cninfo.com.cn/finalpage/2026-04-17/1225114274.PDF\t2025年报 p15\t高\t硅片应用，不作客户/订单推断
""".strip()


@dataclass(frozen=True)
class Company:
    code: str
    name: str
    weight: Decimal

    @property
    def market(self) -> str:
        return "SH" if self.code.startswith(("6", "688")) else "SZ"

    @property
    def ticker(self) -> str:
        return f"{self.code}.{self.market}"


def parse_weights() -> dict[str, Company]:
    result: dict[str, Company] = {}
    for line in WEIGHT_ROWS.splitlines():
        code, name, weight = line.split("\t")
        result[code] = Company(code, name, Decimal(weight))
    if len(result) != 40:
        raise RuntimeError(f"中证标准成分数异常：{len(result)}")
    return result


def parse_applications() -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    headings = ("product", "scenario", "url", "locator", "confidence", "boundary")
    for line in APPLICATION_ROWS.splitlines():
        parts = line.split("\t")
        if len(parts) != 7:
            raise RuntimeError(f"应用场景行列数异常：{line}")
        code, *values = parts
        result[code] = dict(zip(headings, values, strict=True))
    if len(result) != 40:
        raise RuntimeError(f"应用场景覆盖数异常：{len(result)}")
    return result


def now_iso() -> str:
    return datetime.now(CN_TZ).replace(microsecond=0).isoformat()


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    return Decimal(str(value))


def float_or_none(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    raise TypeError(f"无法 JSON 序列化：{type(value)!r}")


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, default=json_default),
        encoding="utf-8",
    )


def request_json(
    url: str,
    *,
    params: dict[str, Any] | None = None,
    form: dict[str, Any] | None = None,
    referer: str,
    timeout: int = 30,
    retries: int = 3,
) -> tuple[dict[str, Any] | None, str | None]:
    if params:
        url = f"{url}{'&' if '?' in url else '?'}{urllib.parse.urlencode(params)}"
    data = urllib.parse.urlencode(form).encode("utf-8") if form is not None else None
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
        "Referer": referer,
        "X-Requested-With": "XMLHttpRequest",
    }
    last_error = ""
    for attempt in range(retries):
        request = urllib.request.Request(url, data=data, headers=headers)
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                content = response.read().decode("utf-8")
            return json.loads(content, parse_float=Decimal), None
        except (
            urllib.error.URLError,
            urllib.error.HTTPError,
            http.client.IncompleteRead,
            OSError,
            TimeoutError,
            json.JSONDecodeError,
        ) as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt + 1 < retries:
                time.sleep(0.8 * (attempt + 1))
    return None, last_error


def request_bytes(url: str, *, referer: str, timeout: int = 60) -> tuple[bytes | None, str | None]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": referer,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read(), None
    except (urllib.error.URLError, urllib.error.HTTPError, http.client.IncompleteRead, OSError, TimeoutError) as exc:
        return None, f"{type(exc).__name__}: {exc}"


def ms_to_date(value: Any) -> str:
    if not value:
        return ""
    try:
        return datetime.fromtimestamp(int(value) / 1000, tz=timezone.utc).astimezone(CN_TZ).date().isoformat()
    except (TypeError, ValueError, OSError):
        return ""


def cninfo_url(item: dict[str, Any]) -> str:
    adjunct = str(item.get("adjunctUrl") or "").lstrip("/")
    return f"{CNINFO_STATIC_PREFIX}{adjunct}" if adjunct else ""


def announcement_items(
    company: Company, start_date: str, end_date: str, page_num: int = 1
) -> tuple[list[dict[str, Any]], str | None]:
    payload, error = request_json(
        CNINFO_ANNOUNCEMENT_ENDPOINT,
        form={
            "pageNum": str(page_num),
            # 巨潮对该接口单页实际限制为 30；显式使用 30 才能可靠翻页。
            "pageSize": "30",
            "tabName": "fulltext",
            "column": "sse" if company.market == "SH" else "szse",
            "stock": "",
            "searchkey": company.name,
            "secid": "",
            "plate": "",
            "category": "",
            "trade": "",
            "seDate": f"{start_date}~{end_date}",
            "sortName": "",
            "sortType": "",
            "isHLtitle": "true",
        },
        referer="https://www.cninfo.com.cn/new/commonUrl?url=disclosure/listed/bulletinDetail",
    )
    if payload is None:
        return [], error
    items = [
        item
        for item in (payload.get("announcements") or [])
        if str(item.get("secCode") or "") == company.code
    ]
    return items, None


def choose_announcement(
    items: list[dict[str, Any]], title_token: str
) -> dict[str, Any] | None:
    matches = [
        item
        for item in items
        if title_token in str(item.get("announcementTitle") or "")
        and "摘要" not in str(item.get("announcementTitle") or "")
    ]
    if not matches:
        return None
    matches.sort(key=lambda item: int(item.get("announcementTime") or 0), reverse=True)
    return matches[0]


def choose_h1_forecast(items: list[dict[str, Any]]) -> dict[str, Any] | None:
    """兼容“业绩预告”“业绩预增/预亏”等交易所标题写法。"""
    matches = []
    for item in items:
        title = str(item.get("announcementTitle") or "")
        if "2026年半年度" not in title:
            continue
        if any(token in title for token in ("业绩预告", "业绩预增", "业绩预亏", "业绩预减", "业绩快报")):
            matches.append(item)
    if not matches:
        return None
    matches.sort(key=lambda item: int(item.get("announcementTime") or 0), reverse=True)
    return matches[0]


def get_prbook(company: Company) -> tuple[str, dict[str, Any] | None, str | None]:
    payload, error = request_json(
        CNINFO_PRBOOK_ENDPOINT,
        form={
            "sectionTime": "2026-06-30",
            "market": "szsh",
            "stockCode": company.code,
            "pagesize": "1000",
            "pagenum": "1",
        },
        referer="https://www.cninfo.com.cn/new/commonUrl?url=data/yypl",
    )
    if payload is None:
        return company.code, None, error
    entries = payload.get("prbookinfos") or []
    entry = entries[0] if entries else None
    return company.code, entry, None


def get_status_announcements(
    company: Company, cutoff: str
) -> tuple[str, list[dict[str, Any]], str | None]:
    items, error = announcement_items(company, "2026-06-01", cutoff)
    return company.code, items, error


def get_period_announcement(
    company: Company, period_key: str
) -> tuple[str, str, dict[str, Any] | None, str | None]:
    windows = {
        "q1_2025": ("2025-04-01", "2025-05-20", ("2025年第一季度报告", "2025年一季度报告")),
        "h1_2025": ("2025-07-01", "2025-09-10", "2025年半年度报告"),
        "q1_2026": ("2026-04-01", "2026-05-20", ("2026年第一季度报告", "2026年一季度报告")),
        "h1_2026": ("2026-07-01", "2026-08-10", "2026年半年度报告"),
    }
    start, end, tokens = windows[period_key]
    tokens = (tokens,) if isinstance(tokens, str) else tokens
    last_error: str | None = None
    # 巨潮单页有时最多只给 30 条；在报告期窗口内向后翻页，直到命中目标定期报告。
    for page_num in range(1, 7):
        items, error = announcement_items(company, start, end, page_num)
        if error:
            last_error = error
            continue
        for token in tokens:
            selected = choose_announcement(items, token)
            if selected:
                return company.code, period_key, selected, None
        if len(items) < 30:
            break
    return company.code, period_key, None, last_error


def fetch_component_snapshot() -> dict[str, Any]:
    payload, error = request_json(
        EASTMONEY_COMPONENT_URL,
        referer="https://quote.eastmoney.com/zz/2.931743.html",
    )
    if payload is None:
        raise RuntimeError(f"东方财富指数成分接口失败：{error}")
    result = payload.get("data") or payload
    diff = result.get("diff") or []
    total = int(result.get("total") or len(diff))
    if int(payload.get("rc") or 0) != 0 or total != 40 or len(diff) != 40:
        raise RuntimeError(
            f"东方财富指数成分返回异常：rc={payload.get('rc')} total={total} diff={len(diff)}"
        )
    return payload


def fetch_financials(codes: list[str]) -> tuple[list[dict[str, Any]], str | None, str]:
    quoted_codes = ",".join(f'"{code}"' for code in codes)
    params = {
        "reportName": "RPT_F10_FINANCE_GINCOME",
        "columns": (
            "SECURITY_CODE,SECURITY_NAME_ABBR,REPORT_DATE,REPORT_TYPE,NOTICE_DATE,"
            "DEDUCT_PARENT_NETPROFIT,PARENT_NETPROFIT"
        ),
        "filter": f"(SECURITY_CODE in ({quoted_codes}))",
        "pageNumber": "1",
        "pageSize": "5000",
        "sortTypes": "-1",
        "sortColumns": "REPORT_DATE",
        "source": "WEB",
        "client": "WEB",
    }
    payload, error = request_json(
        EASTMONEY_FINANCE_ENDPOINT,
        params=params,
        referer="https://data.eastmoney.com/",
    )
    endpoint = f"{EASTMONEY_FINANCE_ENDPOINT}?{urllib.parse.urlencode(params)}"
    if payload is None:
        return [], error, endpoint
    return list((payload.get("result") or {}).get("data") or []), None, endpoint


def fetch_predictions(codes: list[str]) -> tuple[list[dict[str, Any]], str | None, str]:
    quoted_codes = ",".join(f'"{code}"' for code in codes)
    params = {
        "reportName": "RPT_PUBLIC_OP_NEWPREDICT",
        "columns": (
            "SECURITY_CODE,SECURITY_NAME_ABBR,NOTICE_DATE,REPORT_DATE,PREDICT_FINANCE,"
            "PREDICT_AMT_LOWER,PREDICT_AMT_UPPER,PREDICT_CONTENT,"
            "CHANGE_REASON_EXPLAIN,PREDICT_TYPE"
        ),
        "filter": f"(SECURITY_CODE in ({quoted_codes}))",
        "pageNumber": "1",
        "pageSize": "10000",
        "sortTypes": "-1",
        "sortColumns": "NOTICE_DATE",
        "source": "WEB",
        "client": "WEB",
    }
    payload, error = request_json(
        EASTMONEY_FINANCE_ENDPOINT,
        params=params,
        referer="https://data.eastmoney.com/",
    )
    endpoint = f"{EASTMONEY_FINANCE_ENDPOINT}?{urllib.parse.urlencode(params)}"
    if payload is None:
        return [], error, endpoint
    return list((payload.get("result") or {}).get("data") or []), None, endpoint


def numeric_variants(value: Decimal) -> set[str]:
    variants: set[str] = set()
    in_ten_thousands = value / Decimal("10000")
    # 报告首页“主要会计数据”常以万元并保留两位小数，财务报表正文则可能
    # 以元列示；两种显示都允许自动文本核对。
    for candidate in (
        value,
        in_ten_thousands,
        in_ten_thousands.quantize(Decimal("0.01")),
    ):
        plain = format(candidate, "f")
        variants.add(plain)
        variants.add(plain.rstrip("0").rstrip("."))
    return {item for item in variants if item}


def verify_pdf_values(url: str, values: list[Decimal]) -> dict[str, Any]:
    if not url:
        return {"status": "source_missing", "matched": []}
    if PdfReader is None:
        return {"status": "pypdf_unavailable", "matched": []}
    binary, error = request_bytes(url, referer="https://www.cninfo.com.cn/")
    if binary is None:
        return {"status": "download_failed", "error": error, "matched": []}
    try:
        reader = PdfReader(io.BytesIO(binary), strict=False)
        text = "\n".join(
            (page.extract_text() or "") for page in reader.pages[: min(15, len(reader.pages))]
        )
    except Exception as exc:  # PDF text extraction is an auxiliary verification check.
        return {"status": "text_extract_failed", "error": f"{type(exc).__name__}: {exc}", "matched": []}
    normalized = re.sub(r"[\s,，]", "", text)
    matched: list[bool] = []
    for value in values:
        matched.append(any(variant in normalized for variant in numeric_variants(value)))
    return {
        "status": "matched" if all(matched) else "value_not_auto_matched",
        "matched": matched,
        "pages_scanned": min(15, len(reader.pages)),
    }


def latest_schedule(entry: dict[str, Any] | None) -> str:
    if not entry:
        return ""
    if entry.get("f006d_0102"):
        return str(entry["f006d_0102"])
    for key in ("f005d_0102", "f004d_0102", "f003d_0102", "f002d_0102"):
        if entry.get(key):
            return str(entry[key])
    return ""


def style_sheet(ws, *, freeze: str | None = None, filter_row: int | None = None) -> None:
    if freeze:
        ws.freeze_panes = freeze
    if filter_row:
        ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False


def apply_header_style(cells) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in cells:
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_workbook(
    output_path: Path,
    rows: list[dict[str, Any]],
    run_time: str,
    snapshot_time: str,
    financial_endpoint: str,
    prediction_endpoint: str,
    source_verification: dict[str, Any],
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "汇总"
    raw = wb.create_sheet("数据")
    calc = wb.create_sheet("计算")
    notes = wb.create_sheet("说明")

    title = "中证半导体材料设备主题指数（931743）半年报、Q2扣非与产品应用汇总"
    summary.merge_cells("A1:S1")
    summary["A1"] = title
    summary["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    summary.merge_cells("A2:S2")
    summary["A2"] = (
        f"成分：东方财富 {snapshot_time}（40只），中证官方成分表 2026-08-07 逐代码交叉一致；"
        f"权重：中证 2026-07-31 收盘快照；财报核验截至 {run_time[:10]}。"
    )
    summary["A2"].font = Font(name="Microsoft YaHei", size=9, color="404040")
    summary["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    summary.merge_cells("A3:S3")
    summary["A3"] = (
        "计算口径：2026Q2扣非=2026H1扣非−2026Q1扣非；环比=Q2/Q1−1；"
        "同比=Q2/(2025H1−2025Q1)−1；用户定义PE=总市值/(Q2扣非×4)，非TTM PE。"
    )
    summary["A3"].font = Font(name="Microsoft YaHei", size=9, color="9C0006")
    summary["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    summary.row_dimensions[2].height = 26
    summary.row_dimensions[3].height = 28

    summary_headers = [
        "序号", "代码", "名称", "中证权重%", "总市值\n(亿元)", "最新价\n(元)",
        "半年报/预告状态", "公告日期", "当前预约/\n实际披露日", "数值口径",
        "H1扣非\n(亿元)", "Q2扣非\n(亿元)", "Q2扣非环比", "Q2扣非同比",
        "用户定义PE\n(Q2扣非×4)", "产品", "主要应用场景", "场景证据级别", "边界/说明",
    ]
    for column, header in enumerate(summary_headers, 1):
        summary.cell(5, column, header)
    apply_header_style(summary[5])
    summary.row_dimensions[5].height = 38

    raw_headers = [
        "代码", "Ticker", "标准简称", "中证权重%", "最新价(CNY)", "涨跌幅%", "总市值(CNY)",
        "东财行情时间", "半年报/预告状态", "公告日期", "公告标题", "官方公告PDF",
        "首次预约", "第一次变更", "第二次变更", "第三次变更", "实际披露", "当前预约/实际日",
        "数值口径", "H1扣非下限(CNY)", "H1扣非上限(CNY)", "2026Q1扣非(CNY)",
        "2025H1扣非(CNY)", "2025Q1扣非(CNY)", "2026H1来源", "2026Q1来源",
        "2025H1来源", "2025Q1来源", "财务原文数值自动核验", "产品", "主要应用场景",
        "产品场景证据", "场景PDF页码", "场景证据级别", "边界/说明", "预约来源", "财务候选来源",
    ]
    raw.append(raw_headers)
    apply_header_style(raw[1])

    calc_headers = [
        "代码", "披露状态", "数值口径", "H1扣非下限(CNY)", "H1扣非上限(CNY)",
        "2026Q1扣非(CNY)", "2025H1扣非(CNY)", "2025Q1扣非(CNY)",
        "2026Q2扣非下限(CNY)", "2026Q2扣非上限(CNY)", "2025Q2扣非(CNY)",
        "Q2环比下限", "Q2环比上限", "Q2同比下限", "Q2同比上限",
        "用户定义PE下限", "用户定义PE上限",
    ]
    calc.append(calc_headers)
    apply_header_style(calc[1])

    blue_font = Font(name="Microsoft YaHei", size=9, color="0000FF")
    body_font = Font(name="Microsoft YaHei", size=9, color="000000")
    thin_gray = Side(style="thin", color="D9E2F3")

    for index, row in enumerate(rows, 1):
        raw_row = index + 1
        summary_row = index + 5
        values = [
            row["code"], row["ticker"], row["name"], float_or_none(row["weight"]),
            row["price"], row["change_pct"], row["market_cap"], row["snapshot_time"],
            row["disclosure_status"], row["announcement_date"], row["announcement_title"],
            row["announcement_pdf"], row["first_appointment"], row["first_change"],
            row["second_change"], row["third_change"], row["actual_publish"],
            row["scheduled_or_actual"], row["financial_mode"], float_or_none(row["h1_low"]),
            float_or_none(row["h1_high"]), float_or_none(row["q1_2026"]),
            float_or_none(row["h1_2025"]), float_or_none(row["q1_2025"]),
            row["source_h1_2026"], row["source_q1_2026"], row["source_h1_2025"],
            row["source_q1_2025"], row["value_verification"], row["product"],
            row["scenario"], row["application_url"], row["application_locator"],
            row["application_confidence"], row["application_boundary"],
            "巨潮预约披露 getPrbookInfo", financial_endpoint,
        ]
        raw.append(values)
        for cell in raw[raw_row]:
            cell.font = blue_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
        for col in (12, 25, 26, 27, 28, 32):
            source_cell = raw.cell(raw_row, col)
            if source_cell.value:
                source_cell.hyperlink = str(source_cell.value)
                source_cell.style = "Hyperlink"

        calc.append([
            f"='数据'!A{raw_row}", f"='数据'!I{raw_row}", f"='数据'!S{raw_row}",
            f"='数据'!T{raw_row}", f"='数据'!U{raw_row}", f"='数据'!V{raw_row}",
            f"='数据'!W{raw_row}", f"='数据'!X{raw_row}",
            f'=IF(OR(D{raw_row}="",F{raw_row}=""),"",D{raw_row}-F{raw_row})',
            f'=IF(OR(E{raw_row}="",F{raw_row}=""),"",E{raw_row}-F{raw_row})',
            f'=IF(OR(G{raw_row}="",H{raw_row}=""),"",G{raw_row}-H{raw_row})',
            f'=IF(OR(I{raw_row}="",F{raw_row}="",F{raw_row}<=0),"",I{raw_row}/F{raw_row}-1)',
            f'=IF(OR(J{raw_row}="",F{raw_row}="",F{raw_row}<=0),"",J{raw_row}/F{raw_row}-1)',
            f'=IF(OR(I{raw_row}="",K{raw_row}="",K{raw_row}<=0),"",I{raw_row}/K{raw_row}-1)',
            f'=IF(OR(J{raw_row}="",K{raw_row}="",K{raw_row}<=0),"",J{raw_row}/K{raw_row}-1)',
            f'=IF(OR(I{raw_row}="",J{raw_row}="",\'数据\'!G{raw_row}=""),"",IF(OR(I{raw_row}<=0,J{raw_row}<=0),"",\'数据\'!G{raw_row}/(J{raw_row}*4)))',
            f'=IF(OR(I{raw_row}="",J{raw_row}="",\'数据\'!G{raw_row}=""),"",IF(OR(I{raw_row}<=0,J{raw_row}<=0),"",\'数据\'!G{raw_row}/(I{raw_row}*4)))',
        ])
        for cell in calc[raw_row]:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

        summary.cell(summary_row, 1, index)
        summary.cell(summary_row, 2, f"='数据'!A{raw_row}")
        summary.cell(summary_row, 3, f"='数据'!C{raw_row}")
        summary.cell(summary_row, 4, f"='数据'!D{raw_row}")
        summary.cell(summary_row, 5, f"='数据'!G{raw_row}/100000000")
        summary.cell(summary_row, 6, f"='数据'!E{raw_row}")
        summary.cell(summary_row, 7, f"='数据'!I{raw_row}")
        summary.cell(summary_row, 8, f"='数据'!J{raw_row}")
        summary.cell(summary_row, 9, f"='数据'!R{raw_row}")
        summary.cell(summary_row, 10, f"='数据'!S{raw_row}")
        summary.cell(
            summary_row, 11,
            f'=IF(OR(计算!D{raw_row}="",计算!E{raw_row}=""),"N/A",'
            f'IF(ABS(计算!D{raw_row}-计算!E{raw_row})<0.00001,'
            f'TEXT(计算!D{raw_row}/100000000,"0.00"),'
            f'TEXT(计算!D{raw_row}/100000000,"0.00")&"~"&TEXT(计算!E{raw_row}/100000000,"0.00")))'
        )
        summary.cell(
            summary_row, 12,
            f'=IF(OR(计算!I{raw_row}="",计算!J{raw_row}=""),"N/A",'
            f'IF(ABS(计算!I{raw_row}-计算!J{raw_row})<0.00001,'
            f'TEXT(计算!I{raw_row}/100000000,"0.00"),'
            f'TEXT(计算!I{raw_row}/100000000,"0.00")&"~"&TEXT(计算!J{raw_row}/100000000,"0.00")))'
        )
        summary.cell(
            summary_row, 13,
            f'=IF(OR(计算!I{raw_row}="",计算!J{raw_row}=""),"N/A",'
            f'IF(OR(计算!L{raw_row}="",计算!M{raw_row}=""),"N/M",'
            f'IF(ABS(计算!L{raw_row}-计算!M{raw_row})<0.0000001,'
            f'TEXT(计算!L{raw_row},"0.0%"),'
            f'TEXT(计算!L{raw_row},"0.0%")&"~"&TEXT(计算!M{raw_row},"0.0%"))))'
        )
        summary.cell(
            summary_row, 14,
            f'=IF(OR(计算!I{raw_row}="",计算!J{raw_row}=""),"N/A",'
            f'IF(OR(计算!N{raw_row}="",计算!O{raw_row}=""),"N/M",'
            f'IF(ABS(计算!N{raw_row}-计算!O{raw_row})<0.0000001,'
            f'TEXT(计算!N{raw_row},"0.0%"),'
            f'TEXT(计算!N{raw_row},"0.0%")&"~"&TEXT(计算!O{raw_row},"0.0%"))))'
        )
        summary.cell(
            summary_row, 15,
            f'=IF(OR(计算!I{raw_row}="",计算!J{raw_row}=""),"N/A",'
            f'IF(OR(计算!I{raw_row}<=0,计算!J{raw_row}<=0),"N/M",'
            f'IF(ABS(计算!P{raw_row}-计算!Q{raw_row})<0.0000001,'
            f'TEXT(计算!P{raw_row},"0.0")&"x",'
            f'TEXT(计算!P{raw_row},"0.0")&"x~"&TEXT(计算!Q{raw_row},"0.0")&"x")))'
        )
        summary.cell(summary_row, 16, f"='数据'!AD{raw_row}")
        summary.cell(summary_row, 17, f"='数据'!AE{raw_row}")
        summary.cell(summary_row, 18, f"='数据'!AH{raw_row}")
        summary.cell(summary_row, 19, f"='数据'!AI{raw_row}")

        for cell in summary[summary_row]:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
        summary.row_dimensions[summary_row].height = 54
        summary.cell(summary_row, 4).number_format = "0.000"
        summary.cell(summary_row, 5).number_format = "0.00"
        summary.cell(summary_row, 6).number_format = "0.00"

    summary_widths = [6, 11, 13, 11, 13, 10, 19, 12, 15, 30, 15, 15, 15, 15, 20, 23, 45, 13, 32]
    for column, width in enumerate(summary_widths, 1):
        summary.column_dimensions[get_column_letter(column)].width = width
    raw_widths = [10, 12, 13, 11, 12, 10, 17, 23, 18, 12, 38, 42, 12, 12, 12, 12, 12, 15, 35, 18, 18, 18, 18, 18, 35, 35, 35, 35, 30, 24, 55, 45, 18, 12, 38, 25, 45]
    for column, width in enumerate(raw_widths, 1):
        raw.column_dimensions[get_column_letter(column)].width = width
    for column in range(1, len(calc_headers) + 1):
        calc.column_dimensions[get_column_letter(column)].width = 20

    for ws in (summary, raw, calc):
        style_sheet(ws, freeze="A6" if ws is summary else "A2", filter_row=5 if ws is summary else 1)
    summary.conditional_formatting.add(
        f"G6:G{summary.max_row}",
        FormulaRule(formula=['$G6="半年报已发布"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    summary.conditional_formatting.add(
        f"G6:G{summary.max_row}",
        FormulaRule(formula=['$G6="半年业绩预告已发布"'], fill=PatternFill("solid", fgColor="FFEB9C")),
    )
    summary.conditional_formatting.add(
        f"G6:G{summary.max_row}",
        FormulaRule(formula=['$G6="未发布半年报/预告"'], fill=PatternFill("solid", fgColor="FCE4D6")),
    )

    notes.merge_cells("A1:D1")
    notes["A1"] = "口径、来源与证据边界"
    notes["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="17365D")
    notes["A1"].alignment = Alignment(horizontal="center")
    note_rows = [
        ("数据截至", run_time),
        ("指数成分", "东方财富当前成分 40 只；中证官方成分表（2026-08-07）逐代码交叉一致。"),
        ("名称标准化", "采用中证官方简称；东财展示的 XD芯源微→芯源微，中巨芯-U→中巨芯。"),
        ("指数权重", "中证 2026-07-31 收盘权重快照，四舍五入合计 100.003%。"),
        ("半年报状态", "巨潮资讯公司公告原文优先；“未发布”指截至本表截点未命中正式半年报或半年业绩预告原文。"),
        ("预约披露", "巨潮预约披露：实际披露优先；未披露时取第三/第二/第一次变更、再取首次预约的最后有效日期。"),
        ("Q2扣非", "2026Q2=2026H1扣非−2026Q1扣非；2025Q2=2025H1扣非−2025Q1扣非。"),
        ("Q2环比", "Q2/Q1−1；Q1≤0 时记 N/M，避免把亏损基数伪装成可解释增速。"),
        ("Q2同比", "Q2/2025Q2−1；2025Q2≤0 时记 N/M。"),
        ("用户定义PE", "总市值/(Q2单季扣非×4)，使用东方财富抓取时点总市值；非 TTM PE、非一致预期 PE；Q2≤0 时 N/M。"),
        ("预告处理", "预告仅展示预告区间推导的 Q2/增速/PE；未披露实际半年报的公司不标为实际值。"),
        ("产品场景", "仅采用公司2025年报可直接支持的产品—应用描述；N/A/待补为明确证据缺口，未用概念标签、新闻或客户推断填补。"),
        ("财务数值", "东方财富 RPT_F10_FINANCE_GINCOME 作候选；已披露公司回链巨潮PDF，并在“数据”记录自动文本匹配结果。"),
        ("财务审计", "同目录 evidence_audit_input.json / evidence_audit_result.json：对实际半年报公司做官方原文与数值候选交叉、Q2环比/同比口径审计。"),
        ("中证成分表", CSINDEX_COMPONENT_URL),
        ("中证权重表", CSINDEX_WEIGHT_URL),
        ("东方财富成分接口", EASTMONEY_COMPONENT_URL),
        ("财务候选接口", financial_endpoint),
        ("预告候选接口", prediction_endpoint),
    ]
    for r, (label, value) in enumerate(note_rows, 3):
        notes.cell(r, 1, label)
        notes.cell(r, 2, value)
        notes.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        notes.cell(r, 1).font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E78")
        notes.cell(r, 2).font = Font(name="Microsoft YaHei", size=10)
        notes.cell(r, 1).alignment = Alignment(vertical="top", wrap_text=True)
        notes.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)
        notes.row_dimensions[r].height = 34
        if isinstance(value, str) and value.startswith("http"):
            notes.cell(r, 2).hyperlink = value
            notes.cell(r, 2).style = "Hyperlink"
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 70
    notes.column_dimensions["C"].width = 15
    notes.column_dimensions["D"].width = 15
    style_sheet(notes, freeze="A3")

    notes["A23"] = "原文数值自动核验统计"
    notes["A23"].font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E78")
    notes["B23"] = json.dumps(source_verification.get("counts", {}), ensure_ascii=False)
    notes["B23"].alignment = Alignment(wrap_text=True)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output_path)


def duration_period(start: str, end: str, label: str, frequency: str = "quarter") -> dict[str, str]:
    return {"kind": "duration", "start": start, "end": end, "frequency": frequency, "label": label}


def make_audit_input(
    rows: list[dict[str, Any]], run_time: str, output_path: Path
) -> dict[str, Any]:
    """审计实际披露公司的原始输入、Q2同比及环比。

    审计器没有“半年报减一季报”的独立 check 类型，因此 Q2 事实显式标为由
    两份官方原文推导，工作簿公式与单独的 Decimal 算术清单共同复核该差额。
    """
    sources: list[dict[str, Any]] = []
    facts: list[dict[str, Any]] = []
    checks: list[dict[str, Any]] = []

    periods = {
        "h1_2026": ("2026-01-01", "2026-06-30", "H1 FY2026", "h1_low", "half"),
        "q1_2026": ("2026-01-01", "2026-03-31", "Q1 FY2026", "q1_2026", "quarter"),
        "h1_2025": ("2025-01-01", "2025-06-30", "H1 FY2025", "h1_2025", "half"),
        "q1_2025": ("2025-01-01", "2025-03-31", "Q1 FY2025", "q1_2025", "quarter"),
    }
    for row in rows:
        if row["financial_mode"] != "实际值：PDF逐期核验+数值交叉":
            continue
        code = row["code"]
        official_fact_ids: dict[str, str] = {}
        for period_key, (start, end, label, value_key, frequency) in periods.items():
            value = row[value_key]
            source_url = row[f"source_{period_key}"]
            available_at = row["period_dates"].get(period_key, "")
            if value is None or not source_url or not available_at:
                continue
            official_source = f"S_OFFICIAL_{code}_{period_key}"
            vendor_source = f"S_VENDOR_{code}_{period_key}"
            official_fact = f"F_OFFICIAL_{code}_{period_key}"
            vendor_fact = f"F_VENDOR_{code}_{period_key}"
            sources.extend([
                {
                    "id": official_source,
                    "source_type": "official_filing",
                    "origin_id": f"cninfo:{code}:{period_key}",
                    "locator": source_url,
                    "source_date": available_at,
                    "checked_at": run_time,
                    "status": "accepted",
                },
                {
                    "id": vendor_source,
                    "source_type": "market_data_vendor",
                    "origin_id": f"eastmoney:gincome:{code}:{period_key}",
                    "locator": (
                        f"{EASTMONEY_FINANCE_ENDPOINT}?"
                        f"reportName=RPT_F10_FINANCE_GINCOME&code={code}&period={period_key}"
                    ),
                    "source_date": available_at,
                    "checked_at": run_time,
                    "status": "accepted",
                },
            ])
            common = {
                "metric": "deducted_attributable_net_profit",
                "value": str(value),
                "unit": "currency",
                "currency": "CNY",
                "scale": "1",
                "period": duration_period(start, end, label, frequency),
                "available_at": available_at,
                "basis": "reported_consolidated_prc_gaap_deducted_attributable_net_profit",
            }
            facts.append({"id": official_fact, **common, "source_refs": [official_source]})
            facts.append({"id": vendor_fact, **common, "source_refs": [vendor_source]})
            official_fact_ids[period_key] = official_fact
            checks.append({
                "id": f"C_XS_{code}_{period_key}",
                "kind": "cross_source",
                "materiality": "material",
                "target": {"fact_id": vendor_fact},
                "references": [{"fact_id": official_fact}],
                "source_gate": {
                    "min_independent_origins": 1,
                    "counted_tier": "official",
                    "required_anchor_tier": "official",
                },
                "tolerance": {"relative_pct": "0", "absolute_base": "0"},
            })

        if len(official_fact_ids) != 4:
            continue
        q2_2026 = row["h1_low"] - row["q1_2026"]
        q2_2025 = row["h1_2025"] - row["q1_2025"]
        q2_2026_fact = f"F_DERIVED_{code}_q2_2026"
        q2_2025_fact = f"F_DERIVED_{code}_q2_2025"
        common_q2 = {
            "metric": "deducted_attributable_net_profit",
            "unit": "currency",
            "currency": "CNY",
            "scale": "1",
            "basis": "reported_consolidated_prc_gaap_deducted_attributable_net_profit",
        }
        facts.extend([
            {
                "id": q2_2026_fact,
                **common_q2,
                "value": str(q2_2026),
                "period": duration_period("2026-04-01", "2026-06-30", "Q2 FY2026"),
                "available_at": row["period_dates"]["h1_2026"],
                "source_refs": [
                    official_fact_ids["h1_2026"].replace("F_OFFICIAL", "S_OFFICIAL"),
                    official_fact_ids["q1_2026"].replace("F_OFFICIAL", "S_OFFICIAL"),
                ],
            },
            {
                "id": q2_2025_fact,
                **common_q2,
                "value": str(q2_2025),
                "period": duration_period("2025-04-01", "2025-06-30", "Q2 FY2025"),
                "available_at": row["period_dates"]["h1_2025"],
                "source_refs": [
                    official_fact_ids["h1_2025"].replace("F_OFFICIAL", "S_OFFICIAL"),
                    official_fact_ids["q1_2025"].replace("F_OFFICIAL", "S_OFFICIAL"),
                ],
            },
        ])
        checks.extend([
            {
                "id": f"C_QOQ_{code}",
                "kind": "percentage",
                "materiality": "material",
                "mode": "change",
                "current": {"fact_id": q2_2026_fact},
                "base": {"fact_id": official_fact_ids["q1_2026"]},
                "period_relation": "qoq",
                "output_metric": "deducted_attributable_net_profit_qoq_pct",
                "output_basis": "reported_consolidated_prc_gaap_deducted_attributable_net_profit_qoq",
                "source_gate": {
                    "min_independent_origins": 1,
                    "counted_tier": "official",
                    "required_anchor_tier": "official",
                },
            },
            {
                "id": f"C_YOY_{code}",
                "kind": "percentage",
                "materiality": "material",
                "mode": "change",
                "current": {"fact_id": q2_2026_fact},
                "base": {"fact_id": q2_2025_fact},
                "period_relation": "yoy",
                "output_metric": "deducted_attributable_net_profit_yoy_pct",
                "output_basis": "reported_consolidated_prc_gaap_deducted_attributable_net_profit_yoy",
                "source_gate": {
                    "min_independent_origins": 1,
                    "counted_tier": "official",
                    "required_anchor_tier": "official",
                },
            },
        ])
    return {
        "schema_version": "1.0",
        "audit_id": f"931743-q2-deducted-{run_time[:10]}",
        "as_of": run_time,
        "sources": sources,
        "facts": facts,
        "checks": checks,
        "metadata": {
            "workbook": str(output_path),
            "scope": "only actual H1 reports; forecast ranges excluded from actual financial audit",
            "q2_derivation": "H1 deducted attributable net profit minus Q1 deducted attributable net profit",
        },
    }


def build_rows(cutoff: str, out_dir: Path) -> tuple[list[dict[str, Any]], dict[str, Any], str, str]:
    companies = parse_weights()
    applications = parse_applications()
    run_time = now_iso()
    component_payload = fetch_component_snapshot()
    snapshot_data = component_payload.get("data") or {}
    snapshot_time = run_time
    if snapshot_data.get("diff") is None:
        raise RuntimeError("东方财富指数成分未返回 diff")
    out_dir.mkdir(parents=True, exist_ok=True)
    write_json(out_dir / "index_snapshot_eastmoney.json", component_payload)

    market_data: dict[str, dict[str, Any]] = {
        str(item["f12"]): item for item in snapshot_data["diff"]
    }
    if set(market_data) != set(companies):
        raise RuntimeError(
            "东方财富与中证官方成分代码不一致："
            f"only_eastmoney={sorted(set(market_data) - set(companies))} "
            f"only_csindex={sorted(set(companies) - set(market_data))}"
        )

    code_list = list(companies)
    finance_rows, finance_error, finance_endpoint = fetch_financials(code_list)
    prediction_rows, prediction_error, prediction_endpoint = fetch_predictions(code_list)
    if finance_error:
        raise RuntimeError(f"东方财富财务候选接口失败：{finance_error}")
    if prediction_error:
        raise RuntimeError(f"东方财富业绩预告候选接口失败：{prediction_error}")
    write_json(out_dir / "finance_candidates_eastmoney.json", finance_rows)
    write_json(out_dir / "forecast_candidates_eastmoney.json", prediction_rows)

    finance_by_code: dict[str, dict[str, Decimal]] = {code: {} for code in companies}
    for item in finance_rows:
        code = str(item.get("SECURITY_CODE") or "")
        date = str(item.get("REPORT_DATE") or "")[:10]
        if code in finance_by_code and date in {"2026-06-30", "2026-03-31", "2025-06-30", "2025-03-31"}:
            value = decimal_or_none(item.get("DEDUCT_PARENT_NETPROFIT"))
            if value is not None:
                finance_by_code[code][date] = value

    forecast_by_code: dict[str, dict[str, Any]] = {}
    for item in prediction_rows:
        if str(item.get("REPORT_DATE") or "")[:10] != "2026-06-30":
            continue
        code = str(item.get("SECURITY_CODE") or "")
        if code not in companies:
            continue
        metric = str(item.get("PREDICT_FINANCE") or "")
        if "扣除非经常性损益" in metric:
            current = forecast_by_code.get(code)
            if current is None or str(item.get("NOTICE_DATE") or "") > str(current.get("NOTICE_DATE") or ""):
                forecast_by_code[code] = item

    schedule: dict[str, dict[str, Any] | None] = {}
    schedule_errors: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = [pool.submit(get_prbook, company) for company in companies.values()]
        for future in as_completed(futures):
            code, entry, error = future.result()
            schedule[code] = entry
            if error:
                schedule_errors[code] = error
    if len(schedule) != 40:
        raise RuntimeError(f"巨潮预约披露返回覆盖不足：{len(schedule)}")
    write_json(out_dir / "cninfo_prbook_2026h1.json", {"rows": schedule, "errors": schedule_errors})

    status_announcements: dict[str, list[dict[str, Any]]] = {}
    announcement_errors: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = [pool.submit(get_status_announcements, company, cutoff) for company in companies.values()]
        for future in as_completed(futures):
            code, items, error = future.result()
            status_announcements[code] = items
            if error:
                announcement_errors[code] = error
    write_json(
        out_dir / "cninfo_h1_announcements.json",
        {"rows": status_announcements, "errors": announcement_errors},
    )

    status_info: dict[str, dict[str, Any]] = {}
    for code, company in companies.items():
        items = status_announcements.get(code, [])
        actual = choose_announcement(items, "2026年半年度报告")
        forecast = choose_h1_forecast(items)
        if actual:
            status = "半年报已发布"
            selected = actual
        elif forecast:
            status = "半年业绩预告已发布"
            selected = forecast
        else:
            status = "未发布半年报/预告"
            selected = None
        status_info[code] = {
            "status": status,
            "actual": actual,
            "forecast": forecast,
            "selected": selected,
        }

    period_sources: dict[str, dict[str, dict[str, Any] | None]] = {code: {} for code in companies}
    needed_codes = [
        code
        for code, info in status_info.items()
        if info["status"] == "半年报已发布"
        or (info["status"] == "半年业绩预告已发布" and code in forecast_by_code)
    ]
    period_futures = []
    with ThreadPoolExecutor(max_workers=5) as pool:
        for code in needed_codes:
            company = companies[code]
            for key in ("q1_2025", "h1_2025", "q1_2026"):
                period_futures.append(pool.submit(get_period_announcement, company, key))
            if status_info[code]["status"] == "半年报已发布":
                period_futures.append(pool.submit(get_period_announcement, company, "h1_2026"))
        for future in as_completed(period_futures):
            code, key, item, error = future.result()
            period_sources[code][key] = item
            if error:
                announcement_errors[f"{code}:{key}"] = error
    for code, info in status_info.items():
        if info["actual"]:
            period_sources[code]["h1_2026"] = info["actual"]
    write_json(out_dir / "cninfo_period_sources.json", period_sources)

    verification_tasks: list[tuple[str, str, str, list[Decimal]]] = []
    for code in needed_codes:
        values = finance_by_code[code]
        info = status_info[code]
        if info["status"] == "半年报已发布":
            mapping = {
                "h1_2026": "2026-06-30",
                "q1_2026": "2026-03-31",
                "h1_2025": "2025-06-30",
                "q1_2025": "2025-03-31",
            }
            for key, date in mapping.items():
                item = period_sources[code].get(key)
                value = values.get(date)
                if item and value is not None:
                    verification_tasks.append((code, key, cninfo_url(item), [value]))
        elif code in forecast_by_code and info["forecast"]:
            prediction = forecast_by_code[code]
            low = decimal_or_none(prediction.get("PREDICT_AMT_LOWER"))
            high = decimal_or_none(prediction.get("PREDICT_AMT_UPPER"))
            if low is not None and high is not None:
                verification_tasks.append((code, "forecast_h1_2026", cninfo_url(info["forecast"]), [low, high]))

    verification: dict[str, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {
            pool.submit(verify_pdf_values, url, values): (code, key)
            for code, key, url, values in verification_tasks
        }
        for future in as_completed(futures):
            code, key = futures[future]
            verification.setdefault(code, {})[key] = future.result()
    verification_counts: dict[str, int] = {}
    for company_checks in verification.values():
        for result in company_checks.values():
            status = str(result.get("status") or "")
            verification_counts[status] = verification_counts.get(status, 0) + 1
    verification_summary = {"checks": verification, "counts": verification_counts}
    write_json(out_dir / "cninfo_pdf_value_verification.json", verification_summary)

    rows: list[dict[str, Any]] = []
    for company in sorted(companies.values(), key=lambda item: item.weight, reverse=True):
        code = company.code
        market = market_data[code]
        info = status_info[code]
        appointment = schedule.get(code) or {}
        selected = info["selected"]
        financial_mode = "N/A"
        h1_low: Decimal | None = None
        h1_high: Decimal | None = None
        required_actual_verification = ("h1_2026", "q1_2026", "h1_2025", "q1_2025")
        exact_verification = all(
            verification.get(code, {}).get(key, {}).get("status") == "matched"
            for key in required_actual_verification
        )
        if info["status"] == "半年报已发布" and finance_by_code[code].get("2026-06-30") is not None:
            h1_low = h1_high = finance_by_code[code]["2026-06-30"]
            financial_mode = (
                "实际值：PDF逐期核验+数值交叉"
                if exact_verification
                else "实际值：已定位原文；数值文本未全自动匹配"
            )
        elif info["status"] == "半年业绩预告已发布" and code in forecast_by_code:
            prediction = forecast_by_code[code]
            h1_low = decimal_or_none(prediction.get("PREDICT_AMT_LOWER"))
            h1_high = decimal_or_none(prediction.get("PREDICT_AMT_UPPER"))
            forecast_status = verification.get(code, {}).get("forecast_h1_2026", {}).get("status")
            financial_mode = (
                "预告推导区间：预告PDF数值匹配"
                if forecast_status == "matched"
                else "预告推导区间：数值候选待PDF文本复核"
            )
        elif info["status"] == "半年业绩预告已发布":
            financial_mode = "预告已发：未提供扣非范围，Q2扣非/PE N/A"

        sources = period_sources.get(code, {})
        period_dates = {
            "h1_2026": ms_to_date((sources.get("h1_2026") or {}).get("announcementTime")),
            "q1_2026": ms_to_date((sources.get("q1_2026") or {}).get("announcementTime")),
            "h1_2025": ms_to_date((sources.get("h1_2025") or {}).get("announcementTime")),
            "q1_2025": ms_to_date((sources.get("q1_2025") or {}).get("announcementTime")),
        }
        application = applications[code]
        rows.append({
            "code": code,
            "ticker": company.ticker,
            "name": company.name,
            "weight": company.weight,
            "price": decimal_or_none(market.get("f2")) / Decimal("100") if market.get("f2") is not None else None,
            "change_pct": decimal_or_none(market.get("f3")) / Decimal("100") if market.get("f3") is not None else None,
            "market_cap": decimal_or_none(market.get("f20")),
            "snapshot_time": snapshot_time,
            "disclosure_status": info["status"],
            "announcement_date": ms_to_date((selected or {}).get("announcementTime")),
            "announcement_title": str((selected or {}).get("announcementTitle") or ""),
            "announcement_pdf": cninfo_url(selected or {}),
            "first_appointment": str(appointment.get("f002d_0102") or ""),
            "first_change": str(appointment.get("f003d_0102") or ""),
            "second_change": str(appointment.get("f004d_0102") or ""),
            "third_change": str(appointment.get("f005d_0102") or ""),
            "actual_publish": str(appointment.get("f006d_0102") or ""),
            "scheduled_or_actual": latest_schedule(appointment),
            "financial_mode": financial_mode,
            "h1_low": h1_low,
            "h1_high": h1_high,
            "q1_2026": finance_by_code[code].get("2026-03-31"),
            "h1_2025": finance_by_code[code].get("2025-06-30"),
            "q1_2025": finance_by_code[code].get("2025-03-31"),
            "source_h1_2026": cninfo_url(sources.get("h1_2026") or {}),
            "source_q1_2026": cninfo_url(sources.get("q1_2026") or {}),
            "source_h1_2025": cninfo_url(sources.get("h1_2025") or {}),
            "source_q1_2025": cninfo_url(sources.get("q1_2025") or {}),
            "period_dates": period_dates,
            "value_verification": json.dumps(verification.get(code, {}), ensure_ascii=False),
            "product": application["product"],
            "scenario": application["scenario"],
            "application_url": application["url"],
            "application_locator": application["locator"],
            "application_confidence": application["confidence"],
            "application_boundary": application["boundary"],
        })

    if len(rows) != 40:
        raise RuntimeError(f"最终汇总行数异常：{len(rows)}")
    metadata = {
        "run_time": run_time,
        "snapshot_time": snapshot_time,
        "disclosure_counts": {
            status: sum(1 for row in rows if row["disclosure_status"] == status)
            for status in ("半年报已发布", "半年业绩预告已发布", "未发布半年报/预告")
        },
        "finance_error": finance_error,
        "prediction_error": prediction_error,
        "schedule_errors": schedule_errors,
        "announcement_errors": announcement_errors,
        "verification_counts": verification_counts,
        "csindex_component_source": CSINDEX_COMPONENT_URL,
        "csindex_weight_source": CSINDEX_WEIGHT_URL,
    }
    write_json(out_dir / "build_metadata.json", metadata)
    return rows, verification_summary, finance_endpoint, prediction_endpoint


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff", default=datetime.now(CN_TZ).date().isoformat(), help="YYYY-MM-DD")
    parser.add_argument("--out-dir", default="", help="默认 artifacts/931743_semiconductor_materials_YYYYMMDD")
    parser.add_argument("--overwrite", action="store_true", help="允许覆盖本脚本生成的同名工作簿")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        datetime.strptime(args.cutoff, "%Y-%m-%d")
    except ValueError as exc:
        raise SystemExit("--cutoff 必须为 YYYY-MM-DD") from exc
    out_dir = (
        Path(args.out_dir)
        if args.out_dir
        else Path("artifacts") / f"931743_semiconductor_materials_{args.cutoff.replace('-', '')}"
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook = out_dir / f"931743_半导体材料设备_截至{args.cutoff.replace('-', '')}.xlsx"
    if workbook.exists() and not args.overwrite:
        raise SystemExit(f"目标已存在，未覆盖：{workbook}；如需重建请传 --overwrite")
    lock_path = workbook.with_name("~$" + workbook.name)
    if lock_path.exists():
        raise SystemExit(f"检测到 Office/WPS 锁文件，未写入：{lock_path}")

    rows, verification, finance_endpoint, prediction_endpoint = build_rows(args.cutoff, out_dir)
    run_time = now_iso()
    build_workbook(
        workbook,
        rows,
        run_time,
        rows[0]["snapshot_time"],
        finance_endpoint,
        prediction_endpoint,
        verification,
    )
    audit_input = make_audit_input(rows, run_time, workbook)
    write_json(out_dir / "evidence_audit_input.json", audit_input)

    summary = {
        "workbook": str(workbook.resolve()),
        "rows": len(rows),
        "actual_h1": sum(row["disclosure_status"] == "半年报已发布" for row in rows),
        "forecast_only": sum(row["disclosure_status"] == "半年业绩预告已发布" for row in rows),
        "pending": sum(row["disclosure_status"] == "未发布半年报/预告" for row in rows),
        "run_time": run_time,
    }
    write_json(out_dir / "build_summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
