from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = PROJECT_ROOT / "watchlists" / "a_share_company_watchlist.xlsx"


HEADERS = [
    "enabled",
    "ticker",
    "exchange",
    "name",
    "aliases",
    "industry_tags",
    "priority",
    "baseline_status",
    "grok_query_terms",
    "tracking_focus",
    "official_sources_hint",
    "last_baseline_date",
    "last_update_date",
    "notes",
]


ROWS = [
    [
        "Y",
        "002428.SZ",
        "SZSE",
        "云南锗业",
        "云南锗业; Yunnan Germanium; 锗; InP; 红外材料",
        "锗; InP; 半导体材料; 红外",
        1,
        "pending",
        "云南锗业 OR 002428 OR Yunnan Germanium OR germanium OR InP",
        "锗资源、InP/化合物半导体材料、红外材料、公告与产能进展",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "002222.SZ",
        "SZSE",
        "福晶科技",
        "福晶科技; CASTECH; nonlinear crystal; LBO; BBO",
        "光学晶体; 激光; 光通信/光学元件",
        1,
        "pending",
        "福晶科技 OR 002222 OR CASTECH OR LBO OR BBO crystal",
        "非线性光学晶体、激光和光通信相关订单、客户与扩产",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "300476.SZ",
        "SZSE",
        "胜宏科技",
        "胜宏科技; Victory Giant Technology; AI PCB; HDI",
        "AI PCB; HDI; 服务器PCB",
        1,
        "pending",
        "胜宏科技 OR 300476 OR Victory Giant Technology OR AI PCB OR HDI",
        "AI服务器PCB、海外客户、订单兑现、产能与毛利率",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "603256.SH",
        "SSE",
        "宏和科技",
        "宏和科技; Grace Fabric; electronic cloth; glass fabric; T-glass",
        "电子布; 玻纤布; AI PCB上游材料",
        1,
        "pending",
        "宏和科技 OR 603256 OR electronic cloth OR glass fabric OR T-glass",
        "高端电子布/T-glass/Low-CTE材料进展、客户认证和价格变化",
        "CNINFO; SSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "601869.SH",
        "SSE",
        "长飞光纤",
        "长飞光纤; YOFC; Yangtze Optical Fibre; optical fiber; optical cable",
        "光纤光缆; 数据中心; 通信基础设施",
        1,
        "pending",
        "长飞光纤 OR 601869 OR YOFC OR optical fiber OR datacenter fiber",
        "光纤价格、数据中心需求、海外扩张、特种光纤进展",
        "CNINFO; SSE; 公司IR; HKEX if needed",
        "",
        "",
        "",
    ],
    [
        "Y",
        "301511.SZ",
        "SZSE",
        "德福科技",
        "德福科技; Defu Technology; HVLP copper foil; RTF copper foil",
        "HVLP铜箔; AI PCB材料; 铜箔",
        1,
        "pending",
        "德福科技 OR 301511 OR HVLP copper foil OR RTF copper foil OR AI PCB copper foil",
        "PCB级HVLP/RTF铜箔客户认证、批量交付和收入占比",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "002384.SZ",
        "SZSE",
        "东山精密",
        "东山精密; DSBJ; Dongshan Precision; PCB; FPC; AI server",
        "PCB; FPC; AI服务器; 消费电子",
        1,
        "pending",
        "东山精密 OR 002384 OR DSBJ OR Dongshan Precision OR AI server PCB",
        "PCB/FPC业务结构、AI服务器相关订单、海外客户和盈利修复",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "300308.SZ",
        "SZSE",
        "中际旭创",
        "中际旭创; Innolight; 800G; 1.6T; optical module; silicon photonics",
        "光模块; 800G; 1.6T; 硅光",
        1,
        "pending",
        "中际旭创 OR 300308 OR Innolight OR 800G OR 1.6T optical module",
        "800G/1.6T交付、硅光、客户需求、上游物料和毛利率",
        "CNINFO; SZSE; 公司IR",
        "",
        "",
        "",
    ],
    [
        "Y",
        "688498.SH",
        "SSE STAR",
        "源杰科技",
        "源杰科技; Yuanjie Semiconductor; EML; DFB; laser chip",
        "激光芯片; EML; DFB; InP; 光模块上游",
        1,
        "pending",
        "源杰科技 OR 688498 OR EML OR DFB OR laser chip OR InP",
        "高速激光芯片、200G EML、AI客户验证、量产和收入占比",
        "CNINFO; SSE STAR; 公司IR",
        "",
        "",
        "",
    ],
]


SCHEMA_ROWS = [
    ("enabled", "Y/N", "Y means track this company in every run."),
    ("ticker", "text", "A-share ticker with suffix, e.g. 300308.SZ or 603256.SH."),
    ("exchange", "text", "SZSE, SSE, or SSE STAR."),
    ("name", "text", "Company short name."),
    ("aliases", "semicolon text", "Names and aliases for search and Grok/X discovery."),
    ("industry_tags", "semicolon text", "Research tags used for grouping."),
    ("priority", "integer", "1 is highest priority."),
    ("baseline_status", "enum", "pending, done, refresh_needed, disabled."),
    ("grok_query_terms", "text", "Optional search string; automation may expand it."),
    ("tracking_focus", "text", "What the automation should pay special attention to."),
    ("official_sources_hint", "text", "Preferred official source hints."),
    ("last_baseline_date", "date", "Filled by automation after baseline creation."),
    ("last_update_date", "date", "Filled by automation after daily update."),
    ("notes", "text", "User notes."),
]


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "watchlist"
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)

    style_header(ws[1])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 10,
        "B": 13,
        "C": 12,
        "D": 14,
        "E": 48,
        "F": 34,
        "G": 10,
        "H": 18,
        "I": 56,
        "J": 54,
        "K": 26,
        "L": 18,
        "M": 18,
        "N": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"
    for row_idx in range(2, len(ROWS) + 2):
        ws[f"G{row_idx}"].number_format = "0"
        ws[f"L{row_idx}"].number_format = "yyyy-mm-dd"
        ws[f"M{row_idx}"].number_format = "yyyy-mm-dd"

    enabled_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    status_validation = DataValidation(
        type="list", formula1='"pending,done,refresh_needed,disabled"', allow_blank=True
    )
    priority_validation = DataValidation(
        type="whole", operator="between", formula1="1", formula2="3", allow_blank=False
    )
    ws.add_data_validation(enabled_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(priority_validation)
    enabled_validation.add("A2:A200")
    status_validation.add("H2:H200")
    priority_validation.add("G2:G200")

    schema = wb.create_sheet("schema")
    schema.append(["field", "type", "description"])
    for row in SCHEMA_ROWS:
        schema.append(row)
    style_header(schema[1])
    schema.column_dimensions["A"].width = 24
    schema.column_dimensions["B"].width = 18
    schema.column_dimensions["C"].width = 88
    for row in schema.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    schema.freeze_panes = "A2"
    schema.auto_filter.ref = f"A1:C{len(SCHEMA_ROWS) + 1}"

    return wb


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
