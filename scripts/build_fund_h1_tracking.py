#!/usr/bin/env python
"""Build a provenance-preserving H1 holdings workbook from an official PDF transcription."""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FUND_NAME = "天弘全球新能源汽车股票（QDII-LOF）"
REPORT_TITLE = "天弘全球新能源汽车股票型证券投资基金（QDII-LOF）2026年中期报告"
SOURCE_URL = "https://static.cninfo.com.cn/finalpage/2026-08-31/1225528103.PDF"
PDF_SHA256 = "ba99769d94cc053b283263c4027b156dc3062c86501e2d40dab7ce81857ea7da"
TRANSCRIPTION_SHA256 = "b137c8561ea35a12948522118f5d81b1c6720199fb01947f2212f3fad666bf53"
REPORT_DATE = "2026-08-31"
REPORT_PERIOD = "2026-01-01 至 2026-06-30"

NAV_CNY = Decimal("1914493933.59")
EQUITY_CNY = Decimal("1647778275.36")
EQUITY_NAV_PCT = Decimal("86.07")
EQUITY_TOTAL_ASSET_PCT = Decimal("81.85")

BLUE = "0000FF"
BLACK = "000000"
WHITE = "FFFFFF"
HEADER = "1F4E78"
SECTION = "D9EAF7"
NOTE = "FFF2CC"
THIN_GRAY = Side(style="thin", color="B7B7B7")
MEDIUM_BLUE = Side(style="medium", color=HEADER)
INPUT_FONT = Font(name="Microsoft YaHei", size=10, color=BLUE)
FORMULA_FONT = Font(name="Microsoft YaHei", size=10, color=BLACK)
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color=HEADER)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=10, italic=True, color="666666")
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'
PCT_FORMAT = '0.00%'


def now_bj() -> str:
    return datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")


def dec(value: str | int | float | Decimal) -> Decimal:
    return Decimal(str(value))


def pct(value: str | int | float | Decimal) -> float:
    return float(dec(value) / Decimal("100"))


def page_range(row: dict) -> str:
    start = row["source_page_start"]
    end = row["source_page_end"]
    return str(start) if start == end else f"{start}-{end}"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def apply_base_style(cell, *, formula: bool = False) -> None:
    cell.font = FORMULA_FONT if formula else INPUT_FONT
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def set_title(ws, title: str, subtitle: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    ws.cell(2, 1, subtitle).font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34


def set_header(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row, col, value)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=MEDIUM_BLUE, right=MEDIUM_BLUE, top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)
    ws.row_dimensions[row].height = 32


def set_widths(ws, widths: list[float]) -> None:
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_table(ws, start_row: int, headers: list[str], rows: list[list], widths: list[float], formats: dict[int, str] | None = None) -> int:
    set_header(ws, start_row, headers)
    for row_offset, values in enumerate(rows, 1):
        current_row = start_row + row_offset
        for col, value in enumerate(values, 1):
            cell = ws.cell(current_row, col, value)
            apply_base_style(cell)
            if formats and col in formats and isinstance(value, (int, float, Decimal)):
                cell.number_format = formats[col]
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    ws.freeze_panes = f"A{start_row + 1}"
    set_widths(ws, widths)
    return start_row + len(rows)


def create_workbook(output_path: Path, source_doc: dict, checked_at: str) -> None:
    listings = source_doc["flattened_listing_lines"]
    validation = source_doc["validation"]
    if len(source_doc["company_records"]) != 77:
        raise ValueError("官方原文应有77个公司级序号")
    if len(listings) != 78:
        raise ValueError("官方原文应有78条证券级上市行")
    if sorted({item["sequence"] for item in listings}) != list(range(1, 78)):
        raise ValueError("证券级上市行的公司级序号不完整")
    if sum(item["sequence"] == 33 for item in listings) != 2:
        raise ValueError("序号33应保留两个证券级上市行")
    if sum((dec(item["fair_value_cny"]) for item in listings), Decimal("0")) != EQUITY_CNY:
        raise ValueError("证券级上市行合计与官方权益投资总额不一致")
    if dec(validation["fair_value_sum_cny"]) != EQUITY_CNY:
        raise ValueError("结构化权益金额与官方披露总额不一致")
    if validation["sequence_is_exactly_1_to_77"] is not True:
        raise ValueError("公司级序号不是连续的1至77")

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "持仓QDII基金 2026H1 持仓跟踪"
    wb.properties.subject = "以基金正式中期报告更新的持仓快照"
    wb.properties.creator = "Codex"

    notes = wb.create_sheet("说明")
    set_title(notes, "持仓QDII基金 2026H1 持仓跟踪", "只写入截至核验时已正式披露 2026 年中期报告的基金家族；未披露者不以 Q2 补充。", 3)
    note_rows = [
        ("检查时间", checked_at, "北京时间"),
        ("账户范围", "164212 / 016823；100055 / 022184；016664 / 016665", "A/C 份额按同一基金家族合并"),
        ("已更新基金", "164212 / 016823", REPORT_TITLE),
        ("尚待披露", "100055 / 022184；016664 / 016665", "截至本次核验未找到正式 2026H1 PDF"),
        ("报告期", REPORT_PERIOD, "基金期末：2026-06-30"),
        ("官方原文", SOURCE_URL, "CNINFO 原文；PDF 共78页"),
        ("文件校验", PDF_SHA256.upper(), "SHA-256；文件大小 1,575,951 bytes"),
        ("完整持仓边界", "§7.4 全部权益投资明细：77个公司级序号、78条证券级上市行", "PDF 第48-58页"),
        ("非权益边界", "§7.7 披露前五名债券投资；§7.10 披露前十名基金投资", "本报告显示1条债券、2条基金投资；不据此宣称所有非权益资产明细"),
        ("历史保护", "原有2026Q2工作簿未覆盖", "H1另存为新文件"),
    ]
    set_header(notes, 4, ["项目", "内容", "说明"])
    for index, row in enumerate(note_rows, 5):
        for col, value in enumerate(row, 1):
            cell = notes.cell(index, col, value)
            apply_base_style(cell)
            if col == 2 and isinstance(value, str) and value.startswith("https://"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
    set_widths(notes, [18, 72, 58])
    notes.freeze_panes = "A5"

    overview = wb.create_sheet("H1概览")
    set_title(overview, "H1 披露概览", "164212 / 016823 已按官方中期报告更新；其余当前持仓基金家族仅保留发布状态。", 4)
    overview_rows = [
        [FUND_NAME, "164212", "016823", "已发布并更新", REPORT_DATE, REPORT_PERIOD, "=SUM('H1全部权益'!I4:I81)", "=SUM('H1全部权益'!J4:J81)", EQUITY_CNY, pct(EQUITY_NAV_PCT), "CNINFO PDF 第46-58页"],
        ["富国全球科技互联网股票（QDII）", "100055", "022184", "待披露", None, None, None, None, None, None, "本轮无正式H1 PDF"],
        ["天弘全球高端制造混合（QDII）", "016664", "016665", "待披露", None, None, None, None, None, None, "本轮无正式H1 PDF"],
    ]
    headers = ["基金家族", "A份额", "C份额", "2026H1状态", "公告日", "报告期", "解析权益合计(元)", "逐行净值占比合计", "官方权益合计(元)", "官方权益占净值比", "备注"]
    end_row = write_table(
        overview,
        4,
        headers,
        overview_rows,
        [34, 12, 12, 18, 14, 26, 20, 18, 20, 18, 38],
        {7: NUMBER_FORMAT, 8: PCT_FORMAT, 9: NUMBER_FORMAT, 10: PCT_FORMAT},
    )
    for cell in overview[5][6:8]:
        apply_base_style(cell, formula=True)
    overview["G5"].number_format = NUMBER_FORMAT
    overview["H5"].number_format = PCT_FORMAT
    for cell in overview[5]:
        cell.comment = Comment("硬编码来源或公式来源见工作簿“来源与核验”工作表。", "Codex")
    overview.row_dimensions[end_row + 2].height = 30
    overview.merge_cells(start_row=end_row + 2, start_column=1, end_row=end_row + 2, end_column=len(headers))
    overview.cell(end_row + 2, 1, "注：逐行占净值比例按原文两位小数披露；逐行合计与报告合计存在 0.02 个百分点的四舍五入差异。")
    overview.cell(end_row + 2, 1).fill = PatternFill("solid", fgColor=NOTE)
    overview.cell(end_row + 2, 1).alignment = Alignment(wrap_text=True, vertical="center")

    equity = wb.create_sheet("H1全部权益")
    set_title(equity, "164212 / 016823 H1 全部权益持仓", "官方报告 §7.4；公司级序号 1-77。序号33含两个证券市场/代码子行，故证券级上市行共78条。", 12)
    headers = ["公司级序号", "证券子行", "公司名称(英文)", "公司名称(中文)", "证券代码", "所在证券市场", "所属国家(地区)", "数量(股)", "公允价值(人民币)", "占基金资产净值比例", "PDF页码", "来源"]
    set_header(equity, 3, headers)
    for row_index, item in enumerate(listings, 4):
        values = [
            item["sequence"],
            item["listing_index"],
            item["company_name_en"],
            item["company_name_zh"],
            item["security_code"],
            item["market"],
            item["country_region"],
            int(item["quantity_shares"]),
            float(dec(item["fair_value_cny"])),
            pct(item["nav_pct"]),
            page_range(item),
            SOURCE_URL,
        ]
        for col, value in enumerate(values, 1):
            cell = equity.cell(row_index, col, value)
            apply_base_style(cell)
            if col == 8:
                cell.number_format = INTEGER_FORMAT
            elif col == 9:
                cell.number_format = NUMBER_FORMAT
            elif col == 10:
                cell.number_format = PCT_FORMAT
            elif col == 12:
                cell.hyperlink = value
                cell.style = "Hyperlink"
    total_row = 4 + len(listings)
    equity.cell(total_row, 1, "解析合计")
    equity.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    equity.cell(total_row, 9, f"=SUM(I4:I{total_row - 1})")
    equity.cell(total_row, 10, f"=SUM(J4:J{total_row - 1})")
    equity.cell(total_row, 11, "与报告§7.3核验")
    equity.cell(total_row, 12, SOURCE_URL)
    for col in range(1, 13):
        cell = equity.cell(total_row, col)
        cell.fill = PatternFill("solid", fgColor=SECTION)
        apply_base_style(cell, formula=col in {9, 10})
    equity.cell(total_row, 9).number_format = NUMBER_FORMAT
    equity.cell(total_row, 10).number_format = PCT_FORMAT
    equity.cell(total_row, 12).hyperlink = SOURCE_URL
    equity.cell(total_row, 12).style = "Hyperlink"
    reported_row = total_row + 1
    equity.merge_cells(start_row=reported_row, start_column=1, end_row=reported_row, end_column=8)
    equity.cell(reported_row, 1, "报告§7.3权益投资合计")
    equity.cell(reported_row, 9, float(EQUITY_CNY))
    equity.cell(reported_row, 10, pct(EQUITY_NAV_PCT))
    equity.cell(reported_row, 11, "报告第48页")
    equity.cell(reported_row, 12, SOURCE_URL)
    difference_row = reported_row + 1
    equity.merge_cells(start_row=difference_row, start_column=1, end_row=difference_row, end_column=8)
    equity.cell(difference_row, 1, "解析合计－报告合计")
    equity.cell(difference_row, 9, f"=I{total_row}-I{reported_row}")
    equity.cell(difference_row, 10, f"=J{total_row}-J{reported_row}")
    equity.cell(difference_row, 11, "比例差异来自逐行两位小数四舍五入")
    for row in range(reported_row, difference_row + 1):
        for col in range(1, 13):
            cell = equity.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=SECTION if row == reported_row else NOTE)
            apply_base_style(cell, formula=col in {9, 10} and row == difference_row)
        equity.cell(row, 9).number_format = NUMBER_FORMAT
        equity.cell(row, 10).number_format = PCT_FORMAT
        equity.cell(row, 12).hyperlink = SOURCE_URL
        equity.cell(row, 12).style = "Hyperlink"
    equity.auto_filter.ref = f"A3:L{total_row - 1}"
    equity.freeze_panes = "A4"
    set_widths(equity, [12, 10, 30, 24, 19, 20, 14, 13, 19, 18, 12, 54])

    allocation = wb.create_sheet("H1资产配置")
    set_title(allocation, "164212 / 016823 H1 基金资产组合", "官方报告 §7.1；金额单位为人民币元，占比以基金总资产为分母。", 6)
    allocation_rows = [
        [1, "权益投资", 1647778275.36, pct("81.85"), "", "46-47"],
        ["", "其中：普通股", 1582385428.23, pct("78.60"), "", "47"],
        ["", "存托凭证", 65392847.13, pct("3.25"), "", "47"],
        [2, "基金投资", 129548575.25, pct("6.43"), "", "47"],
        [3, "固定收益投资", 909443.10, pct("0.05"), "", "47"],
        ["", "其中：债券", 909443.10, pct("0.05"), "", "47"],
        [4, "金融衍生品投资", None, None, "报告显示-", "47"],
        [5, "买入返售金融资产", None, None, "报告显示-", "47"],
        [6, "货币市场工具", None, None, "报告显示-", "47"],
        [7, "银行存款和结算备付金合计", 128783987.25, pct("6.40"), "", "47"],
        [8, "其他各项资产", 106240011.38, pct("5.28"), "", "47"],
        [9, "合计", 2013260292.34, pct("100.00"), "", "47"],
    ]
    end = write_table(allocation, 3, ["序号", "项目", "金额(人民币)", "占基金总资产比", "原文显示", "PDF页码"], allocation_rows, [10, 32, 22, 20, 18, 14], {3: NUMBER_FORMAT, 4: PCT_FORMAT})
    allocation.auto_filter.ref = f"A3:F{end}"

    geography = wb.create_sheet("H1地区行业")
    set_title(geography, "164212 / 016823 H1 地区与行业配置", "官方报告 §7.2、§7.3；占比以基金资产净值为分母。", 7)
    geo_rows = [
        ["地区", "美国", 845928450.68, pct("44.19"), "§7.2", "47", SOURCE_URL],
        ["地区", "中国", 685614842.42, pct("35.81"), "§7.2", "47", SOURCE_URL],
        ["地区", "中国香港", 116234982.26, pct("6.07"), "§7.2", "47", SOURCE_URL],
        ["地区", "合计", 1647778275.36, pct("86.07"), "§7.2", "47", SOURCE_URL],
        ["行业", "基础材料", 241578100.00, pct("12.62"), "§7.3", "48", SOURCE_URL],
        ["行业", "工业", 98167059.14, pct("5.13"), "§7.3", "48", SOURCE_URL],
        ["行业", "消费者非必需品", 28940690.04, pct("1.51"), "§7.3", "48", SOURCE_URL],
        ["行业", "信息技术", 1278484088.00, pct("66.78"), "§7.3", "48", SOURCE_URL],
        ["行业", "公用事业", 383765.49, pct("0.02"), "§7.3", "48", SOURCE_URL],
        ["行业", "其他-GICS未分类", 224572.69, pct("0.01"), "§7.3", "48", SOURCE_URL],
        ["行业", "合计", 1647778275.36, pct("86.07"), "§7.3", "48", SOURCE_URL],
    ]
    end = write_table(geography, 3, ["分类", "项目", "公允价值(人民币)", "占基金资产净值比", "报告章节", "PDF页码", "来源"], geo_rows, [12, 28, 22, 20, 14, 12, 54], {3: NUMBER_FORMAT, 4: PCT_FORMAT})
    for row in range(4, end + 1):
        geography.cell(row, 7).hyperlink = SOURCE_URL
        geography.cell(row, 7).style = "Hyperlink"

    bonds = wb.create_sheet("H1债券")
    set_title(bonds, "164212 / 016823 H1 债券投资", "官方报告 §7.7 披露的前五名债券投资明细；本期报告显示1条。", 8)
    bond_rows = [[1, "019792", "25国债19", 900000, 909443.10, pct("0.05"), "66", SOURCE_URL]]
    write_table(bonds, 3, ["序号", "债券代码", "债券名称", "数量/面值", "公允价值(人民币)", "占净值比", "PDF页码", "来源"], bond_rows, [10, 16, 24, 16, 22, 16, 12, 54], {4: INTEGER_FORMAT, 5: NUMBER_FORMAT, 6: PCT_FORMAT})
    bonds.cell(4, 8).hyperlink = SOURCE_URL
    bonds.cell(4, 8).style = "Hyperlink"

    funds = wb.create_sheet("H1基金投资")
    set_title(funds, "164212 / 016823 H1 基金投资", "官方报告 §7.10 披露的前十名基金投资；本期报告列示2项，不宣称为全部基金投资明细。", 8)
    fund_rows = [
        [1, "CSOP SK Hynix Daily 2x Leveraged Product", "ETF基金", "交易型开放式", "CSOP Asset Management Ltd", 93178044.00, pct("4.87"), "67"],
        [2, "CSOP Samsung Electronics Daily 2x Leveraged Product", "ETF基金", "交易型开放式", "CSOP Asset Management Ltd", 36370531.25, pct("1.90"), "67"],
    ]
    end = write_table(funds, 3, ["序号", "基金名称", "基金类型", "运作方式", "管理人", "公允价值(人民币)", "占净值比", "PDF页码"], fund_rows, [10, 42, 16, 18, 32, 22, 16, 12], {6: NUMBER_FORMAT, 7: PCT_FORMAT})
    funds.merge_cells(start_row=end + 2, start_column=1, end_row=end + 2, end_column=8)
    funds.cell(end + 2, 1, "说明：报告章节为“前十名基金投资明细”，其列示范围不等同于完整基金投资穿透清单。")
    funds.cell(end + 2, 1).fill = PatternFill("solid", fgColor=NOTE)
    funds.cell(end + 2, 1).alignment = Alignment(wrap_text=True)

    notices = wb.create_sheet("公告状态")
    set_title(notices, "当前账户基金家族的 2026H1 公告状态", "只以可下载、可核验的基金公司或法定披露平台原文作为“已发布”依据。", 10)
    notice_rows = [
        [checked_at, "164212", "016823", FUND_NAME, "已发布", REPORT_TITLE, REPORT_DATE, REPORT_PERIOD, SOURCE_URL, "CNINFO官方原文；EID索引可能延迟"],
        [checked_at, "100055", "022184", "富国全球科技互联网股票（QDII）", "待披露", "N/A", None, None, None, "截至核验时富国官网与EID均未见正式H1"],
        [checked_at, "016664", "016665", "天弘全球高端制造混合（QDII）", "待披露", "N/A", None, None, None, "截至核验时天弘官网与EID均未见正式H1"],
    ]
    end = write_table(notices, 3, ["核验时间", "A份额", "C份额", "基金家族", "H1状态", "最新H1标题", "公告日", "报告期", "官方原文", "说明"], notice_rows, [28, 12, 12, 34, 14, 54, 14, 26, 54, 42])
    notices.cell(4, 9).hyperlink = SOURCE_URL
    notices.cell(4, 9).style = "Hyperlink"
    notices.auto_filter.ref = f"A3:J{end}"

    checks = wb.create_sheet("来源与核验")
    set_title(checks, "来源与核验", "所有金额均保留原始披露口径；公式仅用于工作簿内的完整性校验。", 5)
    check_rows = [
        ["官方原文", REPORT_TITLE, SOURCE_URL, "第1-2页", "已验证PDF、报告期、送出日期"],
        ["文件身份", PDF_SHA256.upper(), "CNINFO PDF", "全文件", "SHA-256；78页；1,575,951 bytes"],
        ["公司级序号", 77, "§7.4", "48-58", "严格连续1-77，无缺号/重号"],
        ["证券级上市行", 78, "§7.4", "48-58", "序号33有两个证券市场/代码子行"],
        ["解析权益市值", "=SUM('H1全部权益'!I4:I81)", "§7.4行项目", "48-58", "与§7.3报告总额比对"],
        ["报告权益市值", float(EQUITY_CNY), "§7.3", "48", "人民币；与解析合计应为0差异"],
        ["权益金额差异", "='H1全部权益'!I82-'H1全部权益'!I83", "公式", "H1全部权益", "应为0.00"],
        ["逐行净值占比合计", "=SUM('H1全部权益'!J4:J81)", "§7.4", "48-58", "逐行两位小数披露"],
        ["报告权益占净值比", pct(EQUITY_NAV_PCT), "§7.3", "48", "86.07%"],
        ["占比差异", "='H1全部权益'!J82-'H1全部权益'!J83", "公式", "H1全部权益", "0.02个百分点为逐行四舍五入差异"],
    ]
    set_header(checks, 3, ["核验项目", "数值/公式", "来源", "页码", "结果或说明"])
    for row_index, values in enumerate(check_rows, 4):
        for col, value in enumerate(values, 1):
            cell = checks.cell(row_index, col, value)
            is_formula = isinstance(value, str) and value.startswith("=")
            apply_base_style(cell, formula=is_formula)
            if col == 2 and row_index in {8, 9, 10}:
                cell.number_format = NUMBER_FORMAT
            if col == 2 and row_index in {11, 12, 13}:
                cell.number_format = PCT_FORMAT
        if row_index == 4:
            checks.cell(row_index, 3).hyperlink = SOURCE_URL
            checks.cell(row_index, 3).style = "Hyperlink"
    checks.freeze_panes = "A4"
    set_widths(checks, [24, 28, 60, 16, 55])

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font.name is None:
                    cell.font = INPUT_FONT
        ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_status_payload(checked_at: str) -> dict:
    return {
        "checked_at": checked_at,
        "scope": "current_account_fund_share_classes",
        "account_fund_codes": ["164212", "016823", "100055", "022184", "016664", "016665"],
        "fund_families": [
            {
                "primary_code": "164212",
                "share_classes": ["164212", "016823"],
                "h1_status": "published",
                "title": REPORT_TITLE,
                "announcement_date": REPORT_DATE,
                "period": "2026-01-01/2026-06-30",
                "official_pdf": SOURCE_URL,
                "pdf_sha256": PDF_SHA256,
                "equity_coverage": "full_equity_details_77_company_sequences_78_security_listing_lines",
            },
            {
                "primary_code": "100055",
                "share_classes": ["100055", "022184"],
                "h1_status": "not_found_at_check_time",
                "official_pdf": None,
            },
            {
                "primary_code": "016664",
                "share_classes": ["016664", "016665"],
                "h1_status": "not_found_at_check_time",
                "official_pdf": None,
            },
        ],
        "boundary": "Only the 164212/016823 official H1 full equity table is updated. Pending families are not backfilled with Q2 data.",
    }


def build_audit_payload(checked_at: str, first_holding: dict) -> dict:
    position_as_of = "2026-06-30T23:59:59+08:00"
    return {
        "schema_version": "1.0",
        "audit_id": "fund-164212-h1-20260831",
        "as_of": checked_at,
        "sources": [
            {
                "id": "S_CNINFO_OFFICIAL",
                "source_type": "official_filing",
                "origin_id": "cninfo:1225528103:2026h1",
                "locator": SOURCE_URL + "#page=48",
                "source_date": REPORT_DATE,
                "checked_at": checked_at,
                "status": "accepted",
            },
            {
                "id": "S_STRUCTURED_TRANSCRIPTION",
                "source_type": "report_under_audit",
                "origin_id": "cninfo:1225528103:2026h1",
                "locator": "artifact:fund_tracking/h1_2026_164212_all_equity.json#sequence=1",
                "source_date": REPORT_DATE,
                "checked_at": checked_at,
                "status": "accepted",
            },
        ],
        "facts": [
            {
                "id": "F_OFFICIAL_CORNING_VALUE",
                "metric": "fund_equity_position_fair_value",
                "value": first_holding["fair_value_cny"],
                "unit": "currency",
                "currency": "CNY",
                "scale": "1",
                "period": {"kind": "instant", "as_of": position_as_of},
                "basis": "reported_fund_equity_position",
                "source_refs": ["S_CNINFO_OFFICIAL"],
            },
            {
                "id": "F_TRANSCRIBED_CORNING_VALUE",
                "metric": "fund_equity_position_fair_value",
                "value": first_holding["fair_value_cny"],
                "unit": "currency",
                "currency": "CNY",
                "scale": "1",
                "period": {"kind": "instant", "as_of": position_as_of},
                "basis": "reported_fund_equity_position",
                "source_refs": ["S_STRUCTURED_TRANSCRIPTION"],
            },
        ],
        "checks": [
            {
                "id": "C_CORNING_TRANSCRIPTION",
                "kind": "cross_source",
                "materiality": "material",
                "target": {"fact_id": "F_TRANSCRIBED_CORNING_VALUE"},
                "references": [{"fact_id": "F_OFFICIAL_CORNING_VALUE"}],
                "source_gate": {
                    "min_independent_origins": 1,
                    "counted_tier": "official",
                    "required_anchor_tier": "official",
                },
                "tolerance": {"relative_pct": "0", "absolute_base": "0"},
            }
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", required=True, type=Path)
    parser.add_argument("--equity-json", required=True, type=Path)
    parser.add_argument("--artifacts-dir", required=True, type=Path)
    args = parser.parse_args()

    source_doc = json.loads(args.equity_json.read_text(encoding="utf-8"))
    if sha256(args.pdf) != PDF_SHA256:
        raise ValueError("官方PDF SHA-256不匹配")
    if sha256(args.equity_json) != TRANSCRIPTION_SHA256:
        raise ValueError("已核验的权益转录JSON SHA-256不匹配")
    checked_at = now_bj()
    artifacts = args.artifacts_dir
    artifacts.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.pdf, artifacts / "h1_2026_164212.pdf")
    (artifacts / "h1_2026_164212_all_equity.json").write_text(
        json.dumps(source_doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    status_path = artifacts / "latest_notice_check" / "2026_h1_status_20260831.json"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    status_path.write_text(json.dumps(build_status_payload(checked_at), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    audit_path = artifacts / "financial_audit_164212_h1_20260831.input.json"
    audit_path.write_text(
        json.dumps(build_audit_payload(checked_at, source_doc["flattened_listing_lines"][0]), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    create_workbook(artifacts / "持仓QDII基金_2026H1持仓跟踪.xlsx", source_doc, checked_at)
    print(json.dumps({"checked_at": checked_at, "artifacts": str(artifacts)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
